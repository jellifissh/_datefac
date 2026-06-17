"""Workbook intake helpers for the 348A Excel audit pilot."""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from datefac_agent.audit.row_type_classifier import classify_row_type
from datefac_agent.schemas.audit_models import SpreadsheetRow, WorkbookIntakeResult

PERIOD_LABEL_RE = re.compile(r"(?:19|20)\d{2}(?:\s*(?:A|E|Q[1-4]|FY))?", re.IGNORECASE)
EVIDENCE_HEADER_HINTS = ("页", "page", "evidence", "source", "出处", "来源")
HEADER_LABEL_HINTS = {"项目", "指标", "会计年度", "公司", "业务板块", "类别", "内容详情", "数据类别", "备注"}
SYNTHETIC_KEY_VALUE_HEADERS = ["field_name", "field_value"]
THIRD_WORKBOOK_REPORT_INFO_SHEET = "\u62a5\u544a\u6838\u5fc3\u4fe1\u606f\u4e0e\u6295\u8d44\u8981\u70b9"
THIRD_WORKBOOK_BUSINESS_MATRIX_SHEET = "\u516c\u53f8\u4e1a\u52a1\u4e0e\u4ea7\u54c1\u77e9\u9635"
THIRD_WORKBOOK_NA_AIDC_SHEET = "\u5317\u7f8eAIDC\u7535\u529b\u4f9b\u9700\u4e0e\u6280\u672f\u8def\u5f84"

THIRD_WORKBOOK_REPORT_METADATA_HINTS = {
    "\u62a5\u544a\u7c7b\u578b",
    "\u6807\u7684\u516c\u53f8",
    "\u6240\u5c5e\u884c\u4e1a",
    "\u53d1\u5e03\u673a\u6784",
    "\u53d1\u5e03\u65e5\u671f",
    "\u6295\u8d44\u8bc4\u7ea7",
}
THIRD_WORKBOOK_REPORT_NARRATIVE_PREFIXES = ("1.", "2.", "3.", "4.", "\u4e8c\u3001", "\u4e09\u3001")
THIRD_WORKBOOK_NA_SECTION_HINTS = {
    "\u88681\uff1a2026\u5e74\u7f8e\u56fd\u7164\u7535\u91cd\u70b9\u9000\u5f79\u8ba1\u5212",
    "\u88682\uff1a2025-2030\u5e74\u7f8e\u56fd\u5206\u7535\u6e90\u7c7b\u578b\u589e\u51cf\uff08\u5355\u4f4d\uff1aGW\uff09",
    "\u88683\uff1aAIDC\u53d1\u7535\u6280\u672f\u8def\u5f84\u5bf9\u6bd4",
    "\u7535\u6e90\u7c7b\u578b",
    "\u6280\u672f\u8def\u7ebf",
}


def _stringify_cell(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _normalize_header(value: Any, column_index: int) -> str:
    text = _stringify_cell(value)
    return text or f"column_{column_index}"


def _extract_unit_hint(metric_name: str) -> str | None:
    match = re.search(r"[（(]([^()（）]+)[)）]", metric_name)
    if match:
        return match.group(1).strip() or None
    return None


def _extract_metric_name(values: list[Any]) -> str:
    for value in values[:2]:
        text = _stringify_cell(value)
        if text:
            return text
    for value in values:
        text = _stringify_cell(value)
        if text:
            return text
    return "UNLABELED_ROW"


def _extract_explicit_evidence_ref(headers: list[str], values: list[Any]) -> str | None:
    for header, value in zip(headers, values):
        lower_header = header.lower()
        if any(hint in lower_header for hint in EVIDENCE_HEADER_HINTS):
            text = _stringify_cell(value)
            if text:
                return text
    return None


def detect_period_labels(column_names: list[str]) -> list[str]:
    """Return period labels found in workbook header cells."""

    labels: list[str] = []
    for name in column_names:
        text = _stringify_cell(name)
        if PERIOD_LABEL_RE.search(text):
            labels.append(text)
    return labels


def _count_non_empty(values: list[Any]) -> int:
    return sum(1 for value in values if _stringify_cell(value))


def _is_header_candidate(values: list[Any]) -> bool:
    texts = [_stringify_cell(value) for value in values]
    non_empty = [text for text in texts if text]
    if len(non_empty) < 2:
        return False
    period_hits = sum(1 for text in non_empty if PERIOD_LABEL_RE.search(text))
    if non_empty[0] in HEADER_LABEL_HINTS:
        return True
    if period_hits >= 2:
        return True
    if period_hits >= 1 and len(non_empty) >= 3:
        return True
    return False


def _find_header_row(sheet_rows: list[tuple[int, list[Any]]]) -> tuple[int, list[str]] | None:
    for row_index, values in sheet_rows[:8]:
        if _is_header_candidate(values):
            header_names = [
                _normalize_header(value, column_index)
                for column_index, value in enumerate(values, start=1)
            ]
            return row_index, header_names
    return None


def _find_key_value_start(sheet_rows: list[tuple[int, list[Any]]]) -> int | None:
    candidate_rows: list[int] = []
    for row_index, values in sheet_rows[:12]:
        texts = [_stringify_cell(value) for value in values[:2]]
        if len(texts) < 2:
            continue
        if texts[0] and texts[1]:
            candidate_rows.append(row_index)
    if len(candidate_rows) >= 3:
        return candidate_rows[0]
    return None


def _should_reset_read_only_dimensions(worksheet: Any) -> bool:
    """Detect read-only worksheets whose cached dimension is obviously undersized."""

    if not hasattr(worksheet, "reset_dimensions") or not hasattr(worksheet, "calculate_dimension"):
        return False
    try:
        dimension = worksheet.calculate_dimension()
    except Exception:
        return False
    return dimension == "A1:A1"


def _refine_third_workbook_row_type(row: SpreadsheetRow, inferred_row_type: str) -> str:
    """Apply workbook-specific row typing only when the generic classifier is unknown."""

    if inferred_row_type != "UNKNOWN_ROW":
        return inferred_row_type

    metric_name = _stringify_cell(row.metric_name)

    if row.sheet_name == THIRD_WORKBOOK_REPORT_INFO_SHEET:
        if metric_name in THIRD_WORKBOOK_REPORT_METADATA_HINTS:
            return "NARRATIVE_ASSERTION"
        if metric_name.startswith(THIRD_WORKBOOK_REPORT_NARRATIVE_PREFIXES):
            return "NARRATIVE_ASSERTION"
        return "NARRATIVE_ASSERTION"

    if row.sheet_name == THIRD_WORKBOOK_BUSINESS_MATRIX_SHEET:
        return "NARRATIVE_ASSERTION"

    if row.sheet_name == THIRD_WORKBOOK_NA_AIDC_SHEET:
        if metric_name in THIRD_WORKBOOK_NA_SECTION_HINTS or metric_name.startswith("\u8868"):
            return "NARRATIVE_ASSERTION"
        return "NARRATIVE_ASSERTION"

    return inferred_row_type


def read_excel_workbook(excel_path: str | Path) -> WorkbookIntakeResult:
    """Read an extracted workbook into lightweight structured rows."""

    path = Path(excel_path)
    workbook = load_workbook(path, read_only=True, data_only=True)
    rows: list[SpreadsheetRow] = []

    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        if _should_reset_read_only_dimensions(worksheet):
            worksheet.reset_dimensions()
        sheet_rows = [
            (row_index, list(raw_row))
            for row_index, raw_row in enumerate(worksheet.iter_rows(values_only=True), start=1)
        ]
        header_candidate = _find_header_row(sheet_rows)
        key_value_start = _find_key_value_start(sheet_rows) if header_candidate is None else None

        if header_candidate is not None:
            header_row_index, header_names = header_candidate
            data_rows = [(row_index, values) for row_index, values in sheet_rows if row_index > header_row_index]
        elif key_value_start is not None:
            header_names = SYNTHETIC_KEY_VALUE_HEADERS
            data_rows = [(row_index, values[:2]) for row_index, values in sheet_rows if row_index >= key_value_start]
        else:
            if not sheet_rows:
                continue
            header_row_index, header_values = sheet_rows[0]
            header_names = [
                _normalize_header(value, column_index)
                for column_index, value in enumerate(header_values, start=1)
            ]
            data_rows = [(row_index, values) for row_index, values in sheet_rows if row_index > header_row_index]

        for row_index, values in data_rows:
            if not any(value is not None and str(value).strip() for value in values):
                continue

            padded_values = values + [None] * (len(header_names) - len(values))
            raw_values = {
                header_names[column_index]: padded_values[column_index]
                for column_index in range(len(header_names))
            }
            metric_name = _extract_metric_name(values)
            period_values = {
                header: raw_values[header]
                for header in detect_period_labels(header_names)
                if raw_values.get(header) is not None and str(raw_values.get(header)).strip() != ""
            }

            row = SpreadsheetRow(
                source_excel_path=str(path),
                sheet_name=sheet_name,
                row_index=row_index,
                column_names=header_names,
                raw_values=raw_values,
                metric_name=metric_name,
                unit_hint=_extract_unit_hint(metric_name),
                period_values=period_values,
                explicit_evidence_ref=_extract_explicit_evidence_ref(header_names, padded_values),
            )
            row.row_type = _refine_third_workbook_row_type(row, classify_row_type(row))
            rows.append(row)

    return WorkbookIntakeResult(
        source_excel_path=str(path),
        sheet_names=list(workbook.sheetnames),
        rows=rows,
        sheet_count=len(workbook.sheetnames),
        row_count_total=len(rows),
    )
