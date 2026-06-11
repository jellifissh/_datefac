from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_SNAPSHOT_SUMMARY",
    "02_MILESTONE_CHAIN",
    "03_KEY_ARTIFACTS",
    "04_DEMO_GUIDE",
    "05_PACKAGE_OVERVIEW",
    "06_TRUST_LEVELS",
    "07_RISK_BOUNDARY",
    "08_COLLISION_SUMMARY",
    "09_BACKLOG_SUMMARY",
    "10_HANDOFF_CHECKLIST",
    "11_NEXT_STEP_OPTIONS",
    "12_343A_READINESS",
    "13_NO_WRITE_BACK",
    "14_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "warning",
        "path",
        "purpose",
        "caveat",
        "meaning",
        "note",
        "value",
        "focus",
        "goal",
        "trigger",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 180)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 220)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 342S Package Audit Snapshot / Demo Handoff",
        "",
        "## 中文摘要",
        "- 342S 是基于 342R 的 package audit snapshot / demo handoff，不是 formal client export。",
        "- 当前 package 共有 130 条 candidate rows，其中 30 条为 HUMAN_REVIEWED，100 条为 SIMULATED。",
        "- 100 条 simulated rows 仍然需要 disclaimer 与 later audit，所以 `formal_client_export_allowed=false`、`client_ready=false`、`production_ready=false` 必须保持不变。",
        "- 当前结果适合内部演示、审计复盘、阶段性交接，不可作为正式客户交付或投资建议。",
        "",
        "## English Summary",
        "- 342S is a package audit snapshot and demo handoff built on top of 342R.",
        "- It keeps the MinerU-first / table-first mainline visible while preserving risk boundaries.",
        "- It is not a formal client export, not final human-review completion, and not production-ready.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- demo_handoff_ready: {summary.get('demo_handoff_ready', False)}",
        f"- ready_for_343a: {summary.get('ready_for_343a', False)}",
        f"- recommended_343a_scope: {summary.get('recommended_343a_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Current Mainline",
        f"- latest_completed_milestone: {summary.get('latest_completed_milestone', '')}",
        f"- current_milestone: {summary.get('current_milestone', '')}",
        f"- current_mainline: {summary.get('current_mainline', '')}",
        f"- old_342e_text_candidate_route: {summary.get('old_342e_text_candidate_route', '')}",
        "",
        "## Package Counts",
        f"- export_candidate_package_row_count: {summary.get('export_candidate_package_row_count', 0)}",
        f"- human_reviewed_candidate_count: {summary.get('human_reviewed_candidate_count', 0)}",
        f"- simulated_candidate_count: {summary.get('simulated_candidate_count', 0)}",
        f"- simulated_direct_candidate_count: {summary.get('simulated_direct_candidate_count', 0)}",
        f"- simulated_corrected_candidate_count: {summary.get('simulated_corrected_candidate_count', 0)}",
        f"- disclaimer_required_count: {summary.get('disclaimer_required_count', 0)}",
        f"- later_audit_required_count: {summary.get('later_audit_required_count', 0)}",
        "",
        "## Risk And Backlog",
        f"- export_risk_level: {summary.get('export_risk_level', '')}",
        f"- collision_logged_count: {summary.get('collision_logged_count', 0)}",
        f"- duplicate_metric_year_source_count: {summary.get('duplicate_metric_year_source_count', 0)}",
        f"- severe_collision_count: {summary.get('severe_collision_count', 0)}",
        f"- unresolved_collision_count: {summary.get('unresolved_collision_count', 0)}",
        f"- still_human_required_count: {summary.get('still_human_required_count', 0)}",
        f"- remaining_review_count: {summary.get('remaining_review_count', 0)}",
        "",
        "## Safety Boundary",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        f"- recommended_open_excel_path: {summary.get('recommended_open_excel_path', '')}",
        f"- recommended_demo_readme_path: {summary.get('recommended_demo_readme_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Recommended Demo Open Order",
            "- First open the 342R workbook.",
            "- Start from `03_EXPORT_CANDIDATES`, then `07_AUDIT_LABELS`, `08_REQUIRED_WARNINGS`, `09_RISK_DISCLOSURE`, `10_COLLISION_CONTEXT`, `11_BACKLOG_CONTEXT`, and `12_342S_READINESS`.",
            "",
            "## Next Recommendation",
            "- Recommended next task: `343A Review Queue Schema And Human Review UI Pilot`.",
            "- Keep Argilla / real LLM response ingestion / Phoenix trace as later options, not current mandatory work.",
            "",
        ]
    )
    return "\n".join(lines)


def demo_readme_markdown(summary: Dict[str, Any], recommended_sheets: Sequence[str]) -> str:
    lines = [
        "# 342S Demo Readme",
        "",
        "一句话定位：",
        "- 342S 是一个基于 342R 的内部 snapshot / demo handoff 包，用来展示当前主线成果和边界，不是正式客户导出。",
        "",
        "当前最该打开的 Excel：",
        f"- {summary.get('recommended_open_excel_path', '')}",
        "",
        "推荐演示顺序：",
        "- Step 1: 先说明当前主线是 MinerU-first / table-first。",
        "- Step 2: 说明真实链路是 table extraction -> human review -> simulated adoption -> audit gate -> candidate package。",
        f"- Step 3: 打开 342R workbook，重点看 {' | '.join(recommended_sheets)}。",
        "- Step 4: 用 data_trust_level 区分 HUMAN_REVIEWED、SIMULATED_DIRECT_ADOPTED、SIMULATED_CORRECTION_ADOPTED。",
        "- Step 5: 打开 warnings / risk / collision / backlog 相关 sheet，解释为什么 export_risk_level 仍然是 HIGH。",
        "- Step 6: 最后明确这是可演示、可审计、可交接，但不是正式 client delivery。",
        "",
        "关键指标：",
        f"- export_candidate_package_row_count = {summary.get('export_candidate_package_row_count', 0)}",
        f"- human_reviewed_candidate_count = {summary.get('human_reviewed_candidate_count', 0)}",
        f"- simulated_candidate_count = {summary.get('simulated_candidate_count', 0)}",
        f"- still_human_required_count = {summary.get('still_human_required_count', 0)}",
        f"- remaining_review_count = {summary.get('remaining_review_count', 0)}",
        "",
        "风险边界：",
        "- formal_client_export_allowed = false",
        "- client_ready = false",
        "- production_ready = false",
        "- not investment advice",
        "- simulated rows require disclaimer and later audit",
        "",
        "不能说的话：",
        "- 不要说 formal client export 已允许。",
        "- 不要说 client_ready=true 或 production_ready=true。",
        "- 不要说 full human review completed。",
        "- 不要说 real LLM review completed。",
        "- 不要说这是投资建议。",
        "",
        "后续路线：",
        "- 推荐下一步是 343A Review Queue Schema And Human Review UI Pilot。",
        "- 343B / 343C / 343D 是后续候选路线，不是当前 snapshot 的一部分。",
        "",
    ]
    return "\n".join(lines)


def handoff_checklist_markdown(
    summary: Dict[str, Any],
    latest_commit_sha: str,
    forbidden_paths: Sequence[str],
    protected_paths: Sequence[str],
    recommended_sheets: Sequence[str],
) -> str:
    lines = [
        "# 342S Handoff Checklist",
        "",
        f"- latest commit before 342S commit: {latest_commit_sha or 'unknown'}",
        f"- latest completed milestone: {summary.get('latest_completed_milestone', '')}",
        f"- current milestone: {summary.get('current_milestone', '')}",
        f"- current mainline: {summary.get('current_mainline', '')}",
        f"- recommended artifact to open: {summary.get('recommended_open_excel_path', '')}",
        f"- recommended sheets: {' | '.join(recommended_sheets)}",
        f"- demo readme: {summary.get('recommended_demo_readme_path', '')}",
        f"- known backlog: remaining_review_count = {summary.get('remaining_review_count', 0)}, still_human_required_count = {summary.get('still_human_required_count', 0)}",
        f"- known risk level: {summary.get('export_risk_level', '')}",
        "- forbidden claims: formal client export allowed, client-ready, production-ready, final client delivery, investment advice",
        "- next recommended task: 343A Review Queue Schema And Human Review UI Pilot",
        "",
        "Forbidden paths:",
    ]
    for path in forbidden_paths:
        lines.append(f"- {path}")
    lines.extend(["", "Protected dirty paths:"])
    for path in protected_paths:
        lines.append(f"- {path}")
    lines.append("")
    return "\n".join(lines)


def next_step_plan_markdown(summary: Dict[str, Any]) -> str:
    lines = [
        "# 342S Next Step Plan",
        "",
        "## Recommended now",
        "- 343A Review Queue Schema And Human Review UI Pilot",
        "  Trigger: high backlog and still-human-required rows need explicit routing and UI assumptions.",
        "  Value: creates a durable queue contract before picking or scaling review tooling.",
        "  Risk: schema mistakes can leak into downstream UI/LLM work if rushed.",
        "",
        "## Optional follow-up pilots",
        "- 343B Argilla Human Review UI Pilot",
        "  Trigger: after 343A defines the queue schema.",
        "  Value: fast human-review UX validation on 50-100 high-risk rows.",
        "  Risk: should remain pluggable, not the core system.",
        "",
        "- 343C Real LLM/VLM Response Ingestion Pilot",
        "  Trigger: when real response traces, parser retention, and human correction loop are ready.",
        "  Value: moves from dry-run helpers to traceable real-response experiments.",
        "  Risk: must preserve response/trace/human correction evidence.",
        "",
        "- 343D Phoenix Observability Trace Pilot",
        "  Trigger: after or alongside real model calls.",
        "  Value: improves prompt/trace/eval visibility.",
        "  Risk: lower value before real model calls exist.",
        "",
        "- 343E Manual Review Expansion / Spot-Check Expansion",
        "  Trigger: when export_risk_level needs to be reduced before broader package use.",
        "  Value: directly reduces high-risk simulated and collision-heavy rows.",
        "  Risk: manual effort remains high without better queue tooling.",
        "",
        "## Not recommended now",
        "- Do not jump straight to lakeFS as the immediate next blocker.",
        "- Do not do a large LangGraph rewrite right now.",
        "- Do not use Dify as the core pipeline at this stage.",
        "",
        "## Current boundary reminder",
        f"- formal_client_export_allowed = {summary.get('formal_client_export_allowed', False)}",
        f"- client_ready = {summary.get('client_ready', False)}",
        f"- production_ready = {summary.get('production_ready', False)}",
        "",
    ]
    return "\n".join(lines)
