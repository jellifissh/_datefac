from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_GATE_SUMMARY",
    "02_INPUT_342L_SUMMARY",
    "03_SPOT_CHECK_TEMPLATE",
    "04_SPOT_CHECK_APPLY",
    "05_LLM_RESPONSE_SCHEMA",
    "06_LLM_RESPONSE_INGEST",
    "07_RULE_LLM_COMPARISON",
    "08_ADOPTION_POLICY",
    "09_ADOPTION_CANDIDATES",
    "10_BLOCKED_CANDIDATES",
    "11_RISK_GATE",
    "12_REDUCTION_AFTER_GATE",
    "13_342N_READINESS",
    "14_NO_WRITE_BACK",
    "15_NEXT_STEPS",
]

TEMPLATE_WORKBOOK_SHEETS = [
    "00_README",
    "01_SPOT_CHECK_REVIEW",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: Iterable[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(_to_jsonable(dict(row)), ensure_ascii=False) + "\n")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "note",
        "warning",
        "reason",
        "risk",
        "path",
        "html",
        "snippet",
        "bbox",
        "schema",
        "recommendation",
        "value",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 180)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 342M LLM Suggestion Spot-Check Gate",
        "",
        "## Chinese Summary",
        "- 342M is an adoption gate, not a true auto-apply stage.",
        "- If there is no reviewed spot-check workbook and no real LLM response JSONL, 342M must stay in `WAITING_FOR_EVIDENCE`.",
        "- Dry-run suggestions are not real LLM outputs.",
        "- Auto-confirm candidates are not final confirmations.",
        "",
        "## English Summary",
        "- 342M only prepares controlled evidence intake for a later stage.",
        "- It can generate templates and schemas without claiming real adoption progress.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- ready_for_342n: {summary.get('ready_for_342n', False)}",
        f"- recommended_342n_scope: {summary.get('recommended_342n_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Key Counts",
        f"- pending_review_count: {summary.get('pending_review_count', 0)}",
        f"- auto_confirm_candidate_count: {summary.get('auto_confirm_candidate_count', 0)}",
        f"- spot_check_sample_count: {summary.get('spot_check_sample_count', 0)}",
        f"- reviewed_spot_check_count: {summary.get('reviewed_spot_check_count', 0)}",
        f"- response_count: {summary.get('response_count', 0)}",
        f"- valid_llm_response_count: {summary.get('valid_llm_response_count', 0)}",
        f"- adoption_candidate_count: {summary.get('adoption_candidate_count', 0)}",
        f"- blocked_candidate_count: {summary.get('blocked_candidate_count', 0)}",
        "",
        "## Reduction After Gate",
        f"- risk_adjusted_reduction_count: {summary.get('risk_adjusted_reduction_count', 0)}",
        f"- required_human_review_after_gate: {summary.get('required_human_review_after_gate', 0)}",
        f"- conservative_reduction_rate_after_gate: {summary.get('conservative_reduction_rate_after_gate', 0)}",
        "",
        "## Safety",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        f"- waiting_for_human_spot_check: {summary.get('waiting_for_human_spot_check', False)}",
        f"- waiting_for_real_llm_responses: {summary.get('waiting_for_real_llm_responses', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Boundaries",
            "- 342M does not rerun MinerU.",
            "- 342M does not call a real LLM API by default.",
            "- 342M does not modify production pipeline / parser / extraction / delivery.",
            "- 342M does not write back to 342J / 342K / 342L workbooks.",
            "- client_ready remains false and production_ready remains false.",
            "",
        ]
    )
    return "\n".join(lines)
