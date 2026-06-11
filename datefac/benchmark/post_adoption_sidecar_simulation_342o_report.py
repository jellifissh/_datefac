from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_SIDECAR_SUMMARY",
    "02_INPUT_342N_SUMMARY",
    "03_SIM_ADOPTED_CELLS",
    "04_DIRECT_ADOPTED",
    "05_CORRECTED_ADOPTED",
    "06_STILL_HUMAN_REQUIRED",
    "07_BEFORE_AFTER_TRACE",
    "08_METRIC_COVERAGE",
    "09_REMAINING_REVIEW",
    "10_RISK_BOUNDARY",
    "11_342P_READINESS",
    "12_NO_WRITE_BACK",
    "13_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "note",
        "warning",
        "reason",
        "risk",
        "path",
        "html",
        "snippet",
        "bbox",
        "recommendation",
        "value",
        "evidence",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 180)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 342O Post-Adoption Sidecar Simulation",
        "",
        "## Chinese Summary",
        "- 342O is a post-adoption sidecar simulation, not final adoption and not final human confirmation.",
        "- 342O merges the 342N direct adoption rows and correction-aware adoption rows into a bounded sidecar result.",
        "- This stage remains no-write-back, not client-ready, and not production-ready.",
        "",
        "## English Summary",
        "- 342O packages simulated adopted cells only.",
        "- It does not write back to upstream workbooks and does not create a formal client export.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- ready_for_342p: {summary.get('ready_for_342p', False)}",
        f"- recommended_342p_scope: {summary.get('recommended_342p_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Key Counts",
        f"- pending_review_count: {summary.get('pending_review_count', 0)}",
        f"- input_adoption_candidate_count: {summary.get('input_adoption_candidate_count', 0)}",
        f"- direct_adopted_count: {summary.get('direct_adopted_count', 0)}",
        f"- corrected_adopted_count: {summary.get('corrected_adopted_count', 0)}",
        f"- simulated_adopted_cell_count: {summary.get('simulated_adopted_cell_count', 0)}",
        f"- still_human_required_count: {summary.get('still_human_required_count', 0)}",
        f"- remaining_review_count: {summary.get('remaining_review_count', 0)}",
        f"- reduction_rate_after_342o: {summary.get('reduction_rate_after_342o', 0)}",
        "",
        "## Coverage",
        f"- metric_covered_count: {summary.get('metric_covered_count', 0)}",
        f"- metric_year_pair_count: {summary.get('metric_year_pair_count', 0)}",
        f"- correction_pattern_count: {summary.get('correction_pattern_count', 0)}",
        f"- REVENUE_AMOUNT_NOT_YOY_count: {summary.get('REVENUE_AMOUNT_NOT_YOY_count', 0)}",
        f"- REVENUE_YOY_PERCENT_count: {summary.get('REVENUE_YOY_PERCENT_count', 0)}",
        f"- NET_PROFIT_YOY_PERCENT_count: {summary.get('NET_PROFIT_YOY_PERCENT_count', 0)}",
        "",
        "## Safety",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Boundaries",
            "- 342O does not rerun MinerU.",
            "- 342O does not call a real LLM API.",
            "- 342O does not modify production pipeline / parser / extraction / delivery.",
            "- 342O does not write back to 342I / 342J / 342M / 342N or earlier workbooks.",
            "- The result is simulation only and must not be treated as investment advice.",
            "",
        ]
    )
    return "\n".join(lines)
