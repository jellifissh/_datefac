from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from datefac.review_queue.spot_check_package_343f import EDITABLE_SPOT_CHECK_COLUMNS


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "value",
        "tags",
        "note",
        "errors",
        "recommendation",
        "risk",
        "boundary",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if str(cell.value or "") in EDITABLE_SPOT_CHECK_COLUMNS:
                    cell.fill = editable_fill
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 343F AI-assisted Review Spot-check Package",
        "",
        "## 中文摘要",
        "- 343F 生成一个等待人工 spot-check 的小包，覆盖全部 30 行 AI-assisted review 结果。",
        "- 它不会导入 spot-check 结果，也不会声称 strict human review 或 human spot-check 已完成。",
        "- formal client export 仍然禁止，production readiness 仍然是 false。",
        "",
        "## English Summary",
        "- 343F creates a waiting-for-spot-check package for AI-assisted review results.",
        "- It does not ingest spot-check results yet and does not claim completed strict human review.",
        "- Formal client export remains forbidden and production readiness remains false.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- waiting_for_spot_check: {summary.get('waiting_for_spot_check', False)}",
        f"- spot_check_result_ingested: {summary.get('spot_check_result_ingested', False)}",
        f"- ready_for_343g: {summary.get('ready_for_343g', False)}",
        f"- recommended_343g_scope: {summary.get('recommended_343g_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Package Metrics",
        f"- input_apply_plan_row_count: {summary.get('input_apply_plan_row_count', 0)}",
        f"- input_simulated_sidecar_row_count: {summary.get('input_simulated_sidecar_row_count', 0)}",
        f"- spot_check_item_count: {summary.get('spot_check_item_count', 0)}",
        f"- simulated_applied_spot_check_count: {summary.get('simulated_applied_spot_check_count', 0)}",
        f"- source_check_required_count: {summary.get('source_check_required_count', 0)}",
        f"- skipped_hold_count: {summary.get('skipped_hold_count', 0)}",
        f"- priority_tier_count: {summary.get('priority_tier_count', 0)}",
        "",
        "## AI-assisted Boundary",
        f"- review_source_type: {summary.get('review_source_type', '')}",
        f"- not_pure_human_review: {summary.get('not_pure_human_review', False)}",
        f"- strict_human_review_completed: {summary.get('strict_human_review_completed', False)}",
        f"- requires_human_spot_check: {summary.get('requires_human_spot_check', False)}",
        f"- apply_mode: {summary.get('apply_mode', '')}",
        "",
        "## Safety Boundary",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        "",
        "## Reviewer Guidance",
        "- Review all 10 simulated-applied rows and all 20 held rows in the dedicated workbook.",
        "- Source-check-required rows should usually remain `SOURCE_CHECK_REQUIRED` unless evidence becomes sufficient.",
        "- No spot-check decision is completed until a later filled workbook ingestion stage.",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    return "\n".join(lines)


def reviewer_instructions_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# 343F Spot-check Reviewer Instructions",
            "",
            "- This workbook is for spot-check preparation only.",
            "- Current source remains AI-assisted review plus simulation-only downstream planning.",
            "- Fill only the `spot_check_*` columns, `spot_checker_id`, and `spot_checked_at`.",
            "- Do not change identity/action/evidence columns.",
            "- Do not claim completed strict human review in this stage.",
            "",
            f"Current decision: {summary.get('decision', '')}",
        ]
    )


def ai_assisted_boundary_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# 343F AI-assisted Boundary",
            "",
            "- review_source_type: AI_ASSISTED_REVIEW",
            "- not_pure_human_review: true",
            "- strict_human_review_completed: false",
            "- requires_human_spot_check: true",
            "- apply_mode: SIMULATION_ONLY",
            "- formal_client_export_allowed: false",
            "- client_ready: false",
            "- production_ready: false",
            "",
            "This package prepares future human spot-check only.",
            "It does not mean the spot-check has already been completed.",
            "",
            f"Current decision: {summary.get('decision', '')}",
        ]
    )
