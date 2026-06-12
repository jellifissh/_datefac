from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows]
    path.write_text("\n".join(lines), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "value",
        "note",
        "reason",
        "risk",
        "limitation",
        "recommendation",
        "gap",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 343H AI-assisted Spot-check Audit Summary And Strict Human Gap Report",
        "",
        "## 中文摘要",
        "- 343H 汇总 343A-343G 的 AI-assisted review 与 AI-assisted spot-check 链路，并生成 strict-human-gap report。",
        "- 当前 10 行只是 AI-assisted spot-check confirmed，不是 strict-human confirmed。",
        "- 当前 19 行仍需 source check，另有 1 行 keep-hold，formal client export 仍然禁止。",
        "",
        "## English Summary",
        "- 343H summarizes the 343A-343G AI-assisted review and AI-assisted spot-check chain and generates a strict-human-gap report.",
        "- The current 10 confirmed rows are AI-assisted spot-check confirmed only, not strict-human confirmed.",
        "- 19 rows still require source check and 1 row remains keep-hold, so formal client export remains forbidden.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- input_spot_check_result_row_count: {summary.get('input_spot_check_result_row_count', 0)}",
        f"- ai_assisted_confirmed_count: {summary.get('ai_assisted_confirmed_count', 0)}",
        f"- source_check_required_count: {summary.get('source_check_required_count', 0)}",
        f"- keep_hold_count: {summary.get('keep_hold_count', 0)}",
        f"- strict_human_gap_item_count: {summary.get('strict_human_gap_item_count', 0)}",
        f"- source_check_backlog_count: {summary.get('source_check_backlog_count', 0)}",
        f"- ready_for_343i: {summary.get('ready_for_343i', False)}",
        f"- recommended_343i_scope: {summary.get('recommended_343i_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Boundary",
        f"- review_source_type: {summary.get('review_source_type', '')}",
        f"- spot_check_source_type: {summary.get('spot_check_source_type', '')}",
        f"- not_pure_human_review: {summary.get('not_pure_human_review', False)}",
        f"- strict_human_review_completed: {summary.get('strict_human_review_completed', False)}",
        f"- requires_strict_human_review: {summary.get('requires_strict_human_review', False)}",
        f"- apply_mode: {summary.get('apply_mode', '')}",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        "",
        "## Next Safe Step",
        "- Default recommendation: build a strict human review package for the 10 AI-assisted confirmed rows.",
        "- Secondary option: resolve the 19 source-check backlog rows.",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    return "\n".join(lines)


def strict_human_gap_report_markdown(
    summary: Dict[str, Any],
    *,
    confirmed_count: int,
    source_check_count: int,
    keep_hold_count: int,
) -> str:
    return "\n".join(
        [
            "# 343H Strict Human Gap Report",
            "",
            "## 中文摘要",
            f"- AI-assisted confirmed 行数: {confirmed_count}",
            f"- source-check backlog 行数: {source_check_count}",
            f"- keep-hold 行数: {keep_hold_count}",
            "- 这些结果仍然不是严格人工审核完成结果。",
            "",
            "## English Summary",
            f"- AI-assisted confirmed row count: {confirmed_count}",
            f"- source-check backlog row count: {source_check_count}",
            f"- keep-hold row count: {keep_hold_count}",
            "- These results are still not strict human review completion.",
            "",
            "## Why export remains blocked",
            "- review_source_type remains AI_ASSISTED_REVIEW.",
            "- spot_check_source_type remains AI_ASSISTED_SPOT_CHECK.",
            "- strict_human_review_completed remains false.",
            "- requires_strict_human_review remains true.",
            "- formal_client_export_allowed/client_ready/production_ready remain false.",
            "",
            f"Current decision: {summary.get('decision', '')}",
        ]
    )
