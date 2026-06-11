from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_PREVIEW_SUMMARY",
    "02_INPUT_342O_SUMMARY",
    "03_INPUT_342J_SUMMARY",
    "04_COMBINED_PREVIEW",
    "05_HUMAN_REVIEWED",
    "06_SIM_DIRECT",
    "07_SIM_CORRECTED",
    "08_STILL_HUMAN_REQUIRED",
    "09_COLLISION_CHECK",
    "10_METRIC_COVERAGE",
    "11_PREVIEW_BOUNDARY",
    "12_342Q_READINESS",
    "13_NO_WRITE_BACK",
    "14_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "warning",
        "path",
        "bbox",
        "html",
        "evidence",
        "recommendation",
        "value",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 180)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 220)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 342P Reviewed Plus Simulated Client Preview Pilot",
        "",
        "## Chinese Summary",
        "- 342P combines real 342J human-reviewed preview rows and 342O simulated adopted rows into one bounded preview pilot.",
        "- HUMAN_REVIEWED rows and SIMULATED rows remain explicitly separated.",
        "- Simulated rows are not final confirmation and still require later audit before any client delivery claim.",
        "",
        "## English Summary",
        "- 342P is a reviewed plus simulated client preview pilot only.",
        "- It remains no-write-back, not client-ready, and not production-ready.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- ready_for_342q: {summary.get('ready_for_342q', False)}",
        f"- recommended_342q_scope: {summary.get('recommended_342q_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Key Counts",
        f"- human_reviewed_preview_count: {summary.get('human_reviewed_preview_count', 0)}",
        f"- simulated_preview_count: {summary.get('simulated_preview_count', 0)}",
        f"- simulated_direct_preview_count: {summary.get('simulated_direct_preview_count', 0)}",
        f"- simulated_corrected_preview_count: {summary.get('simulated_corrected_preview_count', 0)}",
        f"- combined_preview_row_count: {summary.get('combined_preview_row_count', 0)}",
        f"- still_human_required_count: {summary.get('still_human_required_count', 0)}",
        f"- remaining_review_count: {summary.get('remaining_review_count', 0)}",
        "",
        "## Coverage",
        f"- metric_covered_count: {summary.get('metric_covered_count', 0)}",
        f"- metric_year_pair_count: {summary.get('metric_year_pair_count', 0)}",
        f"- human_metric_year_pair_count: {summary.get('human_metric_year_pair_count', 0)}",
        f"- simulated_metric_year_pair_count: {summary.get('simulated_metric_year_pair_count', 0)}",
        "",
        "## Collision Audit",
        f"- duplicate_review_item_id_count: {summary.get('duplicate_review_item_id_count', 0)}",
        f"- duplicate_metric_year_source_count: {summary.get('duplicate_metric_year_source_count', 0)}",
        f"- human_over_simulation_override_count: {summary.get('human_over_simulation_override_count', 0)}",
        f"- simulated_duplicate_dropped_count: {summary.get('simulated_duplicate_dropped_count', 0)}",
        f"- collision_logged_count: {summary.get('collision_logged_count', 0)}",
        "",
        "## Safety",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- output_workbook_path: {summary.get('output_workbook_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Boundaries",
            "- 342P does not rerun MinerU.",
            "- 342P does not rerun 342E / 342F / 342G / 342H / 342I / 342J / 342K / 342L / 342M / 342N / 342O.",
            "- 342P does not call VLM or a real LLM API.",
            "- 342P does not write back to upstream workbooks.",
            "- 342P must not be treated as formal client delivery or investment advice.",
            "",
        ]
    )
    return "\n".join(lines)
