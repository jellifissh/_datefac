from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_CORPUS_SUMMARY",
    "02_PDF_CORPUS",
    "03_DEDUP_AUDIT",
    "04_TIER_ASSIGNMENT",
    "05_SPLIT_PLAN",
    "06_METADATA_AUDIT",
    "07_RUN_READINESS",
    "08_RISK_FLAGS",
    "09_NO_WRITE_BACK_PROOF",
    "10_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "notes",
        "status",
        "warning",
        "reference",
        "hash",
        "risk",
        "goal",
        "output",
        "input",
        "criteria",
        "driver",
        "description",
        "rationale",
        "path",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 180)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 220)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# Real PDF Corpus Intake And Metadata Audit 342B",
        "",
        "342B creates a sidecar corpus-intake and metadata-audit package for the current real-PDF benchmark pool.",
        "It is an intake and readiness stage only, not a MinerU batch parse run and not a client export stage.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- warning_count: {summary.get('warning_count', 0)}",
        f"- ready_for_342c: {summary.get('ready_for_342c', False)}",
        f"- recommended_342c_scope: {summary.get('recommended_342c_scope', '')}",
        f"- recommended_first_run_pdf_count: {summary.get('recommended_first_run_pdf_count', 0)}",
        "",
        "## Corpus Counts",
        f"- current_pdf_count: {summary.get('current_pdf_count', 0)}",
        f"- unique_pdf_count: {summary.get('unique_pdf_count', 0)}",
        f"- duplicate_pdf_count: {summary.get('duplicate_pdf_count', 0)}",
        f"- assigned_tier_count: {summary.get('assigned_tier_count', 0)}",
        f"- unknown_tier_count: {summary.get('unknown_tier_count', 0)}",
        f"- pilot_set_count: {summary.get('pilot_set_count', 0)}",
        f"- benchmark_set_count: {summary.get('benchmark_set_count', 0)}",
        f"- holdout_set_count: {summary.get('holdout_set_count', 0)}",
        "",
        "## Metadata Audit",
        f"- missing_sha256_count: {summary.get('missing_sha256_count', 0)}",
        f"- unreadable_pdf_count: {summary.get('unreadable_pdf_count', 0)}",
        f"- missing_page_count_count: {summary.get('missing_page_count_count', 0)}",
        f"- oversized_pdf_count: {summary.get('oversized_pdf_count', 0)}",
        f"- zero_byte_file_count: {summary.get('zero_byte_file_count', 0)}",
        "",
        "## Safety",
        f"- detected_342a_decision: {summary.get('detected_342a_decision', '')}",
        f"- detected_342a_benchmark_status: {summary.get('detected_342a_benchmark_status', '')}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        f"- output_workbook_path: {summary.get('output_workbook_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.append("")
    return "\n".join(lines)
