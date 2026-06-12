from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from datefac.review_queue.strict_human_review_package_343i import (
    EDITABLE_STRICT_REVIEW_COLUMNS,
)


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "value",
        "note",
        "reason",
        "recommendation",
        "boundary",
        "disclosure",
        "scope",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if str(cell.value or "") in EDITABLE_STRICT_REVIEW_COLUMNS:
                    cell.fill = editable_fill
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 343I Strict Human Review Package For AI-assisted Confirmed Rows",
        "",
        "## 中文摘要",
        "- 343I 只为 10 条 AI-assisted confirmed 行生成严格人工复核包。",
        "- 本阶段不导入严格人工复核结果，也不代表严格人工复核已经完成。",
        "- 19 条 source-check backlog 仍然单独保留，不进入本次可填写 strict review 模板。",
        "- formal client export 仍然禁止，client_ready 和 production_ready 仍然保持 false。",
        "",
        "## English Summary",
        "- 343I creates a strict human review package only for the AI-assisted confirmed rows.",
        "- It does not ingest strict human review results yet and does not mean strict human review is complete.",
        "- The source-check backlog remains separate from this fillable package.",
        "- Formal client export remains forbidden and readiness flags remain false.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- input_ai_assisted_confirmed_count: {summary.get('input_ai_assisted_confirmed_count', 0)}",
        f"- strict_review_item_count: {summary.get('strict_review_item_count', 0)}",
        f"- source_check_backlog_context_count: {summary.get('source_check_backlog_context_count', 0)}",
        f"- strict_human_gap_item_count: {summary.get('strict_human_gap_item_count', 0)}",
        f"- waiting_for_strict_human_review: {summary.get('waiting_for_strict_human_review', False)}",
        f"- strict_human_review_result_ingested: {summary.get('strict_human_review_result_ingested', False)}",
        f"- strict_human_review_completed: {summary.get('strict_human_review_completed', False)}",
        f"- ready_for_343j: {summary.get('ready_for_343j', False)}",
        f"- recommended_343j_scope: {summary.get('recommended_343j_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Safety Boundary",
        f"- requires_strict_human_review: {summary.get('requires_strict_human_review', False)}",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        "",
        "## Reviewer Guidance",
        "- Fill only the strict_review_* columns, strict_reviewer_id, and strict_reviewed_at.",
        "- STRICT_CORRECT requires corrected metric/year/value/unit fields.",
        "- STRICT_REJECT and STRICT_NEEDS_SOURCE_CHECK require notes.",
        "- This package is waiting for strict human review and must not be described as export-ready.",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    return "\n".join(lines)


def reviewer_instructions_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# 343I Strict Human Review Instructions",
            "",
            "- This workbook is for strict human review of AI-assisted confirmed rows only.",
            "- AI-assisted confirmation is not enough for formal client export.",
            "- Check identity, evidence, AI-assisted note, and downstream action context before deciding.",
            "- Fill only the strict_review_* columns, strict_reviewer_id, and strict_reviewed_at.",
            "- Do not edit queue_item_id, review_item_id, resulting_status, or evidence context fields.",
            "",
            "Allowed decisions:",
            "- STRICT_CONFIRM",
            "- STRICT_CORRECT",
            "- STRICT_REJECT",
            "- STRICT_NEEDS_SOURCE_CHECK",
            "- STRICT_DEFER",
            "",
            f"Current decision: {summary.get('decision', '')}",
        ]
    )


def fill_guide_markdown() -> str:
    return "\n".join(
        [
            "# 343I Fill Guide",
            "",
            "- `STRICT_CONFIRM`: strict human reviewer independently confirms the current row.",
            "- `STRICT_CORRECT`: fill corrected metric/year/value/unit fields and explain the correction in `strict_review_note` when useful.",
            "- `STRICT_REJECT`: reject the row and provide a reason in `strict_review_note`.",
            "- `STRICT_NEEDS_SOURCE_CHECK`: current evidence is still insufficient; provide what is missing in `strict_review_note`.",
            "- `STRICT_DEFER`: defer to a later batch and explain why when possible.",
            "",
            "Required for `STRICT_CORRECT`:",
            "- strict_review_metric_standardized",
            "- strict_review_year_standardized",
            "- strict_review_value_numeric",
            "- strict_review_normalized_unit",
            "",
            "Save the filled workbook under:",
            "- D:/_datefac/input/review_queue_strict_human_review_343i_filled/",
        ]
    )


def client_export_boundary_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# 343I Client Export Boundary",
            "",
            "- strict_human_review_completed: false",
            "- requires_strict_human_review: true",
            "- formal_client_export_allowed: false",
            "- client_ready: false",
            "- production_ready: false",
            "",
            "343I only prepares a strict human review package.",
            "It does not ingest results, does not apply changes, and does not allow formal client export.",
            "",
            f"Current decision: {summary.get('decision', '')}",
        ]
    )
