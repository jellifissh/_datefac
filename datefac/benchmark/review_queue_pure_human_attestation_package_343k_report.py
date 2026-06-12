from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from datefac.review_queue.pure_human_attestation_package_343k import (
    EDITABLE_HUMAN_ATTESTATION_COLUMNS,
)


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows),
        encoding="utf-8",
    )


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "value",
        "note",
        "reason",
        "recommendation",
        "evidence",
        "bbox",
        "html",
        "text",
        "snippet",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                if str(cell.value or "") in EDITABLE_HUMAN_ATTESTATION_COLUMNS:
                    cell.fill = editable_fill
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 343K Pure Human Confirmation Attestation Package",
        "",
        "## 中文摘要",
        "- 343K 为 10 条 AI-assisted strict-confirm 行生成 pure-human confirmation attestation package。",
        "- 343K 只生成 package 和 fillable template，不导入 attestation 结果。",
        "- 当前仍不能声称 pure strict human review completed，也不能允许 formal client export。",
        "- 人工审核人必须独立查看 source evidence，再填写 `human_attestation_decision`。",
        "",
        "## English Summary",
        "- 343K creates a pure-human confirmation attestation package for the AI-assisted strict-confirm rows.",
        "- It does not ingest attestation results yet and does not complete pure strict human review.",
        "- A human reviewer must independently inspect the source evidence before attesting any row.",
        "",
        "## Key Metrics",
        f"- decision: {summary.get('decision', '')}",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- input_ai_assisted_strict_review_confirm_count: {summary.get('input_ai_assisted_strict_review_confirm_count', 0)}",
        f"- attestation_item_count: {summary.get('attestation_item_count', 0)}",
        f"- evidence_resolved_count: {summary.get('evidence_resolved_count', 0)}",
        f"- source_pdf_name_available_count: {summary.get('source_pdf_name_available_count', 0)}",
        f"- source_text_snippet_available_count: {summary.get('source_text_snippet_available_count', 0)}",
        f"- waiting_for_pure_human_attestation: {summary.get('waiting_for_pure_human_attestation', False)}",
        f"- pure_human_attestation_result_ingested: {summary.get('pure_human_attestation_result_ingested', False)}",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- ready_for_343l: {summary.get('ready_for_343l', False)}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(
            f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})"
        )
    return "\n".join(lines)
