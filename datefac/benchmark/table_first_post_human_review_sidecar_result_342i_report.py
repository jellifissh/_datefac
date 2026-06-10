from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_RESULT_SUMMARY",
    "02_INPUT_342H_SUMMARY",
    "03_HUMAN_REVIEWED_CELLS",
    "04_FINAL_CONFIRMED",
    "05_FINAL_CORRECTED",
    "06_FINAL_REJECTED",
    "07_PENDING_REVIEW",
    "08_BEFORE_AFTER",
    "09_SOURCE_TRACE",
    "10_METRIC_COVERAGE_AFTER",
    "11_UNIT_YEAR_AFTER",
    "12_REMAINING_RISKS",
    "13_342J_READINESS",
    "14_NO_WRITE_BACK",
    "15_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "note",
        "warning",
        "reference",
        "hash",
        "risk",
        "path",
        "decision",
        "action",
        "recommendation",
        "html",
        "snippet",
        "bbox",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 180)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 220)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 342I Table-First Post-Human-Review Sidecar Result",
        "",
        "中文：",
        "342I 只把当前 342H 已应用的人审结果整理成 post-human-review sidecar result，不会写回任何上游 workbook，也不是正式财务结果。",
        f"本次只覆盖首批 {summary.get('reviewed_row_count', 0)} 条 human-reviewed cells，仍保留 {summary.get('pending_review_count', 0)} 条 pending review。",
        "当前仍保持 client_ready = false、production_ready = false。",
        "",
        "English:",
        "342I packages the currently applied 342H human-review result into a post-human-review sidecar result.",
        f"It covers only the current reviewed batch of {summary.get('reviewed_row_count', 0)} cells and preserves {summary.get('pending_review_count', 0)} pending review rows.",
        "It does not write back upstream workbooks and it is neither client-ready nor production-ready.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- ready_for_342j: {summary.get('ready_for_342j', False)}",
        f"- recommended_342j_scope: {summary.get('recommended_342j_scope', '')}",
        "",
        "## Counts",
        f"- input_review_template_row_count: {summary.get('input_review_template_row_count', 0)}",
        f"- reviewed_row_count: {summary.get('reviewed_row_count', 0)}",
        f"- pending_review_count: {summary.get('pending_review_count', 0)}",
        f"- final_confirmed_cell_count: {summary.get('final_confirmed_cell_count', 0)}",
        f"- final_corrected_cell_count: {summary.get('final_corrected_cell_count', 0)}",
        f"- final_rejected_cell_count: {summary.get('final_rejected_cell_count', 0)}",
        f"- post_human_confirmed_count: {summary.get('post_human_confirmed_count', 0)}",
        f"- metric_covered_after_human_count: {summary.get('metric_covered_after_human_count', 0)}",
        f"- metric_year_pair_after_human_count: {summary.get('metric_year_pair_after_human_count', 0)}",
        f"- remaining_review_count: {summary.get('remaining_review_count', 0)}",
        f"- unit_year_remaining_count: {summary.get('unit_year_remaining_count', 0)}",
        f"- duplicate_remaining_count: {summary.get('duplicate_remaining_count', 0)}",
        f"- growth_row_remaining_count: {summary.get('growth_row_remaining_count', 0)}",
        "",
        "## Safety",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- output_workbook_path: {summary.get('output_workbook_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Boundaries",
            "- 342I does not rerun MinerU.",
            "- 342I does not call VLM/LLM.",
            "- 342I does not modify production pipeline / parser / extraction / delivery.",
            "- 342I remains a no-write-back sidecar result stage.",
            "- Next step can be 342J reviewed client preview pilot or expanded human review batches.",
            "",
        ]
    )
    return "\n".join(lines)
