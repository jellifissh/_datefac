from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows),
        encoding="utf-8",
    )


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "value",
        "note",
        "reason",
        "recommendation",
        "evidence",
        "bbox",
        "html",
        "text",
        "snippet",
        "scope",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 344C Source-check Confirmed Sidecar Apply Simulation And Expanded Trust Gate",
        "",
        "## 中文摘要",
        "- 344C 已模拟应用 344B 的 19 条 source-check resolved rows。",
        "- 其中 10 条为 source-check confirmed，9 条为 source-check corrected。",
        "- corrected rows 延续 344B 的 `YOY / %` 语义，而不是旧的 `revenue / 亿元` 语义。",
        "- 与 343O/343N 既有 10-row demo trusted arc 合并后，expanded trusted candidate coverage 达到 29 行。",
        "- 本任务仍然只是 sidecar apply simulation，不做真实写回，不做 formal client export。",
        "- 下一步应进入 344D expanded trusted export package generation for review/demo only。",
        "",
        "## English Summary",
        "- 344C simulates applying 19 source-check resolved rows from 344B.",
        "- 10 rows are source-check confirmed and 9 rows are source-check corrected.",
        "- The corrected rows retain YOY/% semantics from 344B.",
        "- Combined with the 10-row demo trusted arc, expanded trusted coverage reaches 29 rows.",
        "- No production write-back or formal client export occurred.",
        "",
        "## Key Metrics",
        f"- decision: {summary.get('decision', '')}",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- source_check_input_sidecar_row_count: {summary.get('source_check_input_sidecar_row_count', 0)}",
        f"- source_check_apply_plan_row_count: {summary.get('source_check_apply_plan_row_count', 0)}",
        f"- source_check_apply_confirm_count: {summary.get('source_check_apply_confirm_count', 0)}",
        f"- source_check_apply_correct_count: {summary.get('source_check_apply_correct_count', 0)}",
        f"- source_check_apply_blocked_count: {summary.get('source_check_apply_blocked_count', 0)}",
        f"- source_check_applied_sidecar_row_count: {summary.get('source_check_applied_sidecar_row_count', 0)}",
        f"- corrections_applied_count: {summary.get('corrections_applied_count', 0)}",
        f"- prior_demo_trusted_row_count: {summary.get('prior_demo_trusted_row_count', 0)}",
        f"- source_check_trusted_row_count: {summary.get('source_check_trusted_row_count', 0)}",
        f"- expanded_trusted_candidate_count: {summary.get('expanded_trusted_candidate_count', 0)}",
        f"- deduplicated_expanded_trusted_candidate_count: {summary.get('deduplicated_expanded_trusted_candidate_count', 0)}",
        f"- dedup_conflict_count: {summary.get('dedup_conflict_count', 0)}",
        f"- expanded_trusted_scope: {summary.get('expanded_trusted_scope', '')}",
        f"- source_check_sidecar_apply_simulation_completed: {summary.get('source_check_sidecar_apply_simulation_completed', False)}",
        f"- expanded_trust_gate_evaluated: {summary.get('expanded_trust_gate_evaluated', False)}",
        f"- source_check_backlog_resolved: {summary.get('source_check_backlog_resolved', False)}",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- global_strict_human_review_completed: {summary.get('global_strict_human_review_completed', False)}",
        f"- ready_for_344d: {summary.get('ready_for_344d', False)}",
        f"- recommended_344d_scope: {summary.get('recommended_344d_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(
            f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})"
        )
    return "\n".join(lines)

