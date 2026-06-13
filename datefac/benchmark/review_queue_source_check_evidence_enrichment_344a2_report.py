from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from datefac.review_queue.source_check_backlog_package_344a import (
    EDITABLE_SOURCE_CHECK_COLUMNS_344A,
)


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows),
        encoding="utf-8",
    )


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    wrap_keywords = {
        "description",
        "detail",
        "path",
        "message",
        "warning",
        "value",
        "reason",
        "recommendation",
        "note",
        "snippet",
        "bbox",
        "artifact",
        "html",
        "text",
        "contract",
        "rule",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if str(cell.value or "") in EDITABLE_SOURCE_CHECK_COLUMNS_344A:
                    cell.fill = editable_fill
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 344A2 Source Evidence Enrichment For Source-check Backlog",
        "",
        "## 中文摘要",
        "- 344A2 读取 344A 的 19 条 source-check backlog，并扫描现有 342/343/344 产物补充源证据定位字段。",
        "- 本任务只做 evidence enrichment，不确认、不纠正、不驳回任何 backlog 行，也不预填 source_check_decision。",
        "- 若证据不足，会如实保留为 PARTIAL 或 UNRESOLVED，不伪造 PDF / page / image / snippet。",
        "- formal client export、client_ready、production_ready 仍全部保持 false。",
        "",
        "## English Summary",
        "- 344A2 enriches the 19 source-check backlog rows with source evidence locators from existing artifacts.",
        "- It does not decide row outcomes and does not prefill source-check decisions.",
        "- Missing evidence is reported honestly as partial or unresolved.",
        "",
        "## Key Metrics",
        f"- decision: {summary.get('decision', '')}",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- input_source_check_backlog_item_count: {summary.get('input_source_check_backlog_item_count', 0)}",
        f"- deduplicated_backlog_item_count: {summary.get('deduplicated_backlog_item_count', 0)}",
        f"- evidence_resolved_count: {summary.get('evidence_resolved_count', 0)}",
        f"- evidence_partial_count: {summary.get('evidence_partial_count', 0)}",
        f"- evidence_unresolved_count: {summary.get('evidence_unresolved_count', 0)}",
        f"- source_pdf_name_available_count: {summary.get('source_pdf_name_available_count', 0)}",
        f"- page_number_available_count: {summary.get('page_number_available_count', 0)}",
        f"- image_path_available_count: {summary.get('image_path_available_count', 0)}",
        f"- source_text_snippet_available_count: {summary.get('source_text_snippet_available_count', 0)}",
        f"- match_candidate_count: {summary.get('match_candidate_count', 0)}",
        f"- high_confidence_match_count: {summary.get('high_confidence_match_count', 0)}",
        f"- medium_confidence_match_count: {summary.get('medium_confidence_match_count', 0)}",
        f"- low_confidence_match_count: {summary.get('low_confidence_match_count', 0)}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(
            f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})"
        )
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json["warnings"]:
            lines.append(f"- {warning}")
    return "\n".join(lines)


def unresolved_report_markdown(
    summary: Dict[str, Any],
    unresolved_rows: list[Dict[str, Any]],
) -> str:
    lines = [
        "# 344A2 Unresolved Evidence Report",
        "",
        "## 中文摘要",
        f"- evidence_resolved_count: {summary.get('evidence_resolved_count', 0)}",
        f"- evidence_partial_count: {summary.get('evidence_partial_count', 0)}",
        f"- evidence_unresolved_count: {summary.get('evidence_unresolved_count', 0)}",
        "- 这些行尚未被自动补足可审阅证据，必须保守处理，不得假装已完成 source check。",
        "",
        "## English Summary",
        f"- Resolved rows: {summary.get('evidence_resolved_count', 0)}",
        f"- Partial rows: {summary.get('evidence_partial_count', 0)}",
        f"- Unresolved rows: {summary.get('evidence_unresolved_count', 0)}",
    ]
    if not unresolved_rows:
        lines.extend(["", "- No fully unresolved rows remain in the current enriched package."])
        return "\n".join(lines)
    lines.extend(["", "## Unresolved Items"])
    for row in unresolved_rows:
        lines.append(
            f"- {row.get('queue_item_id', '')} / {row.get('review_item_id', '')}: {row.get('evidence_gap_reason', '')}"
        )
    return "\n".join(lines)


def reviewer_instructions_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# 344A2 Reviewer Instructions",
            "",
            "## 中文说明",
            "- 344A2 只补证据，不做 source-check 决策。",
            "- 请优先查看 `04_REVIEW_TEMPLATE`，按 `evidence_resolution_status` 排序后从 `RESOLVED` 行开始复核。",
            "- 使用 `source_pdf_name / page_number / table_id / bbox / image_path / source_text_snippet / source_html_snippet` 作为定位线索。",
            "- 若证据仍不足，请保持保守决策，例如 `SOURCE_STILL_INSUFFICIENT` 或 `SOURCE_DEFER`。",
            "",
            "## English Notes",
            "- 344A2 enriches evidence only and does not decide row outcomes.",
            f"- waiting_for_source_check_review = {summary.get('waiting_for_source_check_review', False)}",
            "- Save the filled workbook under `D:/_datefac/input/review_queue_source_check_evidence_344a2_filled/` for later ingestion.",
        ]
    )


def fill_guide_markdown() -> str:
    return "\n".join(
        [
            "# 344A2 Fill Guide",
            "",
            "1. Open `04_REVIEW_TEMPLATE` in the enriched workbook.",
            "2. Keep all non-editable evidence columns unchanged.",
            "3. Fill only the editable source-check columns.",
            "4. Allowed `source_check_decision` values:",
            "- `SOURCE_CONFIRM`",
            "- `SOURCE_CORRECT`",
            "- `SOURCE_REJECT`",
            "- `SOURCE_STILL_INSUFFICIENT`",
            "- `SOURCE_DEFER`",
            "5. For `SOURCE_CORRECT`, fill corrected metric/year/value/unit fields.",
            "6. Save the filled workbook into `D:/_datefac/input/review_queue_source_check_evidence_344a2_filled/`.",
        ]
    )


def artifact_search_report_markdown(
    searched_artifacts: list[Dict[str, Any]],
    match_candidate_count: int,
) -> str:
    lines = [
        "# 344A2 Artifact Search Report",
        "",
        "## Summary",
        f"- searched_artifact_count: {len(searched_artifacts)}",
        f"- match_candidate_count: {match_candidate_count}",
        "",
        "## Artifacts",
    ]
    for artifact in searched_artifacts:
        lines.append(
            f"- {artifact.get('source_stage', '')}: {artifact.get('artifact_path', '')} "
            f"(type={artifact.get('artifact_type', '')}, rows={artifact.get('row_count', 0)})"
        )
    return "\n".join(lines)


def scope_boundary_markdown(lines: list[str]) -> str:
    body = ["# 344A2 Scope Boundary", ""]
    body.extend(f"- {line}" for line in lines)
    return "\n".join(body)

