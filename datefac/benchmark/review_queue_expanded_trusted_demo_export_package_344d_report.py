from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows), encoding="utf-8")


def write_csv(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows).to_csv(path, index=False, encoding="utf-8-sig")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "value",
        "note",
        "reason",
        "recommendation",
        "evidence",
        "bbox",
        "html",
        "text",
        "snippet",
        "labels",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 344D Expanded Trusted Export Package Generation For Review/Demo Only",
        "",
        "## 中文摘要",
        "- 344D 生成了一个 29 行的 expanded trusted review/demo package。",
        "- 该包合并了 10 条 prior demo trusted rows 与 19 条 source-check resolved rows。",
        "- 其中 10 条 source-check rows 为 confirmed，9 条为 corrected，corrected rows 保留 YOY/% 语义。",
        "- 该包仅用于 review/demo，不是 formal client export，不是 production-ready，也没有发生生产写回。",
        "",
        "## English Summary",
        "- 344D generates a 29-row expanded trusted package for review/demo only.",
        "- It combines the 10-row demo arc with 19 source-check resolved rows.",
        "- Ten source-check rows are confirmed and nine are corrected with YOY/% semantics preserved.",
        "- No production write-back or formal client export occurred.",
        "",
        "## Key Metrics",
        f"- decision: {summary.get('decision', '')}",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- input_expanded_trusted_candidate_count: {summary.get('input_expanded_trusted_candidate_count', 0)}",
        f"- expanded_export_row_count: {summary.get('expanded_export_row_count', 0)}",
        f"- audit_label_row_count: {summary.get('audit_label_row_count', 0)}",
        f"- prior_demo_trusted_row_count: {summary.get('prior_demo_trusted_row_count', 0)}",
        f"- source_check_trusted_row_count: {summary.get('source_check_trusted_row_count', 0)}",
        f"- source_check_confirmed_row_count: {summary.get('source_check_confirmed_row_count', 0)}",
        f"- source_check_corrected_row_count: {summary.get('source_check_corrected_row_count', 0)}",
        f"- correction_row_count: {summary.get('correction_row_count', 0)}",
        f"- dedup_conflict_count: {summary.get('dedup_conflict_count', 0)}",
        f"- expanded_export_scope: {summary.get('expanded_export_scope', '')}",
        f"- export_usage: {summary.get('export_usage', '')}",
        f"- expanded_review_demo_package_generated: {summary.get('expanded_review_demo_package_generated', False)}",
        f"- expanded_demo_handoff_ready: {summary.get('expanded_demo_handoff_ready', False)}",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- ready_for_344e: {summary.get('ready_for_344e', False)}",
        f"- recommended_344e_scope: {summary.get('recommended_344e_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    return "\n".join(lines)
