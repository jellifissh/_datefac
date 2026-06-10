from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_TABLE_QUALITY_SUMMARY",
    "02_PDF_TABLE_SIGNAL",
    "03_ALL_TABLE_BLOCKS",
    "04_TABLE_TYPE_COVERAGE",
    "05_CORE_EXTRACTABLE",
    "06_METADATA_EXTRACTABLE",
    "07_EXCLUDED_TABLES",
    "08_SOURCE_TRACE_AUDIT",
    "09_342F_READINESS",
    "10_NO_WRITE_BACK_PROOF",
    "11_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "notes",
        "warning",
        "reference",
        "hash",
        "risk",
        "goal",
        "output",
        "input",
        "criteria",
        "driver",
        "description",
        "rationale",
        "path",
        "command",
        "error",
        "excerpt",
        "context",
        "html",
        "caption",
        "footnote",
        "bbox",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 180)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 220)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# Core Metric Candidate Quality Audit 342E",
        "",
        "342E now runs as a table-first MinerU sidecar audit. It reads existing MinerU table evidence and classifies tables into core-extractable, metadata-only, excluded, or manual-review buckets.",
        "It does not continue from the older 435 text candidates as a formal extraction source.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- ready_for_342f: {summary.get('ready_for_342f', '')}",
        f"- recommended_342f_scope: {summary.get('recommended_342f_scope', '')}",
        "",
        "## Table Counts",
        f"- audited_pdf_count: {summary.get('audited_pdf_count', 0)}",
        f"- total_table_block_count: {summary.get('total_table_block_count', 0)}",
        f"- core_extractable_table_count: {summary.get('core_extractable_table_count', 0)}",
        f"- metadata_extractable_table_count: {summary.get('metadata_extractable_table_count', 0)}",
        f"- excluded_table_count: {summary.get('excluded_table_count', 0)}",
        f"- manual_review_required_count: {summary.get('manual_review_required_count', 0)}",
        f"- pdf_with_core_extractable_table_count: {summary.get('pdf_with_core_extractable_table_count', 0)}",
        f"- table_source_file_count: {summary.get('table_source_file_count', 0)}",
        "",
        "## Safety",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        "",
        "## Output",
        f"- output_workbook_path: {summary.get('output_workbook_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.append("")
    return "\n".join(lines)
