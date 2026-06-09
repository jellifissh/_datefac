from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_RETRY_SUMMARY",
    "02_RETRY_PARSE_RESULTS",
    "03_OUTPUT_ARTIFACT_AUDIT",
    "04_ORIG_342C_RECAP",
    "05_RETRY_FAILURE_AUDIT",
    "06_EMPTY_OUTPUT_AUDIT",
    "07_342D_READINESS",
    "08_ENV_CONTEXT",
    "09_NO_WRITE_BACK_PROOF",
    "10_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "notes",
        "status",
        "warning",
        "reference",
        "hash",
        "risk",
        "goal",
        "output",
        "input",
        "criteria",
        "driver",
        "description",
        "rationale",
        "path",
        "command",
        "error",
        "excerpt",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 180)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 220)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# MinerU Pilot Retry With Verified Local Environment 342C2",
        "",
        "342C2 reruns the 342B pilot_set with the verified local MinerU environment to test whether 342C failed because of runner environment or model-cache setup.",
        "It is still a sidecar parse benchmark only. It does not modify production pipeline behavior and does not generate client export assets.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- ready_for_342d: {summary.get('ready_for_342d', '')}",
        f"- recommended_next_scope: {summary.get('recommended_next_scope', '')}",
        "",
        "## Original 342C Failure Recap",
        f"- original_342c_failure_detected: {summary.get('original_342c_failure_detected', False)}",
        f"- original_342c_ssl_failure_detected: {summary.get('original_342c_ssl_failure_detected', False)}",
        f"- original_342c_huggingface_detected: {summary.get('original_342c_huggingface_detected', False)}",
        f"- original_342c_empty_output_count: {summary.get('original_342c_empty_output_count', 0)}",
        "",
        "## Retry Counts",
        f"- retry_pilot_total_count: {summary.get('retry_pilot_total_count', 0)}",
        f"- retry_mineru_success_count: {summary.get('retry_mineru_success_count', 0)}",
        f"- retry_mineru_failed_count: {summary.get('retry_mineru_failed_count', 0)}",
        f"- empty_output_count: {summary.get('empty_output_count', 0)}",
        "",
        "## Verified Environment",
        f"- verified_working_lab_dir: {summary.get('verified_working_lab_dir', '')}",
        f"- verified_model_cache_dir: {summary.get('verified_model_cache_dir', '')}",
        f"- mineru_command: {summary.get('mineru_command', '')}",
        f"- output_workbook_path: {summary.get('output_workbook_path', '')}",
        "",
        "## Safety",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.append("")
    return "\n".join(lines)
