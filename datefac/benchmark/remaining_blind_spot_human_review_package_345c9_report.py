from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


WORKBOOK_SHEETS = [
    "review_required",
    "context_only",
    "blocked_or_too_generic",
    "decision_options",
    "reviewer_checklist",
    "package_summary",
]

EDITABLE_COLUMNS = {
    "human_blind_spot_review_decision",
    "approved_standard_metric",
    "approved_new_standard_metric",
    "needs_alias_family_expansion",
    "needs_source_context",
    "reviewer",
    "reviewed_at",
    "review_notes",
    "alias_rule_update_allowed",
}


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_csv(path: Path, rows: List[Dict[str, Any]], fieldnames: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    key: json.dumps(value, ensure_ascii=False)
                    if isinstance(value, (dict, list))
                    else value
                    for key, value in row.items()
                }
            )


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    wrap_keywords = {
        "message",
        "note",
        "reason",
        "excerpt",
        "risk",
        "pdf_names",
        "sample_row_ids",
        "source_stages",
        "source_artifacts",
        "value",
        "decision",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                if str(cell.value or "") in EDITABLE_COLUMNS:
                    cell.fill = editable_fill
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )
    _format_workbook(path)


def reviewer_checklist_markdown(summary: Dict[str, Any], workbook_path: str) -> str:
    reviewed_copy_path = workbook_path.replace(".xlsx", "_reviewed.xlsx")
    return "\n".join(
        [
            "# 345C9 Reviewer Checklist",
            "",
            "1. Open `remaining_blind_spot_human_review_package_345c9.xlsx`.",
            "2. Fill only the human review fields in the `review_required` sheet.",
            "3. Do not edit evidence fields, source fields, impact estimates, or group labels.",
            "4. Use only the allowed human decision options from `decision_options`.",
            "5. Do not set `alias_rule_update_allowed` to true in this package.",
            "6. Use `context_only` as reference only, not as actionable review rows.",
            "7. Use `blocked_or_too_generic` to understand what is intentionally deferred or blocked.",
            "8. Save the reviewed workbook separately.",
            "",
            "Example reviewed workbook path:",
            reviewed_copy_path,
            "",
            "Current boundary:",
            f"- formal_client_export_allowed = {summary.get('formal_client_export_allowed', False)}",
            f"- client_ready = {summary.get('client_ready', False)}",
            f"- production_ready = {summary.get('production_ready', False)}",
            f"- global_strict_human_review_completed = {summary.get('global_strict_human_review_completed', False)}",
            "- 345C9 does not modify normalization rules.",
            "- 345C9 does not modify official alias assets.",
            "- 345C9 does not apply aliases to upstream data.",
        ]
    )


def decision_options_markdown() -> str:
    return "\n".join(
        [
            "# 345C9 Human Decision Options",
            "",
            "- `APPROVE_EXISTING_MAPPING`: requires `approved_standard_metric`.",
            "- `APPROVE_NEW_STANDARD`: requires `approved_new_standard_metric`.",
            "- `REJECT_TOO_GENERIC`: should include `review_notes`.",
            "- `NEEDS_SOURCE_CONTEXT`: should explain missing source context in `review_notes`.",
            "- `DEFER`: should include `review_notes` when possible.",
            "",
            "Reviewer warning:",
            "- Do not approve generic fragments such as `其他`, `变动`, `成本`, empty names, unit-like fragments, or suffix-like fragments unless source context proves a stable financial meaning.",
            "- `alias_rule_update_allowed` must remain false in this package.",
        ]
    )


def executive_summary_markdown(summary: Dict[str, Any], top_rows: List[Dict[str, Any]]) -> str:
    lines = [
        "# 345C9 Remaining Blind Spot Human Review Package",
        "",
        "## Input 345C8 Context",
        f"- input_345c8_decision: {summary.get('input_345c8_decision', '')}",
        f"- input_alias_branch_stop_or_continue_decision: {summary.get('input_alias_branch_stop_or_continue_decision', '')}",
        f"- selected_candidate_count: {summary.get('selected_candidate_count', 0)}",
        "",
        "## Why 345C9 Exists",
        "- 345C8 selected a bounded second review batch from the remaining alias blind spots.",
        "- 345C9 packages those candidates for human review only.",
        "- 345C9 does not perform semantic adjudication by itself.",
        "",
        "## Package Totals",
        f"- review_required_row_count: {summary.get('review_required_row_count', 0)}",
        f"- context_only_row_count: {summary.get('context_only_row_count', 0)}",
        f"- blocked_or_too_generic_row_count: {summary.get('blocked_or_too_generic_row_count', 0)}",
        f"- review_required_estimated_row_impact_total: {summary.get('review_required_estimated_row_impact_total', 0)}",
        f"- review_required_estimated_coverage_delta_total: {summary.get('review_required_estimated_coverage_delta_total', 0)}",
        "",
        "## Risk Distribution",
        f"- review_required_high_risk_count: {summary.get('review_required_high_risk_count', 0)}",
        f"- review_required_medium_risk_count: {summary.get('review_required_medium_risk_count', 0)}",
        f"- review_required_low_risk_count: {summary.get('review_required_low_risk_count', 0)}",
        "",
        "## Top Actionable Rows",
    ]
    for row in top_rows:
        lines.append(
            f"- {row.get('raw_metric_name', '')}: rows={row.get('remaining_row_count', 0)}, "
            f"priority={row.get('candidate_priority', '')}, risk={row.get('risk_level', '')}, "
            f"recommendation={row.get('review_recommendation', '')}"
        )
    lines.extend(
        [
            "",
            "## Boundary",
            "- No normalization rules were changed.",
            "- No official alias assets were changed.",
            f"- formal_client_export_allowed = {summary.get('formal_client_export_allowed', False)}",
            f"- client_ready = {summary.get('client_ready', False)}",
            f"- production_ready = {summary.get('production_ready', False)}",
            f"- global_strict_human_review_completed = {summary.get('global_strict_human_review_completed', False)}",
            "",
            "## Next",
            "- Human reviewer fills the 345C9 workbook.",
            "- 345C10 Second Batch Reviewed Alias Decision Ingestion.",
            "- 345C11 Second Batch Alias Apply Simulation.",
            "- Then decide whether alias governance stops and work returns to 345D Full Structured Demo Export Package.",
            "- 344G still waits for a genuinely human-filled 344F workbook.",
        ]
    )
    return "\n".join(lines)


def artifact_index_markdown(rows: Iterable[Dict[str, Any]]) -> str:
    lines = [
        "# 345C9 Artifact Index",
        "",
        "| Artifact | Path | Use |",
        "| --- | --- | --- |",
    ]
    for row in rows:
        lines.append(
            f"| {row.get('artifact_name', '')} | {row.get('path', '')} | {row.get('purpose', '')} |"
        )
    return "\n".join(lines)


def next_plan_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# 345C9 Next Plan",
            "",
            "- Human reviewer fills the `review_required` sheet in the 345C9 workbook.",
            "- 345C10 Second Batch Reviewed Alias Decision Ingestion.",
            "- 345C11 Second Batch Alias Apply Simulation.",
            "- After second-batch review impact is measured, decide whether alias governance stops and work returns to `345D Full Structured Demo Export Package`.",
            "",
            "Current status reminder:",
            f"- input_alias_branch_stop_or_continue_decision: {summary.get('input_alias_branch_stop_or_continue_decision', '')}",
            f"- full_structured_demo_export_reasonable_after_345c9 = {summary.get('full_structured_demo_export_reasonable_after_345c9', False)}",
            "- 344G still waits for a genuinely human-filled 344F workbook.",
        ]
    )

