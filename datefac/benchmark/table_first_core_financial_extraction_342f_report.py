from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_EXTRACTION_SUMMARY",
    "02_INPUT_CORE_TABLES",
    "03_LONG_FORM_CELLS",
    "04_TRUSTED_CELLS",
    "05_REVIEW_REQUIRED",
    "06_REJECTED_CELLS",
    "07_METRIC_COVERAGE",
    "08_UNIT_NORMALIZATION",
    "09_TABLE_TRACE",
    "10_342G_READINESS",
    "11_NO_WRITE_BACK_PROOF",
    "12_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "notes",
        "warning",
        "reference",
        "hash",
        "risk",
        "goal",
        "output",
        "input",
        "criteria",
        "driver",
        "description",
        "rationale",
        "path",
        "command",
        "error",
        "excerpt",
        "context",
        "html",
        "snippet",
        "bbox",
        "flags",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 180)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 220)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# Table-First Core Financial Extraction 342F",
        "",
        "342F is a table-first sidecar long-form extraction pilot built only from 342E core-extractable tables. It does not modify production extraction logic and does not generate client export assets.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- ready_for_342g: {summary.get('ready_for_342g', '')}",
        f"- recommended_342g_scope: {summary.get('recommended_342g_scope', '')}",
        "",
        "## Extraction Counts",
        f"- audited_pdf_count: {summary.get('audited_pdf_count', 0)}",
        f"- input_core_extractable_table_count: {summary.get('input_core_extractable_table_count', 0)}",
        f"- parsed_core_table_count: {summary.get('parsed_core_table_count', 0)}",
        f"- html_parse_failed_table_count: {summary.get('html_parse_failed_table_count', 0)}",
        f"- long_form_cell_count: {summary.get('long_form_cell_count', 0)}",
        f"- trusted_cell_count: {summary.get('trusted_cell_count', 0)}",
        f"- review_required_cell_count: {summary.get('review_required_cell_count', 0)}",
        f"- rejected_cell_count: {summary.get('rejected_cell_count', 0)}",
        "",
        "## Risks",
        f"- metric_covered_count: {summary.get('metric_covered_count', 0)}",
        f"- metric_year_pair_count: {summary.get('metric_year_pair_count', 0)}",
        f"- unit_issue_count: {summary.get('unit_issue_count', 0)}",
        f"- year_header_issue_count: {summary.get('year_header_issue_count', 0)}",
        f"- duplicate_cell_count: {summary.get('duplicate_cell_count', 0)}",
        f"- table_trace_count: {summary.get('table_trace_count', 0)}",
        "",
        "## Safety",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        "",
        "## Output",
        f"- output_workbook_path: {summary.get('output_workbook_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.append("")
    return "\n".join(lines)
