from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {"description", "detail", "rule", "path", "message", "warning", "value", "note", "reason", "recommendation", "evidence", "bbox", "html", "text", "snippet"}
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 343M Human-confirmed Sidecar Apply Simulation And Limited Export Gate",
        "",
        "## 中文摘要",
        "- 343M 只模拟把 10 条 package-level human-confirmed rows 应用到 sidecar，不做真实写回。",
        "- limited export candidate 只限 `343K_PACKAGE_ONLY` 范围，不能视为 formal client export。",
        "- remaining backlog/source-check rows 仍然存在，因此 global strict human review 仍为 false。",
        "- formal client export、client_ready、production_ready 都继续保持 false。",
        "",
        "## English Summary",
        "- 343M simulates applying the human-confirmed package rows into a sidecar-only result set.",
        "- The limited candidate is restricted to `343K_PACKAGE_ONLY` and is not a formal client export.",
        "- Remaining backlog still blocks global readiness.",
        "",
        "## Key Metrics",
        f"- decision: {summary.get('decision', '')}",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- input_human_attested_row_count: {summary.get('input_human_attested_row_count', 0)}",
        f"- valid_human_attested_row_count: {summary.get('valid_human_attested_row_count', 0)}",
        f"- sidecar_row_count: {summary.get('sidecar_row_count', 0)}",
        f"- sidecar_human_accept_count: {summary.get('sidecar_human_accept_count', 0)}",
        f"- sidecar_human_correct_count: {summary.get('sidecar_human_correct_count', 0)}",
        f"- sidecar_blocked_count: {summary.get('sidecar_blocked_count', 0)}",
        f"- limited_export_candidate_row_count: {summary.get('limited_export_candidate_row_count', 0)}",
        f"- remaining_source_check_backlog_count: {summary.get('remaining_source_check_backlog_count', 0)}",
        f"- limited_package_export_candidate_allowed: {summary.get('limited_package_export_candidate_allowed', False)}",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- ready_for_343n: {summary.get('ready_for_343n', False)}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    return "\n".join(lines)
