from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from datefac.review_queue.spot_check_package_343f import EDITABLE_SPOT_CHECK_COLUMNS


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "value",
        "tags",
        "note",
        "errors",
        "recommendation",
        "risk",
        "gap",
        "disclosure",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if str(cell.value or "") in EDITABLE_SPOT_CHECK_COLUMNS:
                    cell.fill = editable_fill
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 343G AI-assisted Review Spot-check Result Ingestion",
        "",
        "## 中文摘要",
        "- 343G 读取已填写的 343F spot-check workbook，并把每行 decision 校验后转换成 spot-check result sidecar。",
        "- 当前 workbook 属于 AI-assisted spot-check，不是 strict pure human spot-check，必须继续保留披露字段。",
        "- 343G 只做 ingestion 和 validation，不做真实写回、不做 production apply、不做 formal client export。",
        "",
        "## English Summary",
        "- 343G ingests a filled 343F spot-check workbook and converts validated decisions into a spot-check result sidecar.",
        "- The workbook is AI-assisted spot-check, not strict pure human spot-check, so disclosure must remain explicit.",
        "- This stage performs ingestion and validation only, with no real write-back, no production apply, and no formal client export.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- filled_workbook_path: {summary.get('filled_workbook_path', '')}",
        f"- spot_check_result_ingested: {summary.get('spot_check_result_ingested', False)}",
        f"- ready_for_343h: {summary.get('ready_for_343h', False)}",
        f"- recommended_343h_scope: {summary.get('recommended_343h_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Ingestion Metrics",
        f"- filled_spot_check_row_count: {summary.get('filled_spot_check_row_count', 0)}",
        f"- valid_row_count: {summary.get('valid_row_count', 0)}",
        f"- invalid_row_count: {summary.get('invalid_row_count', 0)}",
        f"- confirm_ai_assisted_result_count: {summary.get('confirm_ai_assisted_result_count', 0)}",
        f"- correct_ai_assisted_result_count: {summary.get('correct_ai_assisted_result_count', 0)}",
        f"- reject_ai_assisted_result_count: {summary.get('reject_ai_assisted_result_count', 0)}",
        f"- source_check_required_count: {summary.get('source_check_required_count', 0)}",
        f"- keep_hold_count: {summary.get('keep_hold_count', 0)}",
        f"- skip_spot_check_count: {summary.get('skip_spot_check_count', 0)}",
        f"- validation_error_count: {summary.get('validation_error_count', 0)}",
        f"- validation_warning_count: {summary.get('validation_warning_count', 0)}",
        "",
        "## AI-assisted Spot-check Disclosure",
        f"- review_source_type: {summary.get('review_source_type', '')}",
        f"- spot_check_source_type: {summary.get('spot_check_source_type', '')}",
        f"- not_pure_human_review: {summary.get('not_pure_human_review', False)}",
        f"- strict_human_review_completed: {summary.get('strict_human_review_completed', False)}",
        f"- requires_strict_human_review: {summary.get('requires_strict_human_review', False)}",
        f"- apply_mode: {summary.get('apply_mode', '')}",
        "",
        "## Why formal client export remains forbidden",
        "- The source workbook is AI-assisted spot-check rather than strict pure human verification.",
        "- Strict human review is still incomplete and remains explicitly required.",
        "- 343G only serializes a validated sidecar result JSONL plus audit workbook.",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    return "\n".join(lines)


def ai_assisted_spot_check_disclosure_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# 343G AI-assisted Spot-check Disclosure",
            "",
            "- review_source_type: AI_ASSISTED_REVIEW",
            "- spot_check_source_type: AI_ASSISTED_SPOT_CHECK",
            "- not_pure_human_review: true",
            "- strict_human_review_completed: false",
            "- requires_strict_human_review: true",
            "- apply_mode: SIMULATION_ONLY",
            "",
            "This ingestion result must not be described as strict pure human spot-check completion.",
            "It is a useful AI-assisted spot-check sidecar result, but strict human review remains required.",
            "",
            f"Current decision: {summary.get('decision', '')}",
        ]
    )
