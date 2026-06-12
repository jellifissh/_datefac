from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_ROUND_TRIP_SUMMARY",
    "02_INPUT_343A_SUMMARY",
    "03_REVIEW_TEMPLATE_SPEC",
    "04_EXPORT_TEMPLATE_ROWS",
    "05_IMPORT_SIMULATION",
    "06_VALIDATION_RULES",
    "07_STATUS_MAPPING",
    "08_DECISION_MAPPING",
    "09_ERROR_CASES",
    "10_REVIEWED_RESULT",
    "11_BACKLOG_NOTE",
    "12_343C_READINESS",
    "13_NO_WRITE_BACK",
    "14_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows]
    path.write_text("\n".join(lines), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "example",
        "allowed",
        "validation",
        "requirement",
        "section",
        "value",
        "tags",
        "note",
        "errors",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 343B Excel Round-trip Review Queue Pilot",
        "",
        "## 中文摘要",
        "- 343B 用 Excel round-trip 的方式验证 343A Review Queue schema，可以导出 review template、导入 deterministic review simulation、运行校验，并生成 reviewed result JSONL。",
        "- 343B 不代表真实 human review，不实现 Argilla，不实现完整 UI，也不写回任何上游 workbook。",
        "- 343B 证明 Review Queue contract 已经足够稳定，因此后续 343C 可以把 Argilla 当作 pluggable UI，而不是系统核心。",
        "",
        "## English Summary",
        "- 343B validates the 343A Review Queue schema through an Excel round-trip pilot.",
        "- It exports a review template, simulates import decisions, validates the rows, and produces reviewed-result JSONL.",
        "- It does not represent real human review, does not implement Argilla, and does not write back to upstream artifacts.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- ready_for_343c: {summary.get('ready_for_343c', False)}",
        f"- recommended_343c_scope: {summary.get('recommended_343c_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Round-trip Metrics",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- template_row_count: {summary.get('template_row_count', 0)}",
        f"- import_simulation_row_count: {summary.get('import_simulation_row_count', 0)}",
        f"- reviewed_result_row_count: {summary.get('reviewed_result_row_count', 0)}",
        f"- confirmed_count: {summary.get('confirmed_count', 0)}",
        f"- corrected_count: {summary.get('corrected_count', 0)}",
        f"- rejected_count: {summary.get('rejected_count', 0)}",
        f"- needs_source_check_count: {summary.get('needs_source_check_count', 0)}",
        f"- skipped_count: {summary.get('skipped_count', 0)}",
        f"- validation_error_count: {summary.get('validation_error_count', 0)}",
        f"- validation_warning_count: {summary.get('validation_warning_count', 0)}",
        "",
        "## Safety Boundary",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        "",
        "## Why formal client export remains forbidden",
        "- 343B only validates the review-queue contract and deterministic Excel round-trip mechanics.",
        "- The import rows are simulated reviewer actions, not real human-reviewed evidence.",
        "- Upstream package risk and backlog boundaries from 342R/342S still remain.",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Recommended Open Order",
            "- First open `review_queue_excel_round_trip_343b.xlsx` and inspect `01_ROUND_TRIP_SUMMARY`, `04_EXPORT_TEMPLATE_ROWS`, `05_IMPORT_SIMULATION`, and `10_REVIEWED_RESULT`.",
            "- Then open the dedicated review template workbook and reviewed-result JSONL.",
            "",
            "## Next Recommendation",
            "- Recommended next step: `343C Argilla Human Review UI Pilot`.",
            "- 343C should treat Argilla as a pluggable UI on top of the validated queue contract, not as a replacement for the queue schema.",
            "",
        ]
    )
    return "\n".join(lines)
