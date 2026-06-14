from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill


WORKBOOK_SHEETS = [
    "00_README",
    "01_REVIEW_SUMMARY",
    "02_REVIEW_ROWS",
    "03_DECISION_OPTIONS",
    "04_REVIEWER_CHECKLIST",
    "05_LLM_SUGGESTION_SUM",
    "06_PRIORITY_SUMMARY",
    "07_NO_WRITE_BACK",
    "08_NEXT_STEPS",
]

EDITABLE_COLUMNS = {
    "human_alias_review_decision",
    "approved_standard_metric",
    "approved_new_standard_metric",
    "alias_reviewer",
    "alias_reviewed_at",
    "alias_review_notes",
}


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_csv(path: Path, rows: List[Dict[str, Any]], fieldnames: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(
                {
                    key: json.dumps(value, ensure_ascii=False)
                    if isinstance(value, (dict, list))
                    else value
                    for key, value in row.items()
                }
            )


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    wrap_keywords = {
        "message",
        "note",
        "reason",
        "excerpt",
        "focus",
        "flags",
        "pdf_names",
        "sample_row_ids",
        "source_stages",
        "value",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if str(cell.value or "") in EDITABLE_COLUMNS:
                    cell.fill = editable_fill
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def reviewer_checklist_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# 345C4 Reviewer Checklist",
            "",
            "1. This package is for human review of LLM alias suggestions only.",
            "2. LLM suggestions are not approved by default.",
            "3. Review raw metric name, frequency, source stages, sample row ids, LLM action, reason, evidence excerpt, confidence, and validation status.",
            "4. Fill only the human review fields.",
            "5. Do not edit LLM evidence fields or source fields.",
            "6. This task does not modify normalization rules or official alias assets.",
            "7. 345C5 or later should ingest reviewed decisions.",
            "",
            "## Current Boundary",
            f"- formal_client_export_allowed = {summary.get('formal_client_export_allowed', False)}",
            f"- client_ready = {summary.get('client_ready', False)}",
            f"- production_ready = {summary.get('production_ready', False)}",
            f"- global_strict_human_review_completed = {summary.get('global_strict_human_review_completed', False)}",
        ]
    )


def decision_options_markdown() -> str:
    return "\n".join(
        [
            "# 345C4 Human Decision Options",
            "",
            "- `APPROVE_EXISTING_MAPPING`: requires `approved_standard_metric`.",
            "- `APPROVE_NEW_STANDARD`: requires `approved_new_standard_metric` and should usually include notes.",
            "- `REJECT_ALIAS`: should include a note.",
            "- `NEEDS_MORE_CONTEXT`: should describe missing context.",
            "- `DEFER`: means later tasks should not use this row for apply simulation.",
            "",
            "Generated rows keep `alias_rule_update_allowed = false` by default.",
        ]
    )


def executive_summary_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# 345C4 Alias Suggestion Human Review Package",
            "",
            "## Input 345C2 Live Result",
            f"- input_345c2_decision: {summary.get('input_345c2_decision', '')}",
            f"- input_llm_mode: {summary.get('input_llm_mode', '')}",
            f"- input_suggestion_row_count: {summary.get('input_suggestion_row_count', 0)}",
            "",
            "## Why Every Suggestion Needs Human Review",
            "- 345C2 live suggestions are semantic sidecar proposals only.",
            "- No suggestion is pre-approved for alias-rule updates.",
            "- Validation failures, low confidence, and new-standard proposals require explicit human review.",
            "",
            "## Review Package Totals",
            f"- review_row_count: {summary.get('review_row_count', 0)}",
            f"- llm_propose_new_standard_count: {summary.get('llm_propose_new_standard_count', 0)}",
            f"- llm_insufficient_evidence_count: {summary.get('llm_insufficient_evidence_count', 0)}",
            f"- validation_failed_count: {summary.get('validation_failed_count', 0)}",
            "",
            "## Boundary",
            "- No normalization rules were changed.",
            "- No official alias assets were changed.",
            f"- formal_client_export_allowed = {summary.get('formal_client_export_allowed', False)}",
            f"- client_ready = {summary.get('client_ready', False)}",
            f"- production_ready = {summary.get('production_ready', False)}",
            "",
            "## Next",
            "- 345C5 Reviewed Alias Decision Ingestion",
            "- 345C6 Reviewed Alias Apply Simulation",
            "- 345D Full Structured Demo Export Package only after reviewed alias impact is measured",
            "- 344G still waits for a genuinely human-filled 344F workbook.",
        ]
    )


def artifact_index_markdown(rows: Iterable[Dict[str, Any]]) -> str:
    lines = [
        "# 345C4 Artifact Index",
        "",
        "| Artifact | Path | Use |",
        "| --- | --- | --- |",
    ]
    for row in rows:
        lines.append(
            f"| {row.get('artifact_name', '')} | {row.get('path', '')} | {row.get('purpose', '')} |"
        )
    return "\n".join(lines)


def next_plan_markdown() -> str:
    return "\n".join(
        [
            "# 345C4 Next Plan",
            "",
            "- 345C5 Reviewed Alias Decision Ingestion",
            "- 345C6 Reviewed Alias Apply Simulation",
            "- 345D Full Structured Demo Export Package only after reviewed alias impact is measured",
            "",
            "Boundary reminder:",
            "- 345C4 does not modify normalization rules.",
            "- 345C4 does not modify official alias assets.",
            "- 344G still waits for a genuinely human-filled 344F workbook.",
        ]
    )
