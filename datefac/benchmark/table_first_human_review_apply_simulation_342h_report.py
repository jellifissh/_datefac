from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_APPLY_SUMMARY",
    "02_INPUT_REVIEW_STATUS",
    "03_VALIDATED_DECISIONS",
    "04_CONFIRMED_CELLS",
    "05_CORRECTED_CELLS",
    "06_REJECTED_CELLS",
    "07_STILL_REVIEW",
    "08_NEEDS_SOURCE_CHECK",
    "09_PENDING_REVIEW",
    "10_REVIEW_ERRORS",
    "11_BEFORE_AFTER",
    "12_SOURCE_TRACE",
    "13_342I_READINESS",
    "14_NO_WRITE_BACK",
    "15_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "note",
        "warning",
        "reference",
        "hash",
        "risk",
        "path",
        "command",
        "error",
        "html",
        "snippet",
        "bbox",
        "status",
        "decision",
        "action",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 180)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 220)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    waiting = summary.get("decision", "").endswith("WAITING_FOR_HUMAN_REVIEW")
    lines = [
        "# 342H Table-First Human Review Apply Simulation",
        "",
        "中文：",
        (
            "342H 当前处于 waiting-for-human-review 状态，因为 reviewed workbook 还不存在。"
            if waiting
            else "342H 已读取 reviewed workbook，并执行了 human review apply simulation dry-run。"
        ),
        "342H 不是正式财务结果，不会写回任何上游 workbook，也仍然不是 client-ready / production-ready。",
        "",
        "English:",
        (
            "342H is currently waiting for human review because the reviewed workbook does not exist yet."
            if waiting
            else "342H consumed a reviewed workbook and performed a human-review apply simulation dry run."
        ),
        "It does not write back to upstream workbooks and is not client-ready or production-ready.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- ready_for_342i: {summary.get('ready_for_342i', False)}",
        f"- recommended_342i_scope: {summary.get('recommended_342i_scope', '')}",
        f"- recommended_next_action: {summary.get('recommended_next_action', '')}",
        "",
        "## Counts",
        f"- input_review_template_row_count: {summary.get('input_review_template_row_count', 0)}",
        f"- reviewed_row_count: {summary.get('reviewed_row_count', 0)}",
        f"- pending_review_count: {summary.get('pending_review_count', 0)}",
        f"- confirmed_cell_count: {summary.get('confirmed_cell_count', 0)}",
        f"- corrected_cell_count: {summary.get('corrected_cell_count', 0)}",
        f"- rejected_cell_count: {summary.get('rejected_cell_count', 0)}",
        f"- not_core_metric_count: {summary.get('not_core_metric_count', 0)}",
        f"- still_review_required_count: {summary.get('still_review_required_count', 0)}",
        f"- needs_source_check_count: {summary.get('needs_source_check_count', 0)}",
        f"- validation_error_count: {summary.get('validation_error_count', 0)}",
        "",
        "## Safety",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- output_workbook_path: {summary.get('output_workbook_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Boundaries",
            "- 342H does not rerun MinerU.",
            "- 342H does not call VLM/LLM.",
            "- 342H does not modify production pipeline / parser / extraction / delivery.",
            "- 342H remains a no-write-back dry-run sidecar stage.",
            "- Next step is fill_342g_review_template_first or 342I, depending on human review availability and validation quality.",
            "",
        ]
    )
    return "\n".join(lines)
