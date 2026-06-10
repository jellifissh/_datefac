from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_REVIEW_SUMMARY",
    "02_INPUT_342F_SUMMARY",
    "03_REVIEW_QUEUE",
    "04_TRUSTED_AUDIT",
    "05_UNIT_YEAR_ISSUES",
    "06_DUPLICATE_ISSUES",
    "07_GROWTH_ROW_ISSUES",
    "08_TABLE_TRACE",
    "09_REVIEW_GUIDE",
    "10_REVIEW_TEMPLATE",
    "11_DECISION_OPTIONS",
    "12_342H_READINESS",
    "13_NO_WRITE_BACK",
    "14_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "notes",
        "warning",
        "reference",
        "hash",
        "risk",
        "goal",
        "output",
        "input",
        "criteria",
        "driver",
        "description",
        "rationale",
        "path",
        "command",
        "error",
        "excerpt",
        "context",
        "html",
        "snippet",
        "bbox",
        "flags",
        "option",
        "meaning",
        "guide",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 180)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 220)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 342G Table-First Extraction Review Package",
        "",
        "中文：",
        "342G 是建立在 342F 之上的 sidecar review package，用来把 table-first long-form extraction 结果整理成可人工复核的 workbook。它不是正式财务结果，也不会写回任何上游 workbook。",
        "",
        "English:",
        "342G is a sidecar review package built on top of completed 342F outputs. It organizes table-first long-form extraction into a human-review workbook and does not write back to upstream workbooks.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- ready_for_342h: {summary.get('ready_for_342h', '')}",
        f"- recommended_342h_scope: {summary.get('recommended_342h_scope', '')}",
        "",
        "## Input Counts",
        f"- audited_pdf_count: {summary.get('audited_pdf_count', 0)}",
        f"- input_long_form_cell_count: {summary.get('input_long_form_cell_count', 0)}",
        f"- input_trusted_cell_count: {summary.get('input_trusted_cell_count', 0)}",
        f"- input_review_required_cell_count: {summary.get('input_review_required_cell_count', 0)}",
        f"- input_rejected_cell_count: {summary.get('input_rejected_cell_count', 0)}",
        "",
        "## Review Package Counts",
        f"- review_queue_count: {summary.get('review_queue_count', 0)}",
        f"- trusted_audit_sample_count: {summary.get('trusted_audit_sample_count', 0)}",
        f"- unit_year_issue_count: {summary.get('unit_year_issue_count', 0)}",
        f"- duplicate_issue_count: {summary.get('duplicate_issue_count', 0)}",
        f"- growth_row_issue_count: {summary.get('growth_row_issue_count', 0)}",
        f"- high_priority_review_count: {summary.get('high_priority_review_count', 0)}",
        f"- medium_priority_review_count: {summary.get('medium_priority_review_count', 0)}",
        f"- low_priority_review_count: {summary.get('low_priority_review_count', 0)}",
        f"- review_template_row_count: {summary.get('review_template_row_count', 0)}",
        "",
        "## Safety",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        "",
        "## Output",
        f"- output_workbook_path: {summary.get('output_workbook_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Boundaries",
            "- 342G does not rerun MinerU.",
            "- 342G does not call VLM/LLM.",
            "- 342G does not modify production pipeline / parser / extraction / delivery.",
            "- 342G remains not client-ready and not production-ready.",
            "- Next step is 342H Table-First Human Review Apply Simulation.",
            "",
        ]
    )
    return "\n".join(lines)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 342G Table-First Extraction Review Package",
        "",
        "中文：",
        "342G 是建立在 342F 之上的 sidecar review package，用来把 table-first long-form extraction 结果整理成可人工复核的 workbook。它不是正式财务结果，也不会写回任何上游 workbook。",
        "",
        "English:",
        "342G is a sidecar review package built on top of completed 342F outputs. It organizes table-first long-form extraction into a human-review workbook and does not write back to upstream workbooks.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- ready_for_342h: {summary.get('ready_for_342h', '')}",
        f"- recommended_342h_scope: {summary.get('recommended_342h_scope', '')}",
        "",
        "## Input Counts",
        f"- audited_pdf_count: {summary.get('audited_pdf_count', 0)}",
        f"- input_long_form_cell_count: {summary.get('input_long_form_cell_count', 0)}",
        f"- input_trusted_cell_count: {summary.get('input_trusted_cell_count', 0)}",
        f"- input_review_required_cell_count: {summary.get('input_review_required_cell_count', 0)}",
        f"- input_rejected_cell_count: {summary.get('input_rejected_cell_count', 0)}",
        "",
        "## Review Package Counts",
        f"- review_queue_count: {summary.get('review_queue_count', 0)}",
        f"- trusted_audit_sample_count: {summary.get('trusted_audit_sample_count', 0)}",
        f"- unit_year_issue_count: {summary.get('unit_year_issue_count', 0)}",
        f"- duplicate_issue_count: {summary.get('duplicate_issue_count', 0)}",
        f"- growth_row_issue_count: {summary.get('growth_row_issue_count', 0)}",
        f"- high_priority_review_count: {summary.get('high_priority_review_count', 0)}",
        f"- medium_priority_review_count: {summary.get('medium_priority_review_count', 0)}",
        f"- low_priority_review_count: {summary.get('low_priority_review_count', 0)}",
        f"- review_template_row_count: {summary.get('review_template_row_count', 0)}",
        "",
        "## Safety",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        "",
        "## Output",
        f"- output_workbook_path: {summary.get('output_workbook_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Boundaries",
            "- 342G does not rerun MinerU.",
            "- 342G does not call VLM/LLM.",
            "- 342G does not modify production pipeline / parser / extraction / delivery.",
            "- 342G remains not client-ready and not production-ready.",
            "- Next step is 342H Table-First Human Review Apply Simulation.",
            "",
        ]
    )
    return "\n".join(lines)
