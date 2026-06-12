from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from datefac.review_queue.source_evidence_enrichment_343i2 import (
    EDITABLE_STRICT_REVIEW_COLUMNS,
)


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "value",
        "note",
        "reason",
        "recommendation",
        "evidence",
        "bbox",
        "html",
        "text",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if str(cell.value or "") in EDITABLE_STRICT_REVIEW_COLUMNS:
                    cell.fill = editable_fill
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 343I2 Source Evidence Enrichment For Strict Human Review Package",
        "",
        "## 中文摘要",
        "- 343I 暴露出 strict human review 模板缺少原始证据定位信息，导致无法独立回看 PDF/page/table。",
        "- 343I2 基于可用的 343D 与 342R 证据链，给 strict review items 补充 source locator 字段。",
        "- 没找到的证据不会伪造，会明确标记为 PARTIAL 或 UNRESOLVED，并写出 evidence gap reason。",
        "- strict human review 仍未完成，formal client export 仍然禁止。",
        "",
        "## English Summary",
        "- 343I2 enriches the strict human review package with any available source evidence locators.",
        "- Missing evidence is not fabricated; it is marked as partial or unresolved.",
        "- Strict human review is still pending and formal client export remains forbidden.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- input_strict_review_item_count: {summary.get('input_strict_review_item_count', 0)}",
        f"- enriched_review_item_count: {summary.get('enriched_review_item_count', 0)}",
        f"- evidence_resolved_count: {summary.get('evidence_resolved_count', 0)}",
        f"- evidence_partial_count: {summary.get('evidence_partial_count', 0)}",
        f"- evidence_unresolved_count: {summary.get('evidence_unresolved_count', 0)}",
        f"- source_pdf_name_available_count: {summary.get('source_pdf_name_available_count', 0)}",
        f"- source_pdf_path_available_count: {summary.get('source_pdf_path_available_count', 0)}",
        f"- page_number_available_count: {summary.get('page_number_available_count', 0)}",
        f"- image_path_available_count: {summary.get('image_path_available_count', 0)}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Reviewer Guidance",
        "- RESOLVED: reviewer can inspect the located PDF/page/table evidence before deciding.",
        "- PARTIAL: prefer `STRICT_NEEDS_SOURCE_CHECK` unless the missing locator can be independently recovered.",
        "- UNRESOLVED: keep the row conservative and do not claim strict human confirmation.",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    return "\n".join(lines)


def evidence_gap_report_markdown(summary: Dict[str, Any], unresolved_items: list[Dict[str, Any]]) -> str:
    lines = [
        "# 343I2 Evidence Gap Report",
        "",
        "## 中文摘要",
        f"- evidence_resolved_count: {summary.get('evidence_resolved_count', 0)}",
        f"- evidence_partial_count: {summary.get('evidence_partial_count', 0)}",
        f"- evidence_unresolved_count: {summary.get('evidence_unresolved_count', 0)}",
        "- 未解析出的证据不会伪造，会保留为显式 gap。",
        "",
        "## English Summary",
        f"- Resolved evidence rows: {summary.get('evidence_resolved_count', 0)}",
        f"- Partial evidence rows: {summary.get('evidence_partial_count', 0)}",
        f"- Unresolved evidence rows: {summary.get('evidence_unresolved_count', 0)}",
    ]
    if not unresolved_items:
        lines.extend(["", "- No fully unresolved evidence rows remain in the current enriched package."])
        return "\n".join(lines)
    lines.extend(["", "## Unresolved Items"])
    for item in unresolved_items:
        lines.append(
            f"- {item.get('queue_item_id', '')} / {item.get('review_item_id', '')}: {item.get('evidence_gap_reason', '')}"
        )
    return "\n".join(lines)
