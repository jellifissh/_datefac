from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from datefac.review_queue.real_excel_review_343c import EDITABLE_REVIEWER_COLUMNS, REVIEW_QUEUE_REAL_EXCEL_SHEETS


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "value",
        "tags",
        "note",
        "errors",
        "recommendation",
    }
    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if str(cell.value or "") in EDITABLE_REVIEWER_COLUMNS:
                    cell.fill = editable_fill
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 343C Real Excel Review Queue Pilot",
        "",
        "## 中文摘要",
        "- 343C 只生成一个真实可填写的 Excel review 包，用于后续人工审核准备。",
        "- 343C 不导入 reviewed result，不代表人工审核已经完成，也不实现 Argilla 或正式 UI。",
        "- 343C 仍然禁止 formal client export；`client_ready=false`、`production_ready=false` 保持不变。",
        "",
        "## English Summary",
        "- 343C prepares a real fillable Excel review package for later human review.",
        "- It does not ingest reviewed results yet and does not claim completed human review.",
        "- It is not Argilla integration, not a formal client export, and not a write-back stage.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- waiting_for_human_review: {summary.get('waiting_for_human_review', False)}",
        f"- reviewed_result_ingested: {summary.get('reviewed_result_ingested', False)}",
        f"- ready_for_343d: {summary.get('ready_for_343d', False)}",
        f"- recommended_343d_scope: {summary.get('recommended_343d_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Pilot Review Counts",
        f"- real_review_template_row_count: {summary.get('real_review_template_row_count', 0)}",
        f"- fillable_review_row_count: {summary.get('fillable_review_row_count', 0)}",
        f"- human_reviewed_audit_row_count: {summary.get('human_reviewed_audit_row_count', 0)}",
        f"- simulated_direct_review_row_count: {summary.get('simulated_direct_review_row_count', 0)}",
        f"- simulated_corrected_review_row_count: {summary.get('simulated_corrected_review_row_count', 0)}",
        f"- summary_derived_review_row_count: {summary.get('summary_derived_review_row_count', 0)}",
        f"- allowed_decision_count: {summary.get('allowed_decision_count', 0)}",
        "",
        "## Safety Boundary",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        "",
        "## Reviewer Guidance",
        "- Fill only the reviewer columns highlighted in the workbook.",
        "- Use `CORRECT` only when all corrected metric/year/value/unit fields can be filled confidently.",
        "- Treat simulated rows with extra caution; they are not real human-reviewed evidence yet.",
        "- Save the filled workbook for 343D ingestion later.",
        "",
        "## Why formal client export remains forbidden",
        "- 343C prepares a waiting-for-review workbook only.",
        "- No reviewed result is ingested yet.",
        "- This stage remains sidecar-only and non-production.",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Recommended Open Order",
            "- First open `review_queue_real_excel_review_343c_review_template.xlsx` and fill `04_FILLABLE_REVIEW`.",
            "- Then read `review_queue_real_excel_review_343c_reviewer_instructions.md` and `review_queue_real_excel_review_343c_fill_guide.md`.",
            "- Keep `review_queue_real_excel_review_343c.xlsx` as the audit/reference workbook.",
            "",
            "## Next Recommendation",
            "- Next action is user human review on the fillable workbook.",
            "- After a real filled workbook exists, the next task is `343D real Excel review result ingestion`.",
            "",
        ]
    )
    return "\n".join(lines)


def reviewer_instructions_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# 343C Reviewer Instructions",
            "",
            "## 中文说明",
            "- 这是一个真实可填写的 review workbook，用来准备后续人工审核结果导入。",
            "- 当前并没有完成人工审核；343C 只是把待审样本、规则、字段说明整理成可填写模板。",
            "- 只能填写 reviewer 列：`reviewer_decision`、`reviewer_metric_standardized`、`reviewer_year_standardized`、`reviewer_value_numeric`、`reviewer_normalized_unit`、`reviewer_note`、`reviewer_id`、`reviewed_at`。",
            "- 如果选择 `CORRECT`，必须同时填写所有 corrected 字段。",
            "- `SIMULATED_*` 行必须谨慎处理，它们不是人审证据，只是候选数据。",
            "- 该模板不是正式 client export，`formal_client_export_allowed=false` 仍然成立。",
            "",
            "## English Notes",
            "- This workbook is a real fillable review package, not a completed review result.",
            "- Fill only the reviewer columns and keep identity/evidence columns unchanged.",
            "- `CORRECT` requires corrected metric/year/value/unit fields.",
            f"- Current decision: {summary.get('decision', '')}",
            f"- Waiting state: {summary.get('waiting_for_human_review', False)}",
        ]
    )


def fill_guide_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# 343C Fill Guide",
            "",
            "1. 打开 `review_queue_real_excel_review_343c_review_template.xlsx`。",
            "2. 在 `04_FILLABLE_REVIEW` 中逐行查看证据列、风险列、来源列。",
            "3. 只填写高亮 reviewer 列。",
            "4. `CONFIRM` 表示保留候选值；`CORRECT` 表示必须写入修正后的 metric/year/value/unit；`REJECT` 表示丢弃；`NEEDS_SOURCE_CHECK` 表示需要回看来源；`SKIP` 表示本轮暂不处理。",
            "5. 填完后保存一份人工填写版本，供 343D 后续 ingestion 使用。",
            "",
            "## Important boundaries",
            "- Do not change `queue_item_id`, `review_item_id`, or other identity columns.",
            "- Do not treat this workbook as a client delivery file.",
            f"- ready_for_343d remains: {summary.get('ready_for_343d', False)}",
            f"- recommended_343d_scope: {summary.get('recommended_343d_scope', '') or 'wait until workbook is filled'}",
        ]
    )


WORKBOOK_SHEETS = REVIEW_QUEUE_REAL_EXCEL_SHEETS
