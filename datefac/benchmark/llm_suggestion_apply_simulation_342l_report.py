from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_SIM_SUMMARY",
    "02_INPUT_342K_SUMMARY",
    "03_AUTO_CANDIDATES",
    "04_SPOT_CHECK_SAMPLE",
    "05_PREFILL_REVIEW_DRAFT",
    "06_HUMAN_REQUIRED",
    "07_CONFLICT_BLOCKERS",
    "08_REDUCTION_SIMULATION",
    "09_RISK_AUDIT",
    "10_PROMPT_REQUEST_TRACE",
    "11_DECISION_POLICY",
    "12_342M_READINESS",
    "13_NO_WRITE_BACK",
    "14_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "note",
        "warning",
        "reference",
        "hash",
        "risk",
        "path",
        "decision",
        "action",
        "recommendation",
        "html",
        "snippet",
        "bbox",
        "prompt",
        "json",
        "schema",
        "evidence",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 180)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 220)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 342L LLM Suggestion Apply Or Human Spot-Check Simulation",
        "",
        "中文：",
        "342L 是 suggestion apply simulation，不是真实 LLM apply，也不是人工审核完成结果。",
        "342L 只基于 342K 的 rule baseline、dry-run suggestions、auto-confirm candidates 与 human-required rows 生成模拟结果。",
        "auto-confirm candidates 仍然只是 candidate，必须经过人工 spot-check 才能进入更激进的后续自动化策略。",
        "",
        "English:",
        "342L is a suggestion-apply simulation, not a real LLM apply and not a completed human-review result.",
        "It only converts 342K simulation artifacts into a controlled spot-check and prefill package.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- ready_for_342m: {summary.get('ready_for_342m', False)}",
        f"- recommended_342m_scope: {summary.get('recommended_342m_scope', '')}",
        "",
        "## Counts",
        f"- pending_review_count: {summary.get('pending_review_count', 0)}",
        f"- auto_confirm_candidate_count: {summary.get('auto_confirm_candidate_count', 0)}",
        f"- spot_check_sample_count: {summary.get('spot_check_sample_count', 0)}",
        f"- human_required_count: {summary.get('human_required_count', 0)}",
        f"- conflict_count: {summary.get('conflict_count', 0)}",
        f"- prefill_review_draft_count: {summary.get('prefill_review_draft_count', 0)}",
        f"- prompt_pack_count: {summary.get('prompt_pack_count', 0)}",
        f"- request_pack_count: {summary.get('request_pack_count', 0)}",
        f"- jsonl_parse_error_count: {summary.get('jsonl_parse_error_count', 0)}",
        "",
        "## Reduction Simulation",
        f"- theoretical_review_reduction_count: {summary.get('theoretical_review_reduction_count', 0)}",
        f"- risk_adjusted_reduction_count: {summary.get('risk_adjusted_reduction_count', 0)}",
        f"- required_human_review_after_strategy: {summary.get('required_human_review_after_strategy', 0)}",
        f"- reduction_rate: {summary.get('reduction_rate', 0)}",
        f"- conservative_reduction_rate: {summary.get('conservative_reduction_rate', 0)}",
        "",
        "## Safety",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Boundaries",
            "- 342L does not rerun MinerU.",
            "- 342L does not call VLM.",
            "- 342L does not call a real LLM API by default.",
            "- 342L does not modify production pipeline / parser / extraction / delivery.",
            "- 342L remains a no-write-back simulation-only sidecar.",
            "",
        ]
    )
    return "\n".join(lines)
