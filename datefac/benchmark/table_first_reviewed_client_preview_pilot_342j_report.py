from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_PREVIEW_SUMMARY",
    "02_INPUT_342I_SUMMARY",
    "03_REVIEWED_PREVIEW",
    "04_CONFIRMED_PREVIEW",
    "05_CORRECTED_PREVIEW",
    "06_METRIC_YEAR_MATRIX",
    "07_BEFORE_AFTER",
    "08_SOURCE_TRACE",
    "09_REMAINING_REVIEW",
    "10_DEMO_NOTES",
    "11_LIMITATIONS",
    "12_342K_READINESS",
    "13_NO_WRITE_BACK",
    "14_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "note",
        "warning",
        "reference",
        "hash",
        "risk",
        "path",
        "decision",
        "action",
        "recommendation",
        "html",
        "snippet",
        "bbox",
        "limit",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 180)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 220)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 342J Table-First Reviewed Client Preview Pilot",
        "",
        "中文：",
        "342J 只把当前已经通过人工确认或人工修正的 table-first cells 整理成 reviewed client preview pilot。",
        f"本次基于已处理的 {summary.get('reviewed_row_count', 0)} 条 review rows，其中 {summary.get('reviewed_preview_row_count', 0)} 条进入 preview。",
        f"仍有 {summary.get('pending_review_count', 0)} 条 pending review，因此它不是全量人审完成，也不是正式客户交付。",
        "当前必须继续保持 client_ready = false、production_ready = false，并明确 not investment advice。",
        "",
        "English:",
        "342J packages the currently human-confirmed and human-corrected table-first rows into a reviewed client preview pilot.",
        f"It is based on {summary.get('reviewed_row_count', 0)} reviewed rows, with {summary.get('reviewed_preview_row_count', 0)} rows entering the preview.",
        f"{summary.get('pending_review_count', 0)} rows still remain pending review, so this is not full human-review completion and not formal client delivery.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- ready_for_342k: {summary.get('ready_for_342k', False)}",
        f"- recommended_342k_scope: {summary.get('recommended_342k_scope', '')}",
        "",
        "## Counts",
        f"- input_review_template_row_count: {summary.get('input_review_template_row_count', 0)}",
        f"- reviewed_row_count: {summary.get('reviewed_row_count', 0)}",
        f"- pending_review_count: {summary.get('pending_review_count', 0)}",
        f"- reviewed_preview_row_count: {summary.get('reviewed_preview_row_count', 0)}",
        f"- confirmed_preview_row_count: {summary.get('confirmed_preview_row_count', 0)}",
        f"- corrected_preview_row_count: {summary.get('corrected_preview_row_count', 0)}",
        f"- rejected_in_batch_count: {summary.get('rejected_in_batch_count', 0)}",
        f"- metric_covered_count: {summary.get('metric_covered_count', 0)}",
        f"- metric_year_pair_count: {summary.get('metric_year_pair_count', 0)}",
        f"- remaining_review_count: {summary.get('remaining_review_count', 0)}",
        f"- unit_year_remaining_count: {summary.get('unit_year_remaining_count', 0)}",
        f"- duplicate_remaining_count: {summary.get('duplicate_remaining_count', 0)}",
        f"- growth_row_remaining_count: {summary.get('growth_row_remaining_count', 0)}",
        "",
        "## Safety",
        f"- source_trace_missing_count: {summary.get('source_trace_missing_count', 0)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- output_workbook_path: {summary.get('output_workbook_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Boundaries",
            "- 342J does not rerun MinerU.",
            "- 342J does not call VLM/LLM.",
            "- 342J does not modify production pipeline / parser / extraction / delivery.",
            "- 342J remains a no-write-back reviewed preview pilot.",
            "- Next step can be 342K LLM-assisted adjudication or reviewed preview polish, but still within pilot scope.",
            "",
        ]
    )
    return "\n".join(lines)
