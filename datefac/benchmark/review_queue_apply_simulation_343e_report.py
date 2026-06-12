from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "value",
        "tags",
        "note",
        "errors",
        "recommendation",
        "risk",
        "boundary",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 343E AI-assisted Review Result Apply Simulation And Audit Gate",
        "",
        "## 中文摘要",
        "- 343E 只做 simulation-only apply，不做真实写回、不做生产应用、不做 formal client export。",
        "- 输入仍然是 `AI_ASSISTED_REVIEW`，所以即便是 simulated apply rows，也必须保留 human spot-check 边界。",
        "- 当前结果只允许进入后续 spot-check package，不允许进入 production export。",
        "",
        "## English Summary",
        "- 343E performs simulation-only application planning on AI-assisted reviewed results.",
        "- No real apply, no formal client export, and no production application is performed.",
        "- Human spot-check remains required even for simulation-eligible rows.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- apply_mode: {summary.get('apply_mode', '')}",
        f"- apply_simulation_completed: {summary.get('apply_simulation_completed', False)}",
        f"- audit_gate_passed_for_spot_check_package: {summary.get('audit_gate_passed_for_spot_check_package', False)}",
        f"- ready_for_343f: {summary.get('ready_for_343f', False)}",
        f"- recommended_343f_scope: {summary.get('recommended_343f_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Simulation Metrics",
        f"- input_reviewed_result_row_count: {summary.get('input_reviewed_result_row_count', 0)}",
        f"- apply_plan_row_count: {summary.get('apply_plan_row_count', 0)}",
        f"- simulated_sidecar_row_count: {summary.get('simulated_sidecar_row_count', 0)}",
        f"- hold_row_count: {summary.get('hold_row_count', 0)}",
        f"- simulate_confirm_apply_count: {summary.get('simulate_confirm_apply_count', 0)}",
        f"- simulate_correction_apply_count: {summary.get('simulate_correction_apply_count', 0)}",
        f"- hold_rejected_count: {summary.get('hold_rejected_count', 0)}",
        f"- hold_source_check_required_count: {summary.get('hold_source_check_required_count', 0)}",
        f"- hold_skipped_count: {summary.get('hold_skipped_count', 0)}",
        "",
        "## AI-assisted Boundary",
        f"- review_source_type: {summary.get('review_source_type', '')}",
        f"- not_pure_human_review: {summary.get('not_pure_human_review', False)}",
        f"- strict_human_review_completed: {summary.get('strict_human_review_completed', False)}",
        f"- requires_human_spot_check: {summary.get('requires_human_spot_check', False)}",
        "",
        "## Safety Boundary",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        "",
        "## Why export remains forbidden",
        "- The source review is AI-assisted, not strict pure human review.",
        "- 19 rows still require source check and 1 row remains skipped.",
        "- Simulation-only results are not formal client delivery evidence.",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    return "\n".join(lines)


def ai_assisted_boundary_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# 343E AI-assisted Boundary",
            "",
            "- review_source_type: AI_ASSISTED_REVIEW",
            "- not_pure_human_review: true",
            "- strict_human_review_completed: false",
            "- requires_human_spot_check: true",
            "- apply_mode: SIMULATION_ONLY",
            "- formal_client_export_allowed: false",
            "- client_ready: false",
            "- production_ready: false",
            "",
            "This stage only simulates what reviewed rows would do downstream.",
            "No upstream workbook write-back and no production application is allowed.",
            "",
            f"Current decision: {summary.get('decision', '')}",
        ]
    )
