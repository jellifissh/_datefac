from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_SCHEMA_SUMMARY",
    "02_INPUT_342S_SUMMARY",
    "03_QUEUE_FIELDS",
    "04_STATUS_LIFECYCLE",
    "05_REASON_CODES",
    "06_PRIORITY_RULES",
    "07_TRUST_MAPPING",
    "08_SAMPLE_QUEUE_ITEMS",
    "09_EXCEL_TEMPLATE",
    "10_ARGILLA_MAPPING",
    "11_UI_CONTRACT",
    "12_BACKLOG_STRATEGY",
    "13_343B_READINESS",
    "14_NO_WRITE_BACK",
    "15_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    lines = [json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows]
    path.write_text("\n".join(lines), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "example",
        "allowed",
        "validation",
        "requirement",
        "section",
        "value",
        "tags",
        "note",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 343A Review Queue Schema And Human Review UI Pilot",
        "",
        "## 中文摘要",
        "- 343A 定义了一个 durable Review Queue schema，用来承接未来的人审、spot-check、高风险样本处理、Excel round-trip、Argilla pilot 和后续自定义 UI。",
        "- 343A 只做 schema 和 pilot package，不实现完整 Argilla 集成，不实现完整前端 UI，不改生产 pipeline/parser/extraction/delivery。",
        "- Review Queue 才是稳定核心，Argilla 只是可插拔的人审界面候选之一。",
        "- 343A 把 342R/342S 的 export candidate package 映射成可审阅的 queue items，同时继续保持 formal/client/production readiness 全部为 false。",
        "",
        "## English Summary",
        "- 343A defines a durable Review Queue schema and a pilot package.",
        "- It does not implement a full Argilla integration or a full UI.",
        "- Review Queue is the stable system contract; Argilla is only a pluggable interface candidate.",
        "- Formal client export remains forbidden and no upstream workbook is written back.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- ready_for_343b: {summary.get('ready_for_343b', False)}",
        f"- recommended_343b_scope: {summary.get('recommended_343b_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Schema Metrics",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- field_count: {summary.get('field_count', 0)}",
        f"- required_field_count: {summary.get('required_field_count', 0)}",
        f"- status_count: {summary.get('status_count', 0)}",
        f"- reason_code_count: {summary.get('reason_code_count', 0)}",
        f"- priority_level_count: {summary.get('priority_level_count', 0)}",
        "",
        "## Sample Queue",
        f"- sample_queue_item_count: {summary.get('sample_queue_item_count', 0)}",
        f"- human_reviewed_sample_count: {summary.get('human_reviewed_sample_count', 0)}",
        f"- simulated_sample_count: {summary.get('simulated_sample_count', 0)}",
        f"- summary_derived_sample_count: {summary.get('summary_derived_sample_count', 0)}",
        "",
        "## Boundaries",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        "",
        "## Why formal client export stays forbidden",
        "- 342R/342S still carry HIGH export risk, later-audit requirements, and an unresolved larger review backlog.",
        "- Simulated rows are visible for demo/audit scope only and cannot be treated as final confirmed export rows.",
        "- 343A only prepares the queue contract for future review tools such as 343B Argilla pilot or an Excel round-trip pilot.",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Recommended Open Order",
            "- First open the 343A workbook and inspect `01_SCHEMA_SUMMARY`, `03_QUEUE_FIELDS`, `08_SAMPLE_QUEUE_ITEMS`, and `13_343B_READINESS`.",
            "- Then open the JSON schema and UI contract artifacts for downstream implementation handoff.",
            "",
            "## Next Recommendation",
            "- Preferred next step: `343B Argilla Human Review UI Pilot`.",
            "- Safe alternative: `343B Excel Round-trip Review Queue Pilot` if UI tooling should remain lighter.",
            "",
        ]
    )
    return "\n".join(lines)


def ui_contract_markdown(contract: Dict[str, Any]) -> str:
    lines = [
        "# 343A Review Queue UI Contract",
        "",
        "## Positioning",
        f"- {contract.get('positioning', '')}",
        "",
        "## List View",
    ]
    for item in contract.get("list_view_columns", []):
        lines.append(f"- {item}")
    lines.extend(["", "## Detail View"])
    for item in contract.get("detail_view_sections", []):
        lines.append(f"- {item}")
    lines.extend(["", "## Review Actions"])
    for item in contract.get("review_actions", []):
        lines.append(f"- {item}")
    lines.extend(["", "## Validation Rules"])
    for item in contract.get("validation_rules", []):
        lines.append(f"- {item}")
    lines.extend(["", "## Export Contract"])
    for item in contract.get("export_contract", []):
        lines.append(f"- {item}")
    lines.extend(["", "## Audit Log Expectations"])
    for item in contract.get("audit_log_expectations", []):
        lines.append(f"- {item}")
    lines.append("")
    return "\n".join(lines)
