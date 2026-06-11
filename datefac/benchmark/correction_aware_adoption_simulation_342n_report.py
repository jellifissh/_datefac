from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_ADOPTION_SUMMARY",
    "02_INPUT_342M_SUMMARY",
    "03_SPOT_CHECK_PATTERNS",
    "04_ADOPTION_INPUT",
    "05_DIRECT_ADOPT_SIM",
    "06_CORRECTION_ADOPT_SIM",
    "07_STILL_HUMAN_REQUIRED",
    "08_PATTERN_APPLICATION",
    "09_RISK_REVIEW",
    "10_BEFORE_AFTER_SIM",
    "11_REDUCTION_SIM",
    "12_342O_READINESS",
    "13_NO_WRITE_BACK",
    "14_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "note",
        "warning",
        "reason",
        "risk",
        "path",
        "html",
        "snippet",
        "bbox",
        "recommendation",
        "value",
        "evidence",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 180)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 342N Correction-Aware Adoption Simulation",
        "",
        "## Chinese Summary",
        "- 342N is a correction-aware adoption simulation, not final adoption.",
        "- 33 corrected rows out of 50 spot-check rows mean raw bulk adoption is unsafe.",
        "- 342N only simulates safe direct rows and explicit correction patterns.",
        "",
        "## English Summary",
        "- 342N is a gated simulation layer only.",
        "- It does not create final confirmed results and does not write back upstream workbooks.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- ready_for_342o: {summary.get('ready_for_342o', False)}",
        f"- recommended_342o_scope: {summary.get('recommended_342o_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Key Counts",
        f"- pending_review_count: {summary.get('pending_review_count', 0)}",
        f"- input_adoption_candidate_count: {summary.get('input_adoption_candidate_count', 0)}",
        f"- spot_check_sample_count: {summary.get('spot_check_sample_count', 0)}",
        f"- spot_check_confirm_count: {summary.get('spot_check_confirm_count', 0)}",
        f"- spot_check_correct_count: {summary.get('spot_check_correct_count', 0)}",
        f"- direct_adopt_sim_count: {summary.get('direct_adopt_sim_count', 0)}",
        f"- correction_adopt_sim_count: {summary.get('correction_adopt_sim_count', 0)}",
        f"- still_human_required_count: {summary.get('still_human_required_count', 0)}",
        f"- adoption_sim_total_count: {summary.get('adoption_sim_total_count', 0)}",
        "",
        "## Reduction",
        f"- risk_adjusted_reduction_count: {summary.get('risk_adjusted_reduction_count', 0)}",
        f"- required_human_review_after_342n: {summary.get('required_human_review_after_342n', 0)}",
        f"- conservative_reduction_rate_after_342n: {summary.get('conservative_reduction_rate_after_342n', 0)}",
        "",
        "## Safety",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Boundaries",
            "- 342N does not rerun MinerU.",
            "- 342N does not call a real LLM API by default.",
            "- 342N does not modify production pipeline / parser / extraction / delivery.",
            "- 342N does not write back to 342G / 342H / 342I / 342J / 342K / 342L / 342M workbooks.",
            "- client_ready remains false and production_ready remains false.",
            "",
        ]
    )
    return "\n".join(lines)
