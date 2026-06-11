from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_PACKAGE_SUMMARY",
    "02_INPUT_342Q_SUMMARY",
    "03_EXPORT_CANDIDATES",
    "04_HUMAN_REVIEWED",
    "05_SIMULATED_DIRECT",
    "06_SIMULATED_CORRECTED",
    "07_AUDIT_LABELS",
    "08_REQUIRED_WARNINGS",
    "09_RISK_DISCLOSURE",
    "10_COLLISION_CONTEXT",
    "11_BACKLOG_CONTEXT",
    "12_342S_READINESS",
    "13_NO_WRITE_BACK",
    "14_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "warning",
        "path",
        "evidence",
        "recommendation",
        "value",
        "disclaimer",
        "note",
        "risk",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 180)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 220)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 342R Audit-Labeled Export Candidate Package",
        "",
        "## 中文摘要",
        "- 342R 把 342Q 已放行的 export candidate scope 打包成一个带审计标签、带风险披露、带 trust-level 区分的候选包。",
        "- 342R 不是正式 client export，不是最终人审完成，也不是 production-ready output。",
        "- 包内 simulated rows 仍然需要 later audit，且必须保持 `formal_client_export_allowed=false`、`client_ready=false`、`production_ready=false`。",
        "",
        "## English Summary",
        "- 342R packages the 342Q-approved export candidate scope into an audit-labeled candidate package.",
        "- It is not a formal client export and keeps simulation boundary warnings plus later-audit requirements.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- ready_for_342s: {summary.get('ready_for_342s', False)}",
        f"- recommended_342s_scope: {summary.get('recommended_342s_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Key Counts",
        f"- export_candidate_package_row_count: {summary.get('export_candidate_package_row_count', 0)}",
        f"- human_reviewed_candidate_count: {summary.get('human_reviewed_candidate_count', 0)}",
        f"- simulated_candidate_count: {summary.get('simulated_candidate_count', 0)}",
        f"- simulated_direct_candidate_count: {summary.get('simulated_direct_candidate_count', 0)}",
        f"- simulated_corrected_candidate_count: {summary.get('simulated_corrected_candidate_count', 0)}",
        f"- disclaimer_required_count: {summary.get('disclaimer_required_count', 0)}",
        f"- later_audit_required_count: {summary.get('later_audit_required_count', 0)}",
        f"- package_row_fail_count: {summary.get('package_row_fail_count', 0)}",
        "",
        "## Risk And Backlog",
        f"- export_risk_level: {summary.get('export_risk_level', '')}",
        f"- collision_logged_count: {summary.get('collision_logged_count', 0)}",
        f"- duplicate_metric_year_source_count: {summary.get('duplicate_metric_year_source_count', 0)}",
        f"- severe_collision_count: {summary.get('severe_collision_count', 0)}",
        f"- human_over_simulation_override_count: {summary.get('human_over_simulation_override_count', 0)}",
        f"- simulated_duplicate_dropped_count: {summary.get('simulated_duplicate_dropped_count', 0)}",
        f"- still_human_required_count: {summary.get('still_human_required_count', 0)}",
        f"- remaining_review_count: {summary.get('remaining_review_count', 0)}",
        "",
        "## Safety",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- export_candidate_scope_allowed: {summary.get('export_candidate_scope_allowed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        f"- output_workbook_path: {summary.get('output_workbook_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Boundaries",
            "- 342R does not rerun MinerU or rebuild the candidate scope from old upstream logic.",
            "- 342R does not call VLM or a real LLM API.",
            "- 342R does not write back to upstream workbooks.",
            "- 342R must not be used as formal client delivery or investment advice.",
            "",
        ]
    )
    return "\n".join(lines)
