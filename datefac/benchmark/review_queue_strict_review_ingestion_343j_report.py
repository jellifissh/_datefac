from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from datefac.review_queue.strict_human_review_package_343i import (
    EDITABLE_STRICT_REVIEW_COLUMNS,
)


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows),
        encoding="utf-8",
    )


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "value",
        "note",
        "reason",
        "recommendation",
        "evidence",
        "bbox",
        "html",
        "text",
        "disclosure",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                if str(cell.value or "") in EDITABLE_STRICT_REVIEW_COLUMNS:
                    cell.fill = editable_fill
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 343J Strict Review Result Ingestion From Enriched Workbook",
        "",
        "## 中文摘要",
        "- 343J 已读取并校验 343I2 enriched strict review workbook 的填写结果。",
        "- 当前输入 workbook 属于 `AI_ASSISTED_EVIDENCE_CHECK`，不是纯人工严格复核。",
        "- 即使当前 10 行都是 `STRICT_CONFIRM`，也不能据此声明 pure strict human review completed。",
        "- formal client export 仍然禁止，下一步应进入 343K pure human confirmation attestation。",
        "",
        "## English Summary",
        "- 343J ingests the filled enriched strict review workbook and validates the strict-review decisions.",
        "- The current workbook remains AI-assisted evidence-check input, not pure human strict review.",
        "- Even if all rows are `STRICT_CONFIRM`, pure human confirmation is still required.",
        "- Formal client export remains forbidden and the next task is 343K.",
        "",
        "## Key Metrics",
        f"- decision: {summary.get('decision', '')}",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- filled_workbook_path: {summary.get('filled_workbook_path', '')}",
        f"- filled_row_count: {summary.get('filled_row_count', 0)}",
        f"- valid_row_count: {summary.get('valid_row_count', 0)}",
        f"- invalid_row_count: {summary.get('invalid_row_count', 0)}",
        f"- strict_confirm_count: {summary.get('strict_confirm_count', 0)}",
        f"- strict_correct_count: {summary.get('strict_correct_count', 0)}",
        f"- strict_reject_count: {summary.get('strict_reject_count', 0)}",
        f"- strict_needs_source_check_count: {summary.get('strict_needs_source_check_count', 0)}",
        f"- strict_defer_count: {summary.get('strict_defer_count', 0)}",
        f"- strict_review_input_source_type: {summary.get('strict_review_input_source_type', '')}",
        f"- not_pure_human_review: {summary.get('not_pure_human_review', False)}",
        f"- ai_assisted_strict_review_confirm_count: {summary.get('ai_assisted_strict_review_confirm_count', 0)}",
        f"- pure_strict_human_confirm_count: {summary.get('pure_strict_human_confirm_count', 0)}",
        f"- strict_review_result_ingested: {summary.get('strict_review_result_ingested', False)}",
        f"- pure_strict_human_review_completed: {summary.get('pure_strict_human_review_completed', False)}",
        f"- strict_human_review_completed: {summary.get('strict_human_review_completed', False)}",
        f"- requires_pure_human_confirmation: {summary.get('requires_pure_human_confirmation', False)}",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- ready_for_343k: {summary.get('ready_for_343k', False)}",
        f"- recommended_343k_scope: {summary.get('recommended_343k_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(
            f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})"
        )
    return "\n".join(lines)


def reviewer_source_disclosure_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# 343J Reviewer-Source Disclosure",
            "",
            "## 中文说明",
            "- 当前 filled workbook 来自 AI-assisted evidence check，不是纯人工严格复核。",
            "- `strict_review_input_source_type = AI_ASSISTED_EVIDENCE_CHECK`。",
            "- `not_pure_human_review = true`。",
            "- `pure_strict_human_review_completed = false`。",
            "- `strict_human_review_completed = false`。",
            "- `requires_pure_human_confirmation = true`。",
            "- formal client export 仍然禁止。",
            "",
            "## English Note",
            "- This workbook was filled through AI-assisted evidence checking, not pure human strict review.",
            "- Pure human confirmation is still required before any stronger trust claim.",
            "",
            "## Current Gate",
            f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
            f"- client_ready: {summary.get('client_ready', False)}",
            f"- production_ready: {summary.get('production_ready', False)}",
            f"- ready_for_343k: {summary.get('ready_for_343k', False)}",
            f"- recommended_343k_scope: {summary.get('recommended_343k_scope', '')}",
        ]
    )
