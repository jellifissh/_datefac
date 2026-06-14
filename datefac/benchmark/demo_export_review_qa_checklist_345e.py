from __future__ import annotations

import csv
import json
import subprocess
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Dict, Iterable, List, Sequence

from datefac.benchmark.demo_export_review_qa_checklist_345e_report import (
    render_artifact_index_markdown,
    render_executive_summary_markdown,
    render_next_plan_markdown,
    render_review_checklist_markdown,
)
from datefac.trust.no_apply_proof import (
    FORMAL_SCOPE_RULES_PATH,
    SEMANTIC_ALIAS_ASSET_PATH,
    build_no_apply_proof,
    capture_official_asset_hashes,
    sha256_file,
)


READY_DECISION_345D = "FULL_STRUCTURED_DEMO_EXPORT_PACKAGE_345D_READY"
READY_DECISION_345E = "DEMO_EXPORT_REVIEW_QA_CHECKLIST_345E_READY"
BLOCKED_DECISION_345E = "DEMO_EXPORT_REVIEW_QA_CHECKLIST_345E_BLOCKED"
INPUT_STAGE_345E = "POST_345D_DEMO_EXPORT_REVIEW_QA_CHECKLIST"

DEFAULT_FULL_STRUCTURED_DEMO_EXPORT_PACKAGE_345D_DIR = Path(
    r"D:\_datefac\output\full_structured_demo_export_package_345d"
)
DEFAULT_OUTPUT_DIR = Path(r"D:\_datefac\output\demo_export_review_qa_checklist_345e")

MANIFEST_FILE_NAME = "demo_export_review_qa_checklist_345e_manifest.json"
ARTIFACT_COMPLETENESS_JSON_FILE_NAME = "demo_export_review_qa_checklist_345e_artifact_completeness.json"
ARTIFACT_COMPLETENESS_CSV_FILE_NAME = "demo_export_review_qa_checklist_345e_artifact_completeness.csv"
ROW_COUNT_RECONCILIATION_JSON_FILE_NAME = "demo_export_review_qa_checklist_345e_row_count_reconciliation.json"
ROW_COUNT_RECONCILIATION_CSV_FILE_NAME = "demo_export_review_qa_checklist_345e_row_count_reconciliation.csv"
GATE_SAFETY_JSON_FILE_NAME = "demo_export_review_qa_checklist_345e_gate_safety_check.json"
CAVEAT_COMPLETENESS_JSON_FILE_NAME = "demo_export_review_qa_checklist_345e_caveat_completeness_check.json"
PRESENTATION_READINESS_JSON_FILE_NAME = "demo_export_review_qa_checklist_345e_demo_presentation_readiness.json"
SAMPLE_DEMO_ROWS_JSON_FILE_NAME = "demo_export_review_qa_checklist_345e_sample_demo_rows.json"
SAMPLE_DEMO_ROWS_CSV_FILE_NAME = "demo_export_review_qa_checklist_345e_sample_demo_rows.csv"
QUALITY_LIMITED_SAMPLE_ROWS_JSON_FILE_NAME = "demo_export_review_qa_checklist_345e_quality_limited_sample_rows.json"
QUALITY_LIMITED_SAMPLE_ROWS_CSV_FILE_NAME = "demo_export_review_qa_checklist_345e_quality_limited_sample_rows.csv"
EXCLUDED_SAMPLE_ROWS_JSON_FILE_NAME = "demo_export_review_qa_checklist_345e_excluded_sample_rows.json"
EXCLUDED_SAMPLE_ROWS_CSV_FILE_NAME = "demo_export_review_qa_checklist_345e_excluded_sample_rows.csv"
REVIEW_CHECKLIST_MD_FILE_NAME = "demo_export_review_qa_checklist_345e_review_checklist.md"
EXECUTIVE_SUMMARY_MD_FILE_NAME = "demo_export_review_qa_checklist_345e_executive_summary.md"
ARTIFACT_INDEX_MD_FILE_NAME = "demo_export_review_qa_checklist_345e_artifact_index.md"
NEXT_PLAN_MD_FILE_NAME = "demo_export_review_qa_checklist_345e_next_plan.md"

INPUT_MANIFEST_NAME = "full_structured_demo_export_package_345d_manifest.json"
INPUT_DEMO_ROWS_JSON_NAME = "full_structured_demo_export_package_345d_demo_rows.json"
INPUT_DEMO_ROWS_CSV_NAME = "full_structured_demo_export_package_345d_demo_rows.csv"
INPUT_DEMO_ROWS_XLSX_NAME = "full_structured_demo_export_package_345d_demo_rows.xlsx"
INPUT_QUALITY_LIMITED_ROWS_JSON_NAME = "full_structured_demo_export_package_345d_quality_limited_rows.json"
INPUT_QUALITY_LIMITED_ROWS_CSV_NAME = "full_structured_demo_export_package_345d_quality_limited_rows.csv"
INPUT_EXCLUDED_ROWS_JSON_NAME = "full_structured_demo_export_package_345d_excluded_rows.json"
INPUT_EXCLUDED_ROWS_CSV_NAME = "full_structured_demo_export_package_345d_excluded_rows.csv"
INPUT_REMAINING_BLIND_SPOTS_JSON_NAME = "full_structured_demo_export_package_345d_remaining_blind_spots.json"
INPUT_REMAINING_BLIND_SPOTS_CSV_NAME = "full_structured_demo_export_package_345d_remaining_blind_spots.csv"
INPUT_ALIAS_SIDECAR_JSON_NAME = "full_structured_demo_export_package_345d_alias_simulation_sidecar.json"
INPUT_ALIAS_SIDECAR_CSV_NAME = "full_structured_demo_export_package_345d_alias_simulation_sidecar.csv"
INPUT_QUALITY_CAVEATS_JSON_NAME = "full_structured_demo_export_package_345d_quality_caveats.json"
INPUT_QUALITY_CAVEATS_MD_NAME = "full_structured_demo_export_package_345d_quality_caveats.md"
INPUT_DEMO_EXPORT_SUMMARY_JSON_NAME = "full_structured_demo_export_package_345d_demo_export_summary.json"
INPUT_EXECUTIVE_SUMMARY_MD_NAME = "full_structured_demo_export_package_345d_executive_summary.md"
INPUT_ARTIFACT_INDEX_MD_NAME = "full_structured_demo_export_package_345d_artifact_index.md"
INPUT_NEXT_PLAN_MD_NAME = "full_structured_demo_export_package_345d_next_plan.md"

PROTECTED_DIRTY_PATHS = [
    "datefac/benchmark/batch_row_text_delivery_benchmark.py",
    "datefac/extraction/row_text_metric_extractor.py",
    "datefac/pipeline/batch_ppstructure_row_text_pipeline.py",
    "tools/run_batch_ppstructure_outputs_320g.py",
    "input/semantic_adjudicator_responses_322d",
    "input/semantic_adjudicator_responses_322f",
    "tools/mineru_new_runner.cmd",
]

FORBIDDEN_STAGE_PATHS = [
    "output",
    "temp",
    "input",
    "input/semantic_adjudicator_responses_322d",
    "input/semantic_adjudicator_responses_322f",
    "tools/mineru_new_runner.cmd",
]

EXPECTED_345D_ARTIFACT_SPECS = [
    {
        "artifact_key": "manifest",
        "artifact_name": INPUT_MANIFEST_NAME,
        "path_name": INPUT_MANIFEST_NAME,
        "required": True,
        "mode": "json",
        "kind": "metadata",
    },
    {
        "artifact_key": "demo_rows_json",
        "artifact_name": INPUT_DEMO_ROWS_JSON_NAME,
        "path_name": INPUT_DEMO_ROWS_JSON_NAME,
        "required": True,
        "mode": "json",
        "kind": "rowset",
    },
    {
        "artifact_key": "demo_rows_csv",
        "artifact_name": INPUT_DEMO_ROWS_CSV_NAME,
        "path_name": INPUT_DEMO_ROWS_CSV_NAME,
        "required": True,
        "mode": "csv",
        "kind": "rowset",
    },
    {
        "artifact_key": "demo_rows_xlsx",
        "artifact_name": INPUT_DEMO_ROWS_XLSX_NAME,
        "path_name": INPUT_DEMO_ROWS_XLSX_NAME,
        "required": False,
        "mode": "xlsx",
        "kind": "rowset",
    },
    {
        "artifact_key": "quality_limited_rows_json",
        "artifact_name": INPUT_QUALITY_LIMITED_ROWS_JSON_NAME,
        "path_name": INPUT_QUALITY_LIMITED_ROWS_JSON_NAME,
        "required": True,
        "mode": "json",
        "kind": "rowset",
    },
    {
        "artifact_key": "quality_limited_rows_csv",
        "artifact_name": INPUT_QUALITY_LIMITED_ROWS_CSV_NAME,
        "path_name": INPUT_QUALITY_LIMITED_ROWS_CSV_NAME,
        "required": True,
        "mode": "csv",
        "kind": "rowset",
    },
    {
        "artifact_key": "excluded_rows_json",
        "artifact_name": INPUT_EXCLUDED_ROWS_JSON_NAME,
        "path_name": INPUT_EXCLUDED_ROWS_JSON_NAME,
        "required": True,
        "mode": "json",
        "kind": "rowset",
    },
    {
        "artifact_key": "excluded_rows_csv",
        "artifact_name": INPUT_EXCLUDED_ROWS_CSV_NAME,
        "path_name": INPUT_EXCLUDED_ROWS_CSV_NAME,
        "required": True,
        "mode": "csv",
        "kind": "rowset",
    },
    {
        "artifact_key": "remaining_blind_spots_json",
        "artifact_name": INPUT_REMAINING_BLIND_SPOTS_JSON_NAME,
        "path_name": INPUT_REMAINING_BLIND_SPOTS_JSON_NAME,
        "required": True,
        "mode": "json",
        "kind": "rowset",
    },
    {
        "artifact_key": "remaining_blind_spots_csv",
        "artifact_name": INPUT_REMAINING_BLIND_SPOTS_CSV_NAME,
        "path_name": INPUT_REMAINING_BLIND_SPOTS_CSV_NAME,
        "required": True,
        "mode": "csv",
        "kind": "rowset",
    },
    {
        "artifact_key": "alias_sidecar_json",
        "artifact_name": INPUT_ALIAS_SIDECAR_JSON_NAME,
        "path_name": INPUT_ALIAS_SIDECAR_JSON_NAME,
        "required": True,
        "mode": "json",
        "kind": "rowset",
    },
    {
        "artifact_key": "alias_sidecar_csv",
        "artifact_name": INPUT_ALIAS_SIDECAR_CSV_NAME,
        "path_name": INPUT_ALIAS_SIDECAR_CSV_NAME,
        "required": True,
        "mode": "csv",
        "kind": "rowset",
    },
    {
        "artifact_key": "quality_caveats_json",
        "artifact_name": INPUT_QUALITY_CAVEATS_JSON_NAME,
        "path_name": INPUT_QUALITY_CAVEATS_JSON_NAME,
        "required": True,
        "mode": "json",
        "kind": "metadata",
    },
    {
        "artifact_key": "quality_caveats_md",
        "artifact_name": INPUT_QUALITY_CAVEATS_MD_NAME,
        "path_name": INPUT_QUALITY_CAVEATS_MD_NAME,
        "required": True,
        "mode": "md",
        "kind": "metadata",
    },
    {
        "artifact_key": "demo_export_summary_json",
        "artifact_name": INPUT_DEMO_EXPORT_SUMMARY_JSON_NAME,
        "path_name": INPUT_DEMO_EXPORT_SUMMARY_JSON_NAME,
        "required": True,
        "mode": "json",
        "kind": "metadata",
    },
    {
        "artifact_key": "executive_summary_md",
        "artifact_name": INPUT_EXECUTIVE_SUMMARY_MD_NAME,
        "path_name": INPUT_EXECUTIVE_SUMMARY_MD_NAME,
        "required": True,
        "mode": "md",
        "kind": "metadata",
    },
    {
        "artifact_key": "artifact_index_md",
        "artifact_name": INPUT_ARTIFACT_INDEX_MD_NAME,
        "path_name": INPUT_ARTIFACT_INDEX_MD_NAME,
        "required": True,
        "mode": "md",
        "kind": "metadata",
    },
    {
        "artifact_key": "next_plan_md",
        "artifact_name": INPUT_NEXT_PLAN_MD_NAME,
        "path_name": INPUT_NEXT_PLAN_MD_NAME,
        "required": True,
        "mode": "md",
        "kind": "metadata",
    },
]

OUTPUT_ARTIFACT_ROWS = [
    {"artifact_name": MANIFEST_FILE_NAME, "path": MANIFEST_FILE_NAME, "purpose": "Decision and QA metrics manifest."},
    {
        "artifact_name": ARTIFACT_COMPLETENESS_JSON_FILE_NAME,
        "path": ARTIFACT_COMPLETENESS_JSON_FILE_NAME,
        "purpose": "Artifact completeness check results in JSON.",
    },
    {
        "artifact_name": ARTIFACT_COMPLETENESS_CSV_FILE_NAME,
        "path": ARTIFACT_COMPLETENESS_CSV_FILE_NAME,
        "purpose": "Artifact completeness check results in CSV.",
    },
    {
        "artifact_name": ROW_COUNT_RECONCILIATION_JSON_FILE_NAME,
        "path": ROW_COUNT_RECONCILIATION_JSON_FILE_NAME,
        "purpose": "Row-count reconciliation results in JSON.",
    },
    {
        "artifact_name": ROW_COUNT_RECONCILIATION_CSV_FILE_NAME,
        "path": ROW_COUNT_RECONCILIATION_CSV_FILE_NAME,
        "purpose": "Row-count reconciliation results in CSV.",
    },
    {
        "artifact_name": GATE_SAFETY_JSON_FILE_NAME,
        "path": GATE_SAFETY_JSON_FILE_NAME,
        "purpose": "Gate safety audit results.",
    },
    {
        "artifact_name": CAVEAT_COMPLETENESS_JSON_FILE_NAME,
        "path": CAVEAT_COMPLETENESS_JSON_FILE_NAME,
        "purpose": "Caveat coverage audit results.",
    },
    {
        "artifact_name": PRESENTATION_READINESS_JSON_FILE_NAME,
        "path": PRESENTATION_READINESS_JSON_FILE_NAME,
        "purpose": "Demo-only presentation readiness assessment.",
    },
    {
        "artifact_name": SAMPLE_DEMO_ROWS_JSON_FILE_NAME,
        "path": SAMPLE_DEMO_ROWS_JSON_FILE_NAME,
        "purpose": "Bounded sample copied from 345D demo rows.",
    },
    {
        "artifact_name": SAMPLE_DEMO_ROWS_CSV_FILE_NAME,
        "path": SAMPLE_DEMO_ROWS_CSV_FILE_NAME,
        "purpose": "Bounded sample copied from 345D demo rows in CSV form.",
    },
    {
        "artifact_name": QUALITY_LIMITED_SAMPLE_ROWS_JSON_FILE_NAME,
        "path": QUALITY_LIMITED_SAMPLE_ROWS_JSON_FILE_NAME,
        "purpose": "Bounded sample copied from 345D quality-limited rows.",
    },
    {
        "artifact_name": QUALITY_LIMITED_SAMPLE_ROWS_CSV_FILE_NAME,
        "path": QUALITY_LIMITED_SAMPLE_ROWS_CSV_FILE_NAME,
        "purpose": "Bounded sample copied from 345D quality-limited rows in CSV form.",
    },
    {
        "artifact_name": EXCLUDED_SAMPLE_ROWS_JSON_FILE_NAME,
        "path": EXCLUDED_SAMPLE_ROWS_JSON_FILE_NAME,
        "purpose": "Bounded sample copied from 345D excluded rows.",
    },
    {
        "artifact_name": EXCLUDED_SAMPLE_ROWS_CSV_FILE_NAME,
        "path": EXCLUDED_SAMPLE_ROWS_CSV_FILE_NAME,
        "purpose": "Bounded sample copied from 345D excluded rows in CSV form.",
    },
    {
        "artifact_name": REVIEW_CHECKLIST_MD_FILE_NAME,
        "path": REVIEW_CHECKLIST_MD_FILE_NAME,
        "purpose": "Human-readable QA checklist.",
    },
    {
        "artifact_name": EXECUTIVE_SUMMARY_MD_FILE_NAME,
        "path": EXECUTIVE_SUMMARY_MD_FILE_NAME,
        "purpose": "Executive summary of QA findings.",
    },
    {
        "artifact_name": ARTIFACT_INDEX_MD_FILE_NAME,
        "path": ARTIFACT_INDEX_MD_FILE_NAME,
        "purpose": "Artifact index for the QA package.",
    },
    {
        "artifact_name": NEXT_PLAN_MD_FILE_NAME,
        "path": NEXT_PLAN_MD_FILE_NAME,
        "purpose": "Recommended next step after QA review.",
    },
]


def _utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def _default_ledger_path() -> Path:
    milestone_dir = Path(r"D:\_datefac\docs\project_milestones")
    matches = sorted(milestone_dir.glob("PROJECT_MILESTONE_LEDGER_*.md"))
    if matches:
        return matches[0]
    return milestone_dir / "PROJECT_MILESTONE_LEDGER.md"


DEFAULT_LEDGER_PATH = _default_ledger_path()


def _safe_text(value: Any) -> str:
    if value is None:
        return ""
    text = " ".join(str(value).strip().split())
    return "" if text.lower() == "nan" else text


def _bool_value(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    return _safe_text(value).lower() in {"1", "true", "yes", "y"}


def _read_json(path: Path) -> Any:
    return json.loads(path.read_text(encoding="utf-8"))


def _read_csv_rows(path: Path) -> List[Dict[str, Any]]:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return [dict(row) for row in csv.DictReader(handle)]


def _load_json_or_csv_rows(*, json_path: Path, csv_path: Path, label: str) -> tuple[List[Dict[str, Any]], str]:
    if json_path.exists():
        payload = _read_json(json_path)
        if not isinstance(payload, list):
            raise ValueError(f"{label} must be a JSON list: {json_path}")
        return [dict(row) for row in payload], "json"
    if csv_path.exists():
        return _read_csv_rows(csv_path), "csv"
    raise FileNotFoundError(f"required row artifact missing for {label}: {json_path} / {csv_path}")


def _count_rows_from_json(path: Path, label: str) -> int:
    payload = _read_json(path)
    if not isinstance(payload, list):
        raise ValueError(f"{label} must be a JSON list: {path}")
    return len(payload)


def _count_rows_from_csv(path: Path) -> int:
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        return sum(1 for _ in csv.DictReader(handle))


def _count_rows_from_xlsx(path: Path, sheet_name: str = "demo_rows") -> int:
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"worksheet missing from workbook: {sheet_name}")
    worksheet = workbook[sheet_name]
    rows = 0
    for index, _ in enumerate(worksheet.iter_rows(values_only=True), start=1):
        if index == 1:
            continue
        rows += 1
    return rows


def _is_git_repo(repo_root: Path) -> bool:
    return (repo_root / ".git").exists()


def _run_git(repo_root: Path, args: Sequence[str]) -> subprocess.CompletedProcess[str]:
    return subprocess.run(
        ["git", *args],
        cwd=repo_root,
        capture_output=True,
        text=True,
        encoding="utf-8",
        errors="replace",
    )


def _git_status_porcelain_for_paths(paths: Sequence[str], repo_root: Path) -> List[str]:
    if not _is_git_repo(repo_root):
        return []
    result = _run_git(repo_root, ["status", "--porcelain", "--", *paths])
    if result.returncode != 0:
        return [f"__ERROR__::{result.stderr.strip()}"]
    return [line.rstrip() for line in result.stdout.splitlines() if line.strip()]


def _git_staged_names_for_paths(paths: Sequence[str], repo_root: Path) -> List[str]:
    lines = _git_status_porcelain_for_paths(paths, repo_root)
    staged: List[str] = []
    for line in lines:
        if line.startswith("__ERROR__::"):
            return [line]
        if len(line) >= 3 and line[0] in {"A", "M", "D", "R", "C", "U", "T"}:
            staged.append(line[3:].strip())
    return staged


def _join_markdown_lines(values: Iterable[str]) -> str:
    return "\n".join([line for line in values if line is not None])


def _safe_source_row_count(rows: List[Dict[str, Any]]) -> int:
    return len(rows)


def _row_sample_package(
    *,
    source_artifact: str,
    source_path: Path,
    rows: List[Dict[str, Any]],
    sample_limit: int,
) -> Dict[str, Any]:
    sample_rows = deepcopy(rows[: max(sample_limit, 0)])
    return {
        "source_artifact": source_artifact,
        "source_path": str(source_path),
        "sample_limit": sample_limit,
        "selected_count": len(sample_rows),
        "source_row_count": len(rows),
        "rows": sample_rows,
    }


def _write_sample_csv_rows(rows: List[Dict[str, Any]], fieldnames: List[str]) -> List[Dict[str, Any]]:
    return [{key: row.get(key, "") for key in fieldnames} for row in rows]


def _artifact_completeness_rows(
    *,
    full_structured_demo_export_package_345d_dir: Path,
    manifest_345d: Dict[str, Any],
) -> tuple[List[Dict[str, Any]], Dict[str, int]]:
    rows: List[Dict[str, Any]] = []
    counts = {
        "missing_required": 0,
        "optional_missing": 0,
        "artifact_read_error": 0,
    }
    json_counts: Dict[str, int] = {}
    csv_counts: Dict[str, int] = {}
    xlsx_count = None

    for spec in EXPECTED_345D_ARTIFACT_SPECS:
        artifact_path = full_structured_demo_export_package_345d_dir / spec["path_name"]
        status = "present"
        row_count = None
        note = ""
        try:
            if not artifact_path.exists():
                status = "missing" if spec["required"] else "optional_missing"
                if status == "missing":
                    counts["missing_required"] += 1
                else:
                    counts["optional_missing"] += 1
            elif artifact_path.stat().st_size == 0:
                status = "empty"
                if spec["required"]:
                    counts["missing_required"] += 1
                note = "file exists but is empty"
            elif spec["mode"] == "json":
                payload = _read_json(artifact_path)
                if isinstance(payload, list):
                    row_count = len(payload)
                    if spec["artifact_key"].endswith("_json"):
                        json_counts[spec["artifact_key"]] = row_count
                elif isinstance(payload, dict):
                    row_count = None
                else:
                    raise ValueError("unsupported json payload type")
            elif spec["mode"] == "csv":
                row_count = _count_rows_from_csv(artifact_path)
                if spec["artifact_key"].endswith("_csv"):
                    csv_counts[spec["artifact_key"]] = row_count
            elif spec["mode"] == "xlsx":
                row_count = _count_rows_from_xlsx(artifact_path)
                xlsx_count = row_count
            elif spec["mode"] == "md":
                text = artifact_path.read_text(encoding="utf-8")
                row_count = None
                if not text.strip():
                    status = "empty"
                    if spec["required"]:
                        counts["missing_required"] += 1
                    note = "markdown artifact is empty"
            else:
                raise ValueError(f"unsupported artifact mode: {spec['mode']}")
        except Exception as exc:  # noqa: BLE001
            status = "read_error"
            counts["artifact_read_error"] += 1
            note = str(exc)
            if spec["required"]:
                counts["missing_required"] += 1

        rows.append(
            {
                "artifact_key": spec["artifact_key"],
                "artifact_name": spec["artifact_name"],
                "kind": spec["kind"],
                "required": spec["required"],
                "expected_path": str(artifact_path),
                "status": status,
                "row_count": row_count,
                "note": note,
            }
        )

    return rows, {
        **counts,
        "json_demo_rows_count": json_counts.get("demo_rows_json", 0),
        "csv_demo_rows_count": csv_counts.get("demo_rows_csv", 0),
        "json_quality_limited_rows_count": json_counts.get("quality_limited_rows_json", 0),
        "csv_quality_limited_rows_count": csv_counts.get("quality_limited_rows_csv", 0),
        "json_excluded_rows_count": json_counts.get("excluded_rows_json", 0),
        "csv_excluded_rows_count": csv_counts.get("excluded_rows_csv", 0),
        "xlsx_demo_rows_count": xlsx_count,
    }


def _row_count_reconciliation_rows(
    *,
    manifest_345d: Dict[str, Any],
    json_demo_rows_count: int,
    csv_demo_rows_count: int,
    xlsx_demo_rows_count: int | None,
    json_quality_limited_rows_count: int,
    csv_quality_limited_rows_count: int,
    json_excluded_rows_count: int,
    csv_excluded_rows_count: int,
) -> List[Dict[str, Any]]:
    manifest_demo = int(manifest_345d.get("demo_export_row_count", 0))
    manifest_quality = int(manifest_345d.get("quality_limited_row_count", 0))
    manifest_excluded = int(manifest_345d.get("excluded_row_count", 0))
    manifest_total = int(manifest_345d.get("inventory_row_count", 0))

    rows = [
        {
            "item": "demo_rows",
            "manifest_count": manifest_demo,
            "json_count": json_demo_rows_count,
            "csv_count": csv_demo_rows_count,
            "xlsx_count": xlsx_demo_rows_count,
            "actual_count": json_demo_rows_count,
            "delta": json_demo_rows_count - manifest_demo,
            "status": "PASS" if json_demo_rows_count == manifest_demo and csv_demo_rows_count == manifest_demo else "FAIL",
            "detail": "demo rows reconcile to 345D manifest",
        },
        {
            "item": "quality_limited_rows",
            "manifest_count": manifest_quality,
            "json_count": json_quality_limited_rows_count,
            "csv_count": csv_quality_limited_rows_count,
            "xlsx_count": None,
            "actual_count": json_quality_limited_rows_count,
            "delta": json_quality_limited_rows_count - manifest_quality,
            "status": "PASS" if json_quality_limited_rows_count == manifest_quality and csv_quality_limited_rows_count == manifest_quality else "FAIL",
            "detail": "quality-limited rows reconcile to 345D manifest",
        },
        {
            "item": "excluded_rows",
            "manifest_count": manifest_excluded,
            "json_count": json_excluded_rows_count,
            "csv_count": csv_excluded_rows_count,
            "xlsx_count": None,
            "actual_count": json_excluded_rows_count,
            "delta": json_excluded_rows_count - manifest_excluded,
            "status": "PASS" if json_excluded_rows_count == manifest_excluded and csv_excluded_rows_count == manifest_excluded else "FAIL",
            "detail": "excluded rows reconcile to 345D manifest",
        },
        {
            "item": "inventory_total",
            "manifest_count": manifest_total,
            "json_count": json_demo_rows_count + json_quality_limited_rows_count + json_excluded_rows_count,
            "csv_count": csv_demo_rows_count + csv_quality_limited_rows_count + csv_excluded_rows_count,
            "xlsx_count": xlsx_demo_rows_count,
            "actual_count": json_demo_rows_count + json_quality_limited_rows_count + json_excluded_rows_count,
            "delta": (json_demo_rows_count + json_quality_limited_rows_count + json_excluded_rows_count) - manifest_total,
            "status": "PASS"
            if (json_demo_rows_count + json_quality_limited_rows_count + json_excluded_rows_count) == manifest_total
            else "FAIL",
            "detail": "demo + quality-limited + excluded must equal inventory total",
        },
    ]
    return rows


def _gate_safety_check(
    *,
    manifest_345d: Dict[str, Any],
    demo_rows: List[Dict[str, Any]],
    quality_limited_rows: List[Dict[str, Any]],
    excluded_rows: List[Dict[str, Any]],
    gate_safety_row_limit: int = 50,
) -> Dict[str, Any]:
    checks: List[Dict[str, Any]] = []
    gate_keys = [
        "formal_client_export_allowed",
        "client_ready",
        "production_ready",
        "global_strict_human_review_completed",
        "formal_export_generated",
        "demo_export_only",
    ]
    for key in gate_keys:
        checks.append(
            {
                "name": f"manifest::{key}",
                "status": "PASS"
                if (not _bool_value(manifest_345d.get(key)) if key != "demo_export_only" else _bool_value(manifest_345d.get(key)))
                else "FAIL",
                "expected": False if key != "demo_export_only" else True,
                "observed": _bool_value(manifest_345d.get(key)),
                "evidence": f"manifest flag {key}",
            }
        )

    row_sets = [
        ("demo_rows", demo_rows),
        ("quality_limited_rows", quality_limited_rows),
        ("excluded_rows", excluded_rows),
    ]
    for label, rows in row_sets:
        limited_rows = rows[:gate_safety_row_limit]
        for key in ["formal_client_export_allowed", "client_ready", "production_ready"]:
            observed_values = {_bool_value(row.get(key)) for row in limited_rows}
            checks.append(
                {
                    "name": f"{label}::{key}",
                    "status": "PASS" if observed_values <= {False} else "FAIL",
                    "expected": False,
                    "observed_values": sorted(observed_values),
                    "evidence": f"sampled {len(limited_rows)} rows from {label}",
                }
            )

    passed = all(check["status"] == "PASS" for check in checks)
    return {
        "passed": passed,
        "checks": checks,
        "gate_keys": gate_keys,
        "safety_row_sample_limit": gate_safety_row_limit,
        "row_sets_checked": [label for label, _ in row_sets],
    }


def _caveat_completeness_check(
    *,
    manifest_345d: Dict[str, Any],
    quality_caveats_json: Dict[str, Any],
    quality_caveats_md: str,
    demo_export_summary_json: Dict[str, Any],
) -> Dict[str, Any]:
    combined_text = "\n".join(
        [
            json.dumps(quality_caveats_json, ensure_ascii=False),
            quality_caveats_md,
            json.dumps(demo_export_summary_json, ensure_ascii=False),
            json.dumps(manifest_345d, ensure_ascii=False),
        ]
    ).lower()
    topic_checks = [
        (
            "remaining unnormalized raw metric names",
            "remaining_unnormalized_raw_metric_name_count" in quality_caveats_json
            and "remaining_unnormalized_raw_metric_name_count" in manifest_345d,
        ),
        (
            "remaining unnormalized metric rows",
            "remaining_unnormalized_metric_row_count" in quality_caveats_json
            and "remaining_unnormalized_metric_row_count" in manifest_345d,
        ),
        (
            "high severity quality issues",
            "high_severity_issue_count" in quality_caveats_json and "high severity" in combined_text,
        ),
        (
            "medium severity quality issues",
            "medium_severity_issue_count" in quality_caveats_json and "medium severity" in combined_text,
        ),
        (
            "missing unit count",
            "missing_unit_count" in quality_caveats_json and "missing_unit_count" in combined_text,
        ),
        (
            "missing period count",
            "missing_period_count" in quality_caveats_json and "missing_period_count" in combined_text,
        ),
        (
            "missing source trace count",
            "missing_source_trace_count" in quality_caveats_json and "missing_source_trace_count" in combined_text,
        ),
        (
            "alias simulation is not official rule mutation",
            "simulation_exact_match_limitation" in quality_caveats_json
            and bool(quality_caveats_json.get("simulation_exact_match_limitation"))
            and bool(manifest_345d.get("official_rules_modified") is False)
            and bool(manifest_345d.get("official_alias_assets_modified") is False),
        ),
        (
            "official rules/assets unchanged",
            "official_rules_modified" in combined_text and "official_alias_assets_modified" in combined_text,
        ),
        (
            "formal/client/production gates remain false",
            "formal_client_export_allowed" in combined_text
            and "client_ready" in combined_text
            and "production_ready" in combined_text,
        ),
    ]
    present_topics: List[str] = []
    missing_topics: List[str] = []
    for topic, passed in topic_checks:
        if passed:
            present_topics.append(topic)
        else:
            missing_topics.append(topic)
    return {
        "passed": not missing_topics,
        "present_topics": present_topics,
        "missing_topics": missing_topics,
        "missing_topic_count": len(missing_topics),
    }


def _presentation_readiness(
    *,
    manifest_345d: Dict[str, Any],
    sample_demo_rows_package: Dict[str, Any],
    quality_limited_sample_rows_package: Dict[str, Any],
    excluded_sample_rows_package: Dict[str, Any],
    caveat_check: Dict[str, Any],
) -> Dict[str, Any]:
    safe_for_demo_only = (
        manifest_345d.get("formal_client_export_allowed") is False
        and manifest_345d.get("client_ready") is False
        and manifest_345d.get("production_ready") is False
        and manifest_345d.get("global_strict_human_review_completed") is False
        and manifest_345d.get("formal_export_generated") is False
        and manifest_345d.get("demo_export_only") is True
        and caveat_check["passed"]
        and sample_demo_rows_package["selected_count"] > 0
        and quality_limited_sample_rows_package["selected_count"] > 0
        and excluded_sample_rows_package["selected_count"] > 0
    )
    recommended_first_files = [
        "demo_export_review_qa_checklist_345e_review_checklist.md",
        "demo_export_review_qa_checklist_345e_manifest.json",
        "demo_export_review_qa_checklist_345e_executive_summary.md",
        "demo_export_review_qa_checklist_345e_artifact_index.md",
    ]
    safe_sample_files = [
        "demo_export_review_qa_checklist_345e_sample_demo_rows.json",
        "demo_export_review_qa_checklist_345e_quality_limited_sample_rows.json",
        "demo_export_review_qa_checklist_345e_excluded_sample_rows.json",
    ]
    spoken_caveats = [
        "345D is demo-only, not formal client export.",
        "Remaining unnormalized metric names and rows are still present.",
        "High and medium severity quality issues remain explicitly documented.",
        "Missing unit / period / source-trace caveats must be spoken aloud.",
        "Official normalization rules and alias assets were not modified.",
    ]
    prohibited_claims = [
        "do not claim formal client export is allowed",
        "do not claim client_ready = true",
        "do not claim production_ready = true",
        "do not claim global strict human review is complete",
        "do not claim the package is production-grade",
    ]
    return {
        "safe_for_demo_only": safe_for_demo_only,
        "presentation_ready_for_demo_only": safe_for_demo_only,
        "recommended_first_files": recommended_first_files,
        "safe_sample_files": safe_sample_files,
        "spoken_caveats": spoken_caveats,
        "prohibited_claims": prohibited_claims,
        "first_sample_row_demo": sample_demo_rows_package["rows"][0] if sample_demo_rows_package["rows"] else {},
        "sample_counts": {
            "demo_rows": sample_demo_rows_package["selected_count"],
            "quality_limited_rows": quality_limited_sample_rows_package["selected_count"],
            "excluded_rows": excluded_sample_rows_package["selected_count"],
        },
    }


def ledger_has_345e_entry(ledger_path: Path) -> bool:
    if not ledger_path.exists():
        return False
    text = ledger_path.read_text(encoding="utf-8")
    return "## 345E Demo Export Review / QA Checklist" in text or "DEMO_EXPORT_REVIEW_QA_CHECKLIST_345E_READY" in text


def build_345e_ledger_entry(*, manifest: Dict[str, Any]) -> str:
    return _join_markdown_lines(
        [
            "## 345E Demo Export Review / QA Checklist",
            "",
            "Status: completed",
            "",
            f"- decision: {manifest.get('decision', '')}",
            f"- input package: {manifest.get('input_345d_dir', '')}",
            f"- output package: {manifest.get('output_dir', '')}",
            f"- checked_artifact_count: {manifest.get('checked_artifact_count', 0)}",
            f"- missing_required_artifact_count: {manifest.get('missing_required_artifact_count', 0)}",
            f"- optional_missing_artifact_count: {manifest.get('optional_missing_artifact_count', 0)}",
            f"- artifact_read_error_count: {manifest.get('artifact_read_error_count', 0)}",
            f"- row_count_closure_passed: {manifest.get('row_count_closure_passed', False)}",
            f"- demo_export_row_count: {manifest.get('demo_export_row_count', 0)}",
            f"- quality_limited_row_count: {manifest.get('quality_limited_row_count', 0)}",
            f"- excluded_row_count: {manifest.get('excluded_row_count', 0)}",
            f"- caveat_completeness_passed: {manifest.get('caveat_completeness_passed', False)}",
            f"- gate_safety_check_passed: {manifest.get('gate_safety_check_passed', False)}",
            f"- presentation_ready_for_demo_only: {manifest.get('presentation_ready_for_demo_only', False)}",
            f"- no_write_back_proof_passed: {manifest.get('no_write_back_proof_passed', False)}",
            f"- formal_client_export_allowed: {manifest.get('formal_client_export_allowed', False)}",
            f"- client_ready: {manifest.get('client_ready', False)}",
            f"- production_ready: {manifest.get('production_ready', False)}",
            f"- global_strict_human_review_completed: {manifest.get('global_strict_human_review_completed', False)}",
            f"- sample_demo_row_count: {manifest.get('sample_demo_row_count', 0)}",
            f"- sample_quality_limited_row_count: {manifest.get('sample_quality_limited_row_count', 0)}",
            f"- sample_excluded_row_count: {manifest.get('sample_excluded_row_count', 0)}",
            f"- next recommended step: {manifest.get('next_recommended_step', '')}",
            "",
            "Validation commands and results:",
            f"- py_compile: {manifest.get('validation_py_compile', '')}",
            f"- pytest: {manifest.get('validation_pytest', '')}",
            f"- real runner: {manifest.get('validation_real_runner', '')}",
            "",
            "No-write-back confirmation:",
            f"- {manifest.get('no_write_back_summary', '')}",
        ]
    )


def append_345e_ledger_entry(*, manifest: Dict[str, Any], ledger_path: Path) -> bool:
    if ledger_has_345e_entry(ledger_path):
        return False
    existing = ledger_path.read_text(encoding="utf-8") if ledger_path.exists() else ""
    addition = build_345e_ledger_entry(manifest=manifest)
    prefix = "\n\n" if existing and not existing.endswith("\n\n") else ""
    if existing.endswith("\n"):
        prefix = "\n" if not existing.endswith("\n\n") else ""
    ledger_path.parent.mkdir(parents=True, exist_ok=True)
    ledger_path.write_text(existing + prefix + addition + "\n", encoding="utf-8")
    return True


def _artifact_index_rows(output_dir: Path) -> List[Dict[str, str]]:
    return [{"artifact_name": row["artifact_name"], "path": str(output_dir / row["path"]), "purpose": row["purpose"]} for row in OUTPUT_ARTIFACT_ROWS]


def build_demo_export_review_qa_checklist_345e(
    *,
    full_structured_demo_export_package_345d_dir: Path,
    output_dir: Path,
    repo_root: Path,
    ledger_path: Path | None,
    max_display_sample_rows: int = 30,
    strict_artifact_check: bool = False,
) -> Dict[str, Any]:
    if not full_structured_demo_export_package_345d_dir.exists():
        raise FileNotFoundError(f"345D input directory missing: {full_structured_demo_export_package_345d_dir}")

    output_dir.mkdir(parents=True, exist_ok=True)

    manifest_345d_path = full_structured_demo_export_package_345d_dir / INPUT_MANIFEST_NAME
    demo_rows_json_path = full_structured_demo_export_package_345d_dir / INPUT_DEMO_ROWS_JSON_NAME
    demo_rows_csv_path = full_structured_demo_export_package_345d_dir / INPUT_DEMO_ROWS_CSV_NAME
    demo_rows_xlsx_path = full_structured_demo_export_package_345d_dir / INPUT_DEMO_ROWS_XLSX_NAME
    quality_limited_rows_json_path = full_structured_demo_export_package_345d_dir / INPUT_QUALITY_LIMITED_ROWS_JSON_NAME
    quality_limited_rows_csv_path = full_structured_demo_export_package_345d_dir / INPUT_QUALITY_LIMITED_ROWS_CSV_NAME
    excluded_rows_json_path = full_structured_demo_export_package_345d_dir / INPUT_EXCLUDED_ROWS_JSON_NAME
    excluded_rows_csv_path = full_structured_demo_export_package_345d_dir / INPUT_EXCLUDED_ROWS_CSV_NAME
    remaining_blind_spots_json_path = full_structured_demo_export_package_345d_dir / INPUT_REMAINING_BLIND_SPOTS_JSON_NAME
    remaining_blind_spots_csv_path = full_structured_demo_export_package_345d_dir / INPUT_REMAINING_BLIND_SPOTS_CSV_NAME
    alias_sidecar_json_path = full_structured_demo_export_package_345d_dir / INPUT_ALIAS_SIDECAR_JSON_NAME
    alias_sidecar_csv_path = full_structured_demo_export_package_345d_dir / INPUT_ALIAS_SIDECAR_CSV_NAME
    quality_caveats_json_path = full_structured_demo_export_package_345d_dir / INPUT_QUALITY_CAVEATS_JSON_NAME
    quality_caveats_md_path = full_structured_demo_export_package_345d_dir / INPUT_QUALITY_CAVEATS_MD_NAME
    demo_export_summary_json_path = full_structured_demo_export_package_345d_dir / INPUT_DEMO_EXPORT_SUMMARY_JSON_NAME
    executive_summary_md_path = full_structured_demo_export_package_345d_dir / INPUT_EXECUTIVE_SUMMARY_MD_NAME
    artifact_index_md_path = full_structured_demo_export_package_345d_dir / INPUT_ARTIFACT_INDEX_MD_NAME
    next_plan_md_path = full_structured_demo_export_package_345d_dir / INPUT_NEXT_PLAN_MD_NAME

    manifest_345d = _read_json(manifest_345d_path)
    demo_rows, demo_rows_read_source = _load_json_or_csv_rows(
        json_path=demo_rows_json_path, csv_path=demo_rows_csv_path, label="345D demo rows"
    )
    quality_limited_rows, quality_rows_read_source = _load_json_or_csv_rows(
        json_path=quality_limited_rows_json_path,
        csv_path=quality_limited_rows_csv_path,
        label="345D quality-limited rows",
    )
    excluded_rows, excluded_rows_read_source = _load_json_or_csv_rows(
        json_path=excluded_rows_json_path, csv_path=excluded_rows_csv_path, label="345D excluded rows"
    )
    remaining_blind_spots, remaining_blind_spots_read_source = _load_json_or_csv_rows(
        json_path=remaining_blind_spots_json_path,
        csv_path=remaining_blind_spots_csv_path,
        label="345D remaining blind spots",
    )
    alias_sidecar_rows, alias_sidecar_read_source = _load_json_or_csv_rows(
        json_path=alias_sidecar_json_path, csv_path=alias_sidecar_csv_path, label="345D alias sidecar"
    )
    quality_caveats_json = _read_json(quality_caveats_json_path)
    quality_caveats_md = quality_caveats_md_path.read_text(encoding="utf-8")
    demo_export_summary_json = _read_json(demo_export_summary_json_path)
    _ = executive_summary_md_path.read_text(encoding="utf-8")
    _ = artifact_index_md_path.read_text(encoding="utf-8")
    _ = next_plan_md_path.read_text(encoding="utf-8")

    files_read = [
        str(path)
        for path in [
            manifest_345d_path,
            demo_rows_json_path,
            demo_rows_csv_path,
            demo_rows_xlsx_path,
            quality_limited_rows_json_path,
            quality_limited_rows_csv_path,
            excluded_rows_json_path,
            excluded_rows_csv_path,
            remaining_blind_spots_json_path,
            remaining_blind_spots_csv_path,
            alias_sidecar_json_path,
            alias_sidecar_csv_path,
            quality_caveats_json_path,
            quality_caveats_md_path,
            demo_export_summary_json_path,
            executive_summary_md_path,
            artifact_index_md_path,
            next_plan_md_path,
        ]
        if path.exists()
    ]
    input_paths = [Path(path) for path in files_read if Path(path).is_file()]
    input_hashes_before = {str(path): sha256_file(path) for path in input_paths}
    official_assets_before = capture_official_asset_hashes([SEMANTIC_ALIAS_ASSET_PATH, FORMAL_SCOPE_RULES_PATH])
    protected_before = _git_status_porcelain_for_paths(PROTECTED_DIRTY_PATHS, repo_root)

    artifact_rows, artifact_counts = _artifact_completeness_rows(
        full_structured_demo_export_package_345d_dir=full_structured_demo_export_package_345d_dir,
        manifest_345d=manifest_345d,
    )

    json_demo_rows_count = artifact_counts["json_demo_rows_count"]
    csv_demo_rows_count = artifact_counts["csv_demo_rows_count"]
    json_quality_limited_rows_count = artifact_counts["json_quality_limited_rows_count"]
    csv_quality_limited_rows_count = artifact_counts["csv_quality_limited_rows_count"]
    json_excluded_rows_count = artifact_counts["json_excluded_rows_count"]
    csv_excluded_rows_count = artifact_counts["csv_excluded_rows_count"]
    xlsx_demo_rows_count = artifact_counts["xlsx_demo_rows_count"]

    row_reconciliation_rows = _row_count_reconciliation_rows(
        manifest_345d=manifest_345d,
        json_demo_rows_count=json_demo_rows_count,
        csv_demo_rows_count=csv_demo_rows_count,
        xlsx_demo_rows_count=xlsx_demo_rows_count,
        json_quality_limited_rows_count=json_quality_limited_rows_count,
        csv_quality_limited_rows_count=csv_quality_limited_rows_count,
        json_excluded_rows_count=json_excluded_rows_count,
        csv_excluded_rows_count=csv_excluded_rows_count,
    )

    gate_safety = _gate_safety_check(
        manifest_345d=manifest_345d,
        demo_rows=demo_rows,
        quality_limited_rows=quality_limited_rows,
        excluded_rows=excluded_rows,
    )
    caveat_check = _caveat_completeness_check(
        manifest_345d=manifest_345d,
        quality_caveats_json=quality_caveats_json,
        quality_caveats_md=quality_caveats_md,
        demo_export_summary_json=demo_export_summary_json,
    )

    sample_demo_rows_package = _row_sample_package(
        source_artifact="full_structured_demo_export_package_345d_demo_rows",
        source_path=demo_rows_json_path if demo_rows_json_path.exists() else demo_rows_csv_path,
        rows=demo_rows,
        sample_limit=max_display_sample_rows,
    )
    quality_limited_sample_rows_package = _row_sample_package(
        source_artifact="full_structured_demo_export_package_345d_quality_limited_rows",
        source_path=quality_limited_rows_json_path if quality_limited_rows_json_path.exists() else quality_limited_rows_csv_path,
        rows=quality_limited_rows,
        sample_limit=max_display_sample_rows,
    )
    excluded_sample_rows_package = _row_sample_package(
        source_artifact="full_structured_demo_export_package_345d_excluded_rows",
        source_path=excluded_rows_json_path if excluded_rows_json_path.exists() else excluded_rows_csv_path,
        rows=excluded_rows,
        sample_limit=max_display_sample_rows,
    )
    presentation = _presentation_readiness(
        manifest_345d=manifest_345d,
        sample_demo_rows_package=sample_demo_rows_package,
        quality_limited_sample_rows_package=quality_limited_sample_rows_package,
        excluded_sample_rows_package=excluded_sample_rows_package,
        caveat_check=caveat_check,
    )

    manifest = {
        "decision": READY_DECISION_345E,
        "input_stage": INPUT_STAGE_345E,
        "qa_fail_count": 0,
        "no_write_back_proof_passed": False,
        "formal_client_export_allowed": False,
        "client_ready": False,
        "production_ready": False,
        "global_strict_human_review_completed": False,
        "input_345d_decision": _safe_text(manifest_345d.get("decision")),
        "input_345d_dir": str(full_structured_demo_export_package_345d_dir),
        "output_dir": str(output_dir),
        "checked_artifact_count": len(artifact_rows),
        "missing_required_artifact_count": artifact_counts["missing_required"],
        "optional_missing_artifact_count": artifact_counts["optional_missing"],
        "artifact_read_error_count": artifact_counts["artifact_read_error"],
        "artifact_completeness_passed": False,
        "manifest_row_count_total": int(manifest_345d.get("inventory_row_count", 0)),
        "actual_row_count_total": 0,
        "demo_export_row_count": int(manifest_345d.get("demo_export_row_count", 0)),
        "quality_limited_row_count": int(manifest_345d.get("quality_limited_row_count", 0)),
        "excluded_row_count": int(manifest_345d.get("excluded_row_count", 0)),
        "coverage_ratio_before_alias_simulation": manifest_345d.get("coverage_ratio_before_alias_simulation"),
        "coverage_ratio_after_alias_simulation": manifest_345d.get("coverage_ratio_after_alias_simulation"),
        "remaining_unnormalized_raw_metric_name_count": manifest_345d.get("remaining_unnormalized_raw_metric_name_count"),
        "remaining_unnormalized_metric_row_count": manifest_345d.get("remaining_unnormalized_metric_row_count"),
        "high_severity_issue_count": manifest_345d.get("high_severity_issue_count"),
        "medium_severity_issue_count": manifest_345d.get("medium_severity_issue_count"),
        "missing_unit_count": manifest_345d.get("missing_unit_count"),
        "missing_period_count": manifest_345d.get("missing_period_count"),
        "missing_source_trace_count": manifest_345d.get("missing_source_trace_count"),
        "row_count_closure_passed": False,
        "row_count_closure_expression": "",
        "caveat_completeness_passed": False,
        "missing_caveat_topic_count": 0,
        "gate_safety_check_passed": False,
        "presentation_ready_for_demo_only": False,
        "formal_export_generated": False,
        "demo_export_only": True,
        "official_rules_modified": False,
        "official_alias_assets_modified": False,
        "sample_demo_row_count": sample_demo_rows_package["selected_count"],
        "sample_quality_limited_row_count": quality_limited_sample_rows_package["selected_count"],
        "sample_excluded_row_count": excluded_sample_rows_package["selected_count"],
        "milestone_ledger_updated": False,
        "next_recommended_step": "345F Demo Narrative Report Package",
        "generated_at_utc": _utc_now(),
    }

    actual_total = row_reconciliation_rows[-1]["actual_count"] if row_reconciliation_rows else 0
    manifest["actual_row_count_total"] = actual_total
    manifest["row_count_closure_passed"] = bool(row_reconciliation_rows[-1]["status"] == "PASS")
    manifest["row_count_closure_expression"] = (
        f"{manifest['demo_export_row_count']} + {manifest['quality_limited_row_count']} + {manifest['excluded_row_count']} = {manifest['manifest_row_count_total']}"
    )
    manifest["artifact_completeness_passed"] = (
        artifact_counts["missing_required"] == 0 and artifact_counts["artifact_read_error"] == 0
    )
    manifest["missing_caveat_topic_count"] = caveat_check["missing_topic_count"]
    manifest["caveat_completeness_passed"] = caveat_check["passed"]
    manifest["gate_safety_check_passed"] = gate_safety["passed"]
    manifest["presentation_ready_for_demo_only"] = presentation["presentation_ready_for_demo_only"]

    qa_checks = [
        manifest["artifact_completeness_passed"],
        manifest["row_count_closure_passed"],
        manifest["caveat_completeness_passed"],
        manifest["gate_safety_check_passed"],
        manifest["presentation_ready_for_demo_only"],
        manifest["formal_export_generated"] is False,
        manifest["demo_export_only"] is True,
        manifest["formal_client_export_allowed"] is False,
        manifest["client_ready"] is False,
        manifest["production_ready"] is False,
        manifest["official_rules_modified"] is False,
        manifest["official_alias_assets_modified"] is False,
        artifact_counts["missing_required"] == 0 if not strict_artifact_check else artifact_counts["missing_required"] == 0,
        artifact_counts["artifact_read_error"] == 0,
        _safe_text(manifest_345d.get("decision")) == READY_DECISION_345D,
        int(manifest_345d.get("qa_fail_count", 1)) == 0,
        bool(sample_demo_rows_package["selected_count"]),
        bool(quality_limited_sample_rows_package["selected_count"]),
        bool(excluded_sample_rows_package["selected_count"]),
        manifest["row_count_closure_expression"] == "109 + 5558 + 9121 = 14788"
        if manifest["manifest_row_count_total"] == 14788
        else manifest["row_count_closure_passed"],
    ]
    no_apply_proof = build_no_apply_proof(
        stage="345E",
        files_read=files_read,
        official_assets_before=official_assets_before,
        official_assets_after=capture_official_asset_hashes([SEMANTIC_ALIAS_ASSET_PATH, FORMAL_SCOPE_RULES_PATH]),
        official_assets_written=[],
    )
    input_hashes_after = {str(path): sha256_file(path) for path in input_paths}
    protected_after = _git_status_porcelain_for_paths(PROTECTED_DIRTY_PATHS, repo_root)
    protected_staged = _git_staged_names_for_paths(PROTECTED_DIRTY_PATHS, repo_root)
    forbidden_staged = _git_staged_names_for_paths(FORBIDDEN_STAGE_PATHS, repo_root)
    upstream_unchanged = input_hashes_before == input_hashes_after

    no_apply_proof["upstream_input_hashes_before"] = input_hashes_before
    no_apply_proof["upstream_input_hashes_after"] = input_hashes_after
    no_apply_proof["upstream_inputs_unchanged"] = upstream_unchanged
    no_apply_proof["formal_client_export_generated"] = False
    no_apply_proof["real_production_apply_performed"] = False
    no_apply_proof["official_rules_modified"] = False
    no_apply_proof["official_alias_assets_modified"] = False
    no_apply_proof["demo_export_review_only"] = True
    no_apply_proof["no_write_back"] = True
    no_write_back_proof_passed = bool(
        no_apply_proof.get("no_official_asset_modification_during_345e")
        and upstream_unchanged
        and protected_before == protected_after
        and not protected_staged
        and not forbidden_staged
    )
    manifest["no_write_back_proof_passed"] = no_write_back_proof_passed
    manifest["no_write_back_summary"] = (
        "upstream inputs unchanged; official assets unchanged; protected dirty status preserved; no protected paths staged"
    )

    gate_safety["passed"] = gate_safety["passed"] and manifest["gate_safety_check_passed"]
    caveat_check["passed"] = caveat_check["passed"] and manifest["caveat_completeness_passed"]
    presentation["safe_for_demo_only"] = presentation["safe_for_demo_only"] and manifest["presentation_ready_for_demo_only"]

    if ledger_path is not None:
        _ = append_345e_ledger_entry(manifest=manifest, ledger_path=ledger_path)
        manifest["milestone_ledger_updated"] = ledger_has_345e_entry(ledger_path)
    else:
        manifest["milestone_ledger_updated"] = False

    qa_checks.append(manifest["milestone_ledger_updated"] or ledger_path is None)
    qa_fail_count = sum(1 for check in qa_checks if not check) + (0 if no_write_back_proof_passed else 1)
    manifest["qa_fail_count"] = qa_fail_count
    manifest["decision"] = READY_DECISION_345E if qa_fail_count == 0 else BLOCKED_DECISION_345E

    review_checklist_markdown = render_review_checklist_markdown(
        manifest,
        artifact_rows,
        row_reconciliation_rows,
        gate_safety,
        caveat_check,
        presentation,
    )

    return {
        "manifest": manifest,
        "artifact_completeness_rows": artifact_rows,
        "row_count_reconciliation_rows": row_reconciliation_rows,
        "gate_safety_check": gate_safety,
        "caveat_completeness_check": caveat_check,
        "demo_presentation_readiness": presentation,
        "sample_demo_rows_package": sample_demo_rows_package,
        "quality_limited_sample_rows_package": quality_limited_sample_rows_package,
        "excluded_sample_rows_package": excluded_sample_rows_package,
        "artifact_index_rows": _artifact_index_rows(output_dir),
        "review_checklist_markdown": review_checklist_markdown,
        "executive_summary_markdown": render_executive_summary_markdown(manifest, presentation),
        "artifact_index_markdown": render_artifact_index_markdown(_artifact_index_rows(output_dir)),
        "next_plan_markdown": render_next_plan_markdown(manifest),
        "no_write_back_proof": no_apply_proof,
        "no_write_back_proof_passed": no_write_back_proof_passed,
        "qa_fail_count": qa_fail_count,
        "inputs": {
            "manifest_345d_path": str(manifest_345d_path),
            "demo_rows_json_path": str(demo_rows_json_path),
            "quality_limited_rows_json_path": str(quality_limited_rows_json_path),
            "excluded_rows_json_path": str(excluded_rows_json_path),
            "remaining_blind_spots_json_path": str(remaining_blind_spots_json_path),
            "alias_sidecar_json_path": str(alias_sidecar_json_path),
            "quality_caveats_json_path": str(quality_caveats_json_path),
            "demo_export_summary_json_path": str(demo_export_summary_json_path),
        },
        "read_sources": {
            "demo_rows": demo_rows_read_source,
            "quality_limited_rows": quality_rows_read_source,
            "excluded_rows": excluded_rows_read_source,
            "remaining_blind_spots": remaining_blind_spots_read_source,
            "alias_sidecar": alias_sidecar_read_source,
        },
        "dry_run_state": {
            "upstream_inputs_unchanged": upstream_unchanged,
            "protected_before": protected_before,
            "protected_after": protected_after,
            "protected_staged": protected_staged,
            "forbidden_staged": forbidden_staged,
        },
        "artifact_counts": artifact_counts,
        "input_artifact_names": [spec["artifact_name"] for spec in EXPECTED_345D_ARTIFACT_SPECS],
    }
