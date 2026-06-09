from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_BENCHMARK_SUMMARY",
    "02_PDF_INVENTORY",
    "03_SAMPLE_TIERS",
    "04_TARGET_METRICS",
    "05_RUN_PLAN",
    "06_REVIEW_BUDGET",
    "07_SUCCESS_CRITERIA",
    "08_RISK_REGISTER",
    "09_NEXT_STEPS",
    "10_NO_WRITE_BACK_PROOF",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "notes",
        "status",
        "warning",
        "reference",
        "hash",
        "risk",
        "goal",
        "output",
        "input",
        "criteria",
        "driver",
        "description",
        "rationale",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 180)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 200)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# Larger Real-PDF Benchmark Plan 342A",
        "",
        "342A creates a larger real-PDF benchmark planning package from the current repository inputs and validated preview milestone outputs.",
        "It is a planning and inventory stage only, not a parsing rerun and not a client export stage.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- warning_count: {summary.get('warning_count', 0)}",
        f"- benchmark_status: {summary.get('benchmark_status', '')}",
        "",
        "## Inventory",
        f"- current_pdf_count: {summary.get('current_pdf_count', 0)}",
        f"- target_pdf_count_min: {summary.get('target_pdf_count_min', 0)}",
        f"- target_pdf_count_recommended: {summary.get('target_pdf_count_recommended', 0)}",
        f"- target_pdf_count_stretch: {summary.get('target_pdf_count_stretch', 0)}",
        "",
        "## Upstream Detection",
        f"- detected_341a_decision: {summary.get('detected_341a_decision', '')}",
        f"- detected_340g_decision: {summary.get('detected_340g_decision', '')}",
        f"- detected_340f_decision: {summary.get('detected_340f_decision', '')}",
        f"- detected_340g_audit_passed: {summary.get('detected_340g_audit_passed', '')}",
        f"- detected_340f_client_preview_core_metric_count: {summary.get('detected_340f_client_preview_core_metric_count', '')}",
        "",
        "## Safety",
        f"- demo_ready: {summary.get('demo_ready', False)}",
        f"- client_preview_ready: {summary.get('client_preview_ready', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        f"- output_workbook_path: {summary.get('output_workbook_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.append("")
    return "\n".join(lines)
