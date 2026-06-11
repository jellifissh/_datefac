from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_AUDIT_SUMMARY",
    "02_INPUT_342P_SUMMARY",
    "03_PREVIEW_AUDIT",
    "04_TRUST_LEVEL_AUDIT",
    "05_SIM_BOUNDARY_AUDIT",
    "06_COLLISION_AUDIT",
    "07_DROPPED_DUP_AUDIT",
    "08_OVERRIDE_AUDIT",
    "09_EXPORT_RISK_GATE",
    "10_EXPORT_CANDIDATE_SCOPE",
    "11_REMAINING_BACKLOG",
    "12_342R_READINESS",
    "13_NO_WRITE_BACK",
    "14_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "warning",
        "path",
        "evidence",
        "recommendation",
        "value",
        "disclaimer",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 180)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 220)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 342Q Preview Audit And Export Readiness Gate",
        "",
        "## Chinese Summary",
        "- 342Q is a preview audit gate over the 342P reviewed-plus-simulated combined preview.",
        "- It audits trust level, simulation boundary, collision handling, dropped duplicates, and export-risk boundaries.",
        "- It does not generate a formal client export, and it keeps `formal_client_export_allowed=false`, `client_ready=false`, and `production_ready=false`.",
        "",
        "## English Summary",
        "- 342Q audits export readiness boundaries for the 342P bounded preview package.",
        "- It only allows an audit-labeled 342R candidate scope, not a formal client export.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- ready_for_342r: {summary.get('ready_for_342r', False)}",
        f"- recommended_342r_scope: {summary.get('recommended_342r_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Key Counts",
        f"- human_reviewed_preview_count: {summary.get('human_reviewed_preview_count', 0)}",
        f"- simulated_preview_count: {summary.get('simulated_preview_count', 0)}",
        f"- simulated_direct_preview_count: {summary.get('simulated_direct_preview_count', 0)}",
        f"- simulated_corrected_preview_count: {summary.get('simulated_corrected_preview_count', 0)}",
        f"- combined_preview_row_count: {summary.get('combined_preview_row_count', 0)}",
        f"- export_candidate_row_count: {summary.get('export_candidate_row_count', 0)}",
        f"- still_human_required_count: {summary.get('still_human_required_count', 0)}",
        f"- remaining_review_count: {summary.get('remaining_review_count', 0)}",
        "",
        "## Boundary Audit",
        f"- unknown_trust_level_count: {summary.get('unknown_trust_level_count', 0)}",
        f"- trust_level_mismatch_count: {summary.get('trust_level_mismatch_count', 0)}",
        f"- simulated_final_confirmed_true_count: {summary.get('simulated_final_confirmed_true_count', 0)}",
        f"- simulated_client_ready_true_count: {summary.get('simulated_client_ready_true_count', 0)}",
        f"- simulated_production_ready_true_count: {summary.get('simulated_production_ready_true_count', 0)}",
        f"- missing_display_warning_count: {summary.get('missing_display_warning_count', 0)}",
        "",
        "## Collision And Risk",
        f"- collision_logged_count: {summary.get('collision_logged_count', 0)}",
        f"- duplicate_metric_year_source_count: {summary.get('duplicate_metric_year_source_count', 0)}",
        f"- human_over_simulation_override_count: {summary.get('human_over_simulation_override_count', 0)}",
        f"- simulated_duplicate_dropped_count: {summary.get('simulated_duplicate_dropped_count', 0)}",
        f"- unresolved_collision_count: {summary.get('unresolved_collision_count', 0)}",
        f"- severe_collision_count: {summary.get('severe_collision_count', 0)}",
        f"- export_risk_level: {summary.get('export_risk_level', '')}",
        "",
        "## Safety",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- export_candidate_scope_allowed: {summary.get('export_candidate_scope_allowed', False)}",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- output_workbook_path: {summary.get('output_workbook_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Boundaries",
            "- 342Q does not rerun MinerU or upstream 342E-342P stages.",
            "- 342Q does not call VLM or a real LLM API.",
            "- 342Q does not write back to upstream workbooks.",
            "- 342Q does not convert preview outputs into formal client delivery or investment advice.",
            "",
        ]
    )
    return "\n".join(lines)
