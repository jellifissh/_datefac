from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any, Dict, Iterable, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_csv(path: Path, rows: List[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    pd.DataFrame(rows).to_csv(path, index=False, encoding="utf-8-sig")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "value",
        "note",
        "reason",
        "recommendation",
        "evidence",
        "caveat",
        "issue",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, List[Dict[str, Any]]], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            pd.DataFrame(sheets.get(sheet_name, [])).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def executive_summary_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# 345D Full Structured Demo Export Package",
            "",
            "## Input Context",
            f"- input_345a_decision: {summary.get('input_345a_decision', '')}",
            f"- input_345b_decision: {summary.get('input_345b_decision', '')}",
            f"- input_345c_decision: {summary.get('input_345c_decision', '')}",
            f"- input_345c11_decision: {summary.get('input_345c11_decision', '')}",
            "- 345C11 explicitly stopped the alias-governance branch before returning to 345D.",
            "",
            "## Export Package",
            f"- demo_export_row_count: {summary.get('demo_export_row_count', 0)}",
            f"- quality_limited_row_count: {summary.get('quality_limited_row_count', 0)}",
            f"- excluded_row_count: {summary.get('excluded_row_count', 0)}",
            f"- alias_simulated_demo_row_count: {summary.get('alias_simulated_demo_row_count', 0)}",
            "",
            "## Coverage",
            f"- coverage_ratio_before_alias_simulation: {summary.get('coverage_ratio_before_alias_simulation', None)}",
            f"- coverage_ratio_after_alias_simulation: {summary.get('coverage_ratio_after_alias_simulation', None)}",
            f"- cumulative_alias_simulated_newly_normalized_row_count: {summary.get('cumulative_alias_simulated_newly_normalized_row_count', 0)}",
            "",
            "## Caveats",
            f"- remaining_unnormalized_raw_metric_name_count: {summary.get('remaining_unnormalized_raw_metric_name_count', 0)}",
            f"- remaining_unnormalized_metric_row_count: {summary.get('remaining_unnormalized_metric_row_count', 0)}",
            f"- high_severity_issue_count: {summary.get('high_severity_issue_count', 0)}",
            f"- medium_severity_issue_count: {summary.get('medium_severity_issue_count', 0)}",
            f"- missing_unit_count_in_quality_limited_rows: {summary.get('missing_unit_count', 0)}",
            f"- missing_source_trace_count_in_quality_limited_rows: {summary.get('missing_source_trace_count', 0)}",
            "",
            "## Boundary",
            "- 345D only creates a demo-only structured export package.",
            "- Official normalization rules were not modified.",
            "- Official alias assets were not modified.",
            "- Rows normalized through alias simulation remain sidecar-derived, not official rule mutations.",
            f"- formal_client_export_allowed = {summary.get('formal_client_export_allowed', False)}",
            f"- client_ready = {summary.get('client_ready', False)}",
            f"- production_ready = {summary.get('production_ready', False)}",
            "",
            "## Recommendation",
            f"- next_recommended_step: {summary.get('next_recommended_step', '')}",
            "- 344G still waits for a genuinely human-filled 344F workbook.",
        ]
    )


def quality_caveats_markdown(caveats: Dict[str, Any]) -> str:
    lines = [
        "# 345D Quality Caveats",
        "",
        f"- remaining_unnormalized_raw_metric_name_count: {caveats.get('remaining_unnormalized_raw_metric_name_count', 0)}",
        f"- remaining_unnormalized_metric_row_count: {caveats.get('remaining_unnormalized_metric_row_count', 0)}",
        f"- remaining_ready_candidate_count: {caveats.get('remaining_ready_candidate_count', 0)}",
        f"- missing_unit_count: {caveats.get('missing_unit_count', 0)}",
        f"- missing_period_count: {caveats.get('missing_period_count', 0)}",
        f"- missing_source_trace_count: {caveats.get('missing_source_trace_count', 0)}",
        f"- high_severity_issue_count: {caveats.get('high_severity_issue_count', 0)}",
        f"- medium_severity_issue_count: {caveats.get('medium_severity_issue_count', 0)}",
        f"- rejected_or_excluded_count: {caveats.get('rejected_or_excluded_count', 0)}",
        f"- rows_normalized_only_through_simulation_count: {caveats.get('rows_normalized_only_through_simulation_count', 0)}",
        "",
        "## Policy",
        f"- {caveats.get('quality_limited_scope_policy', '')}",
        f"- simulation_exact_match_limitation: {caveats.get('simulation_exact_match_limitation', '')}",
        "- formal_client_export_allowed remains false.",
        "- production_ready remains false.",
        "",
        "## Top Quality Issue Counts",
    ]
    for key, value in sorted((caveats.get("quality_issue_counts") or {}).items(), key=lambda item: (-item[1], item[0]))[:10]:
        lines.append(f"- {key}: {value}")
    lines.append("")
    lines.append("## Top Exclusion Reason Counts")
    for key, value in sorted((caveats.get("excluded_reason_counts") or {}).items(), key=lambda item: (-item[1], item[0]))[:10]:
        lines.append(f"- {key}: {value}")
    return "\n".join(lines)


def artifact_index_markdown(rows: Iterable[Dict[str, Any]]) -> str:
    lines = [
        "# 345D Artifact Index",
        "",
        "| Artifact | Path | Use |",
        "| --- | --- | --- |",
    ]
    for row in rows:
        lines.append(
            f"| {row.get('artifact_name', '')} | {row.get('path', '')} | {row.get('purpose', '')} |"
        )
    return "\n".join(lines)


def next_plan_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# 345D Next Plan",
            "",
            f"- Recommended next scope: {summary.get('next_recommended_step', '')}",
            "- Demo output can be reviewed and narrated, but it must not be reframed as a formal client export.",
            "- Any official rule update must be a separate explicitly approved task, not an automatic consequence of 345D.",
            "- 344G still waits for a genuinely human-filled 344F workbook.",
        ]
    )
