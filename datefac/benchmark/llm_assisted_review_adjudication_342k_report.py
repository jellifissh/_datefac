from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable, Sequence

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_LLM_REVIEW_SUMMARY",
    "02_INPUT_342J_SUMMARY",
    "03_LLM_CANDIDATE_POOL",
    "04_RULE_BASELINE",
    "05_PROMPT_PACKAGE",
    "06_EXPECTED_SCHEMA",
    "07_DRY_RUN_SUGGESTIONS",
    "08_HUMAN_REQUIRED",
    "09_AUTO_CONFIRM_CANDIDATES",
    "10_CONFLICTS",
    "11_RISK_BUCKETS",
    "12_REVIEW_TEMPLATE_DRAFT",
    "13_342L_READINESS",
    "14_NO_WRITE_BACK",
    "15_NEXT_STEPS",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def write_jsonl(path: Path, rows: Sequence[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8") as handle:
        for row in rows:
            handle.write(json.dumps(_to_jsonable(row), ensure_ascii=False) + "\n")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "note",
        "warning",
        "reference",
        "hash",
        "risk",
        "path",
        "decision",
        "action",
        "recommendation",
        "html",
        "snippet",
        "bbox",
        "prompt",
        "json",
        "schema",
        "evidence",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 180)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 220)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 342K LLM-Assisted Review Adjudication Pilot",
        "",
        "中文：",
        "342K 是一个 LLM-assisted review adjudication pilot，但默认不调用真实 LLM API。",
        f"本次主要生成 {summary.get('llm_candidate_pool_count', 0)} 条候选的 rule baseline、prompt/request package，以及 dry-run suggestions。",
        "这些 dry-run suggestions 不是最终 LLM 输出，更不是人工审核结论。",
        f"当前仍有 {summary.get('pending_review_count', 0)} 条 pending review，必须保持 client_ready = false、production_ready = false。",
        "",
        "English:",
        "342K is an LLM-assisted review adjudication pilot, but it defaults to no real LLM API execution.",
        f"This run mainly generates rule baselines, prompt/request packages, and dry-run suggestions for {summary.get('llm_candidate_pool_count', 0)} candidate rows.",
        "Those dry-run suggestions are neither final LLM responses nor human-review outcomes.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- ready_for_342l: {summary.get('ready_for_342l', False)}",
        f"- recommended_342l_scope: {summary.get('recommended_342l_scope', '')}",
        "",
        "## Counts",
        f"- pending_review_count: {summary.get('pending_review_count', 0)}",
        f"- llm_candidate_pool_count: {summary.get('llm_candidate_pool_count', 0)}",
        f"- prompt_package_count: {summary.get('prompt_package_count', 0)}",
        f"- request_pack_count: {summary.get('request_pack_count', 0)}",
        f"- rule_baseline_count: {summary.get('rule_baseline_count', 0)}",
        f"- dry_run_suggestion_count: {summary.get('dry_run_suggestion_count', 0)}",
        f"- human_required_count: {summary.get('human_required_count', 0)}",
        f"- auto_confirm_candidate_count: {summary.get('auto_confirm_candidate_count', 0)}",
        f"- conflict_count: {summary.get('conflict_count', 0)}",
        f"- unit_year_risk_count: {summary.get('unit_year_risk_count', 0)}",
        f"- duplicate_risk_count: {summary.get('duplicate_risk_count', 0)}",
        f"- growth_row_risk_count: {summary.get('growth_row_risk_count', 0)}",
        f"- source_trace_risk_count: {summary.get('source_trace_risk_count', 0)}",
        f"- metric_mapping_risk_count: {summary.get('metric_mapping_risk_count', 0)}",
        "",
        "## Safety",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- prompt_pack_path: {summary.get('prompt_pack_path', '')}",
        f"- request_pack_path: {summary.get('request_pack_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    if qa_json.get("warnings"):
        lines.extend(["", "## Warnings"])
        for warning in qa_json.get("warnings", []):
            lines.append(f"- {warning}")
    lines.extend(
        [
            "",
            "## Boundaries",
            "- 342K does not rerun MinerU.",
            "- 342K does not call VLM.",
            "- 342K does not call a real LLM API by default.",
            "- 342K does not modify production pipeline / parser / extraction / delivery.",
            "- 342K remains a no-write-back adjudication pilot.",
            "",
        ]
    )
    return "\n".join(lines)
