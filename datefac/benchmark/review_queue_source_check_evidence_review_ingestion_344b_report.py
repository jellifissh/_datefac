from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from datefac.review_queue.source_check_backlog_package_344a import (
    EDITABLE_SOURCE_CHECK_COLUMNS_344A,
)


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows),
        encoding="utf-8",
    )


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "value",
        "note",
        "reason",
        "recommendation",
        "evidence",
        "bbox",
        "html",
        "text",
        "disclosure",
        "scope",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                if str(cell.value or "") in EDITABLE_SOURCE_CHECK_COLUMNS_344A:
                    cell.fill = editable_fill
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 344B Source-check Evidence Review Result Ingestion",
        "",
        "## 中文摘要",
        "- 344B 已读取 344A2 enriched source-check review workbook，并将 19 行结果导入 sidecar-only ingestion 产物。",
        "- 当前 happy path 为 10 行 `SOURCE_CONFIRM`，9 行 `SOURCE_CORRECT`。",
        "- 这 9 行 correction 不是收入额行，而是把 `revenue / 亿元` 修正为 `YOY / %`，并保留原 source year/value。",
        "- 344B 只解决 source-check backlog 的 sidecar 审核结果导入，不做真实写回，不做 formal client export。",
        "- 下一步应进入 344C source-check confirmed sidecar apply simulation and expanded trust gate。",
        "",
        "## English Summary",
        "- 344B ingests the filled 344A2 source-check workbook into sidecar-only review results.",
        "- The current happy path confirms 10 rows and corrects 9 rows.",
        "- The 9 corrected rows are YOY percentage rows, not revenue amount rows.",
        "- No production write-back or formal client export occurred.",
        "- The next safe task is 344C sidecar apply simulation and expanded trust gate.",
        "",
        "## Key Metrics",
        f"- decision: {summary.get('decision', '')}",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- filled_workbook_path: {summary.get('filled_workbook_path', '')}",
        f"- filled_row_count: {summary.get('filled_row_count', 0)}",
        f"- valid_row_count: {summary.get('valid_row_count', 0)}",
        f"- invalid_row_count: {summary.get('invalid_row_count', 0)}",
        f"- source_confirm_count: {summary.get('source_confirm_count', 0)}",
        f"- source_correct_count: {summary.get('source_correct_count', 0)}",
        f"- source_reject_count: {summary.get('source_reject_count', 0)}",
        f"- source_still_insufficient_count: {summary.get('source_still_insufficient_count', 0)}",
        f"- source_defer_count: {summary.get('source_defer_count', 0)}",
        f"- validated_sidecar_row_count: {summary.get('validated_sidecar_row_count', 0)}",
        f"- correction_row_count: {summary.get('correction_row_count', 0)}",
        f"- source_check_result_ingested: {summary.get('source_check_result_ingested', False)}",
        f"- source_check_backlog_resolved: {summary.get('source_check_backlog_resolved', False)}",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- global_strict_human_review_completed: {summary.get('global_strict_human_review_completed', False)}",
        f"- ready_for_344c: {summary.get('ready_for_344c', False)}",
        f"- recommended_344c_scope: {summary.get('recommended_344c_scope', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(
            f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})"
        )
    return "\n".join(lines)

