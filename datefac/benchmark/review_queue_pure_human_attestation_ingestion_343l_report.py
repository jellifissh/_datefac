from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from datefac.review_queue.pure_human_attestation_package_343k import (
    EDITABLE_HUMAN_ATTESTATION_COLUMNS,
)


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "isoformat"):
        try:
            return value.isoformat()
        except Exception:
            pass
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any] | list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def write_jsonl(path: Path, rows: list[Dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        "\n".join(json.dumps(_to_jsonable(row), ensure_ascii=False) for row in rows),
        encoding="utf-8",
    )


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    editable_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")
    wrap_keywords = {
        "description",
        "detail",
        "rule",
        "path",
        "message",
        "warning",
        "value",
        "note",
        "reason",
        "recommendation",
        "evidence",
        "bbox",
        "html",
        "text",
        "snippet",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )
                if str(cell.value or "") in EDITABLE_HUMAN_ATTESTATION_COLUMNS:
                    cell.fill = editable_fill
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                text = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(text))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(
                writer,
                sheet_name=sheet_name,
                index=False,
            )
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# 343L Pure Human Attestation Result Ingestion",
        "",
        "## 中文摘要",
        "- 343L 已读取 343K 生成的 pure-human attestation workbook，并校验 10 行人工确认结果。",
        "- 即使全部通过，也只表示 `343K_PACKAGE_ONLY` 范围内的 pure human confirmation 完成。",
        "- 这不代表整个 corpus 已完成严格人工审核。",
        "- formal client export 仍然禁止，client_ready / production_ready 仍为 false。",
        "",
        "## English Summary",
        "- 343L ingests the filled pure-human attestation workbook and validates the package-scoped human confirmation results.",
        "- Even with full acceptance, completion only applies to the 343K package scope.",
        "- Formal client export remains forbidden and the whole corpus is not globally strict-human-complete.",
        "",
        "## Key Metrics",
        f"- decision: {summary.get('decision', '')}",
        f"- review_queue_schema_version: {summary.get('review_queue_schema_version', '')}",
        f"- filled_workbook_path: {summary.get('filled_workbook_path', '')}",
        f"- filled_row_count: {summary.get('filled_row_count', 0)}",
        f"- valid_row_count: {summary.get('valid_row_count', 0)}",
        f"- invalid_row_count: {summary.get('invalid_row_count', 0)}",
        f"- human_accept_count: {summary.get('human_accept_count', 0)}",
        f"- human_correct_count: {summary.get('human_correct_count', 0)}",
        f"- human_reject_count: {summary.get('human_reject_count', 0)}",
        f"- human_needs_source_check_count: {summary.get('human_needs_source_check_count', 0)}",
        f"- human_defer_count: {summary.get('human_defer_count', 0)}",
        f"- pure_strict_human_review_completed_for_package: {summary.get('pure_strict_human_review_completed_for_package', False)}",
        f"- strict_human_review_completed_scope: {summary.get('strict_human_review_completed_scope', '')}",
        f"- global_strict_human_review_completed: {summary.get('global_strict_human_review_completed', False)}",
        f"- formal_client_export_allowed: {summary.get('formal_client_export_allowed', False)}",
        f"- ready_for_343m: {summary.get('ready_for_343m', False)}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(
            f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})"
        )
    return "\n".join(lines)
