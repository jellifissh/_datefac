from __future__ import annotations

import csv
import io
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Dict, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

from datefac.table_bakeoff.structtable_table_normalizer import (
    detect_table_title,
    guess_cell_type,
    normalize_text,
)


TEXT_ENCODINGS = ("utf-8-sig", "utf-8", "gb18030", "gbk")
MOJIBAKE_MARKERS = ("锟", "锛", "鈥", "鍏", "鍒", "褰", "姣", "鐜", "璐", "娑", "绌", "鎹", "鏈")


@dataclass
class StructTableDiscoveredGroup:
    image_name: str
    input_image_path: str
    output_folder: str
    raw_response_path: str
    markdown_path: str
    xlsx_path: str
    csv_path: str
    run_meta_path: str
    error_path: str
    warnings: List[str] = field(default_factory=list)


@dataclass
class StructTableNormalizedCell:
    row_index: int
    col_index: int
    text: str
    normalized_text: str
    cell_type_guess: str
    warnings: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "row_index": self.row_index,
            "col_index": self.col_index,
            "text": self.text,
            "normalized_text": self.normalized_text,
            "cell_type_guess": self.cell_type_guess,
            "warnings": list(self.warnings),
        }


@dataclass
class StructTableNormalizedRow:
    row_index: int
    label: str
    values: List[str]
    warnings: List[str] = field(default_factory=list)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "row_index": self.row_index,
            "label": self.label,
            "values": list(self.values),
            "warnings": list(self.warnings),
        }


@dataclass
class StructTableNormalizedTable:
    tool: str
    recognition_source: str
    image_name: str
    input_image_path: str
    table_id: str
    table_title: str
    columns: List[str]
    rows: List[StructTableNormalizedRow]
    cells: List[StructTableNormalizedCell]
    provenance: Dict[str, Any]
    warnings: List[str]
    parse_status: str
    raw_response_text: str
    raw_response_has_markdown_marker: bool
    raw_response_timeout_warning_count: int
    returncode: Optional[int]
    error_text: str
    csv_grid_rows: List[List[str]]
    markdown_grid_rows: List[List[str]]
    xlsx_grid_rows: List[List[str]]

    def to_dict(self) -> Dict[str, Any]:
        return {
            "tool": self.tool,
            "recognition_source": self.recognition_source,
            "image_name": self.image_name,
            "table_id": self.table_id,
            "table_title": self.table_title,
            "columns": list(self.columns),
            "rows": [row.to_dict() for row in self.rows],
            "cells": [cell.to_dict() for cell in self.cells],
            "provenance": self.provenance,
            "warnings": list(self.warnings),
            "parse_status": self.parse_status,
        }


def _norm(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _decode_score(text: str) -> int:
    cjk_count = sum(1 for char in text if "\u4e00" <= char <= "\u9fff")
    replacement_penalty = text.count("\ufffd") * 20
    mojibake_penalty = sum(text.count(marker) for marker in MOJIBAKE_MARKERS) * 3
    return cjk_count - replacement_penalty - mojibake_penalty


def read_text_with_fallback(path: Path) -> str:
    raw = path.read_bytes()
    best_text = ""
    best_score = -10**9
    for encoding in TEXT_ENCODINGS:
        try:
            text = raw.decode(encoding)
        except Exception:
            continue
        score = _decode_score(text)
        if score > best_score:
            best_text = text
            best_score = score
    if best_text:
        return best_text
    return raw.decode("utf-8", errors="ignore")


def _trim_grid(grid_rows: Sequence[Sequence[str]]) -> List[List[str]]:
    rows = [[_norm(cell) for cell in row] for row in grid_rows]
    while rows and not any(_norm(cell) for cell in rows[-1]):
        rows.pop()
    max_cols = max((len(row) for row in rows), default=0)
    while max_cols > 0 and all((col >= len(row) or not _norm(row[col])) for row in rows for col in [max_cols - 1]):
        max_cols -= 1
    normalized_rows: List[List[str]] = []
    for row in rows:
        trimmed = row[:max_cols] + [""] * max(0, max_cols - len(row))
        normalized_rows.append(trimmed)
    return normalized_rows


def parse_key_value_text(path: Path) -> Dict[str, str]:
    result: Dict[str, str] = {}
    if not path.exists():
        return result
    for line in read_text_with_fallback(path).splitlines():
        if ":" not in line:
            continue
        key, value = line.split(":", 1)
        result[_norm(key)] = _norm(value)
    return result


def parse_csv_grid(path: Path) -> Tuple[List[List[str]], List[str]]:
    warnings: List[str] = []
    if not path.exists():
        return [], ["CSV_FILE_MISSING"]
    try:
        text = read_text_with_fallback(path)
        reader = csv.reader(io.StringIO(text))
        rows = [[normalize_text(cell) for cell in row] for row in reader]
    except Exception as exc:
        return [], [f"CSV_PARSE_FAILED:{exc}"]
    return _trim_grid(rows), warnings


def parse_markdown_grid(path: Path) -> Tuple[List[List[str]], List[str]]:
    warnings: List[str] = []
    if not path.exists():
        return [], ["MARKDOWN_FILE_MISSING"]
    try:
        text = read_text_with_fallback(path)
    except Exception as exc:
        return [], [f"MARKDOWN_READ_FAILED:{exc}"]

    table_lines: List[str] = []
    for line in text.splitlines():
        stripped = line.strip()
        if stripped.startswith("|") and stripped.endswith("|"):
            table_lines.append(stripped)
    if not table_lines:
        return [], ["MARKDOWN_TABLE_NOT_FOUND"]

    rows: List[List[str]] = []
    for line in table_lines:
        cells = [normalize_text(cell) for cell in line.strip().strip("|").split("|")]
        non_empty = [cell for cell in cells if cell]
        if non_empty and all(set(cell.replace(":", "")) <= {"-"} for cell in non_empty):
            continue
        rows.append(cells)
    return _trim_grid(rows), warnings


def parse_xlsx_grid(path: Path) -> Tuple[List[List[str]], List[str]]:
    warnings: List[str] = []
    if not path.exists():
        return [], ["XLSX_FILE_MISSING"]
    try:
        workbook = load_workbook(filename=path, read_only=True, data_only=True)
        sheet = workbook[workbook.sheetnames[0]]
        rows = []
        for row in sheet.iter_rows(values_only=True):
            rows.append([normalize_text(cell) for cell in row])
        workbook.close()
    except Exception as exc:
        return [], [f"XLSX_PARSE_FAILED:{exc}"]
    return _trim_grid(rows), warnings


def discover_structtable_groups(input_image_dir: Path, structtable_output_dir: Path) -> List[StructTableDiscoveredGroup]:
    groups: List[StructTableDiscoveredGroup] = []
    input_images = sorted(path for path in input_image_dir.iterdir() if path.is_file()) if input_image_dir.exists() else []
    for image_path in input_images:
        stem = image_path.stem
        folder = structtable_output_dir / stem
        warnings: List[str] = []
        output_folder = str(folder) if folder.exists() else ""
        raw_response_path = str(folder / "raw_response_markdown_utf8.txt") if (folder / "raw_response_markdown_utf8.txt").exists() else ""
        markdown_path = str(folder / "table_output_markdown.md") if (folder / "table_output_markdown.md").exists() else ""
        xlsx_path = str(folder / "table_output_from_markdown.xlsx") if (folder / "table_output_from_markdown.xlsx").exists() else ""
        csv_path = str(folder / "table_output_from_markdown.csv") if (folder / "table_output_from_markdown.csv").exists() else ""
        run_meta_path = str(folder / "run_meta.txt") if (folder / "run_meta.txt").exists() else ""
        error_path = str(folder / "error.txt") if (folder / "error.txt").exists() else ""
        if not folder.exists():
            warnings.append("OUTPUT_FOLDER_MISSING")
        if not raw_response_path:
            warnings.append("RAW_RESPONSE_MISSING")
        if not markdown_path:
            warnings.append("MARKDOWN_OUTPUT_MISSING")
        if not xlsx_path:
            warnings.append("XLSX_OUTPUT_MISSING")
        if not csv_path:
            warnings.append("CSV_OUTPUT_MISSING")
        if not run_meta_path:
            warnings.append("RUN_META_MISSING")
        groups.append(
            StructTableDiscoveredGroup(
                image_name=image_path.name,
                input_image_path=str(image_path),
                output_folder=output_folder,
                raw_response_path=raw_response_path,
                markdown_path=markdown_path,
                xlsx_path=xlsx_path,
                csv_path=csv_path,
                run_meta_path=run_meta_path,
                error_path=error_path,
                warnings=warnings,
            )
        )
    return groups


def parse_structtable_group(
    group: StructTableDiscoveredGroup,
    batch_summary_row: Optional[Dict[str, Any]] = None,
) -> StructTableNormalizedTable:
    warnings = list(group.warnings)
    run_meta = parse_key_value_text(Path(group.run_meta_path)) if group.run_meta_path else {}
    error_text = read_text_with_fallback(Path(group.error_path)) if group.error_path else ""
    raw_response_text = read_text_with_fallback(Path(group.raw_response_path)) if group.raw_response_path else ""
    csv_rows, csv_warnings = parse_csv_grid(Path(group.csv_path)) if group.csv_path else ([], ["CSV_FILE_MISSING"])
    markdown_rows, markdown_warnings = parse_markdown_grid(Path(group.markdown_path)) if group.markdown_path else ([], ["MARKDOWN_FILE_MISSING"])
    xlsx_rows, xlsx_warnings = parse_xlsx_grid(Path(group.xlsx_path)) if group.xlsx_path else ([], ["XLSX_FILE_MISSING"])
    warnings.extend(csv_warnings)
    warnings.extend(markdown_warnings)
    warnings.extend(xlsx_warnings)

    if csv_rows:
        primary_rows = csv_rows
        source_format = "csv"
        parse_status = "PARSED_FROM_CSV"
    elif markdown_rows:
        primary_rows = markdown_rows
        source_format = "markdown"
        parse_status = "PARSED_FROM_MARKDOWN"
    elif xlsx_rows:
        primary_rows = xlsx_rows
        source_format = "xlsx"
        parse_status = "PARSED_FROM_XLSX"
    else:
        primary_rows = []
        source_format = ""
        parse_status = "PARSE_FAILED"
        warnings.append("NO_PARSEABLE_TABLE_GRID")

    if csv_rows and markdown_rows and (len(csv_rows) != len(markdown_rows) or max((len(row) for row in csv_rows), default=0) != max((len(row) for row in markdown_rows), default=0)):
        warnings.append("CSV_MARKDOWN_SHAPE_MISMATCH")
    if batch_summary_row:
        batch_row_count = int(batch_summary_row.get("row_count") or 0)
        batch_col_count = int(batch_summary_row.get("col_count") or 0)
        actual_row_count = len(primary_rows)
        actual_col_count = max((len(row) for row in primary_rows), default=0)
        title_match_row_delta = 1 if actual_row_count > 0 else 0
        row_count_matches = batch_row_count in {actual_row_count, max(0, actual_row_count - title_match_row_delta)}
        col_count_matches = batch_col_count in {actual_col_count, max(0, actual_col_count - 1)}
        if primary_rows and (not row_count_matches or not col_count_matches):
            warnings.append("BATCH_SUMMARY_COUNT_MISMATCH")

    title, title_row_index = detect_table_title(primary_rows, run_meta.get("table_title", ""))
    title_row_index = -1 if title_row_index is None else title_row_index
    header_row = primary_rows[1] if title_row_index == 0 and len(primary_rows) > 1 else (primary_rows[0] if primary_rows else [])
    columns = [normalize_text(value) for value in header_row]
    rows: List[StructTableNormalizedRow] = []
    cells: List[StructTableNormalizedCell] = []
    for row_index, row in enumerate(primary_rows):
        label = normalize_text(row[0]) if row else ""
        values = [normalize_text(value) for value in row[1:]]
        rows.append(StructTableNormalizedRow(row_index=row_index, label=label, values=values, warnings=[]))
        for col_index, value in enumerate(row):
            normalized = normalize_text(value)
            cells.append(
                StructTableNormalizedCell(
                    row_index=row_index,
                    col_index=col_index,
                    text=value,
                    normalized_text=normalized,
                    cell_type_guess=guess_cell_type(normalized),
                    warnings=[],
                )
            )

    returncode_value = run_meta.get("returncode")
    try:
        returncode = int(returncode_value) if returncode_value not in {"", None} else None
    except Exception:
        returncode = None
        warnings.append("RUN_META_RETURNCODE_INVALID")

    raw_response_has_markdown_marker = "MARKDOWN format output" in raw_response_text or "|--" in raw_response_text or ("\n|" in raw_response_text)
    raw_response_timeout_warning_count = raw_response_text.lower().count("timeout")

    return StructTableNormalizedTable(
        tool="structtable_intervl2",
        recognition_source="STRUCTTABLE_MARKDOWN",
        image_name=group.image_name,
        input_image_path=group.input_image_path,
        table_id=Path(group.input_image_path).stem,
        table_title=title,
        columns=columns,
        rows=rows,
        cells=cells,
        provenance={
            "output_folder": group.output_folder,
            "raw_response_path": group.raw_response_path,
            "markdown_path": group.markdown_path,
            "xlsx_path": group.xlsx_path,
            "csv_path": group.csv_path,
            "run_meta_path": group.run_meta_path,
            "error_path": group.error_path,
            "source_format": source_format,
            "run_meta": run_meta,
            "batch_summary_row": batch_summary_row or {},
        },
        warnings=warnings,
        parse_status=parse_status,
        raw_response_text=raw_response_text,
        raw_response_has_markdown_marker=raw_response_has_markdown_marker,
        raw_response_timeout_warning_count=raw_response_timeout_warning_count,
        returncode=returncode,
        error_text=_norm(error_text),
        csv_grid_rows=csv_rows,
        markdown_grid_rows=markdown_rows,
        xlsx_grid_rows=xlsx_rows,
    )
