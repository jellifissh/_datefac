from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {"evidence", "notes", "message", "reason", "excerpt", "text", "error", "action"}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 100)
            if "excerpt" in header or "text" in header or "notes" in header or "action" in header:
                width = min(max(max_len + 2, 24), 110)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    ordered_names = list(sheet_order) if sheet_order is not None else list(sheets.keys())
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in ordered_names:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def batch_report_markdown(batch_summary: Dict[str, Any]) -> str:
    lines = [
        "# Real Test Per-PDF Debug Package 336B",
        "",
        "## Decision",
        f"- decision: {batch_summary.get('decision', '')}",
        f"- qa_fail_count: {batch_summary.get('qa_fail_count', 0)}",
        "",
        "## Totals",
        f"- pdf_found_count: {batch_summary.get('pdf_found_count', 0)}",
        f"- total_candidate_count: {batch_summary.get('total_candidate_count', 0)}",
        f"- total_reviewed_count: {batch_summary.get('total_reviewed_count', 0)}",
        f"- total_needs_review_count: {batch_summary.get('total_needs_review_count', 0)}",
        f"- total_rejected_count: {batch_summary.get('total_rejected_count', 0)}",
        "",
        "## Per Document",
    ]
    for row in batch_summary.get("documents", []):
        lines.append(
            "- {document}: pages={page_count}, tables={table_count}, candidates={candidate_count}, reviewed={reviewed_count}, needs_review={needs_review_count}, rejected={rejected_count}, failure={likely_failure_reason}".format(
                **row
            )
        )
    lines.append("")
    return "\n".join(lines)


def document_report_markdown(document_summary: Dict[str, Any]) -> str:
    likely_pages = document_summary.get("likely_forecast_pages", [])
    page_failures = document_summary.get("page_failures", [])
    lines = [
        f"# Debug Report: {document_summary.get('pdf_filename', '')}",
        "",
        "## Summary",
        f"- page_count: {document_summary.get('page_count', '')}",
        f"- extracted_page_count: {document_summary.get('extracted_page_count', 0)}",
        f"- detected_table_count: {document_summary.get('detected_table_count', 0)}",
        f"- candidate_count: {document_summary.get('candidate_count', 0)}",
        f"- reviewed_count: {document_summary.get('reviewed_count', 0)}",
        f"- needs_review_count: {document_summary.get('needs_review_count', 0)}",
        f"- rejected_count: {document_summary.get('rejected_count', 0)}",
        f"- likely_failure_reason: {document_summary.get('likely_failure_reason', '')}",
        "",
        "## What Was Found",
        "- `extracted_page_text.xlsx` shows per-page text coverage and keyword/year hits.",
        "- `extracted_tables.xlsx` shows every extracted table row with detected metrics, years, and numbers.",
        "- `metric_candidates.xlsx` shows the generated metric candidates before final routing.",
        "- `routing_preview.xlsx` shows why each candidate was routed to reviewed / needs_review / rejected.",
        "",
        "## Likely Forecast Pages",
        f"- pages: {', '.join(str(page) for page in likely_pages) if likely_pages else 'none detected'}",
        "",
        "## Next Action",
        f"- {document_summary.get('recommended_next_action', '')}",
    ]
    if document_summary.get("candidate_count", 0) == 0:
        lines.extend(
            [
                "",
                "## Zero-Candidate Explanation",
                f"- likely_failure_reason: {document_summary.get('likely_failure_reason', '')}",
            ]
        )
    if document_summary.get("table_extraction_error"):
        lines.extend(
            [
                "",
                "## Table Extraction Error",
                f"- {document_summary.get('table_extraction_error', '')}",
            ]
        )
    if page_failures:
        lines.extend(["", "## Page Failures"])
        lines.extend([f"- page {item.get('page', '')}: {item.get('error', '')}" for item in page_failures])
    lines.append("")
    return "\n".join(lines)
