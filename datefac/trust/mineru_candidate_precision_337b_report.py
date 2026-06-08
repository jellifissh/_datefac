from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


CUSTOMER_WORKBOOK_SHEETS = [
    "00_README",
    "01_REVIEWED_CORE_METRICS",
    "02_NEEDS_REVIEW",
    "03_REJECTED_OR_EXCLUDED",
    "04_SOURCE_TRACE",
    "05_DOCUMENT_SUMMARY",
    "06_TABLE_CLASSIFICATION_SUMMARY",
]

BEFORE_AFTER_SHEETS = [
    "00_SUMMARY",
    "01_BEFORE_COUNTS",
    "02_AFTER_COUNTS",
    "03_REVIEWED_AFTER",
    "04_NEEDS_REVIEW_AFTER",
    "05_REJECTED_AFTER",
    "06_DUPLICATE_TABLES_REMOVED",
    "07_TABLE_ROLE_CLASSIFICATION",
    "08_ROUTE_CHANGE_TRACE",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {"evidence", "notes", "message", "reason", "excerpt", "preview", "trace", "warning", "explanation"}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 110)
            if any(token in header for token in {"excerpt", "notes", "message", "preview", "trace"}):
                width = min(max(max_len + 2, 24), 120)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# MinerU Candidate Precision Calibration 337B",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Before / After",
        f"- reviewed_before: {summary.get('reviewed_before_count', 0)}",
        f"- reviewed_after: {summary.get('reviewed_after_count', 0)}",
        f"- needs_review_before: {summary.get('needs_review_before_count', 0)}",
        f"- needs_review_after: {summary.get('needs_review_after_count', 0)}",
        f"- rejected_before: {summary.get('rejected_before_count', 0)}",
        f"- rejected_after: {summary.get('rejected_after_count', 0)}",
        "",
        "## Calibration Effects",
        f"- duplicate_table_removed_count: {summary.get('duplicate_table_removed_count', 0)}",
        f"- excluded_rating_standard_table_count: {summary.get('excluded_rating_standard_table_count', 0)}",
        f"- excluded_legal_disclosure_table_count: {summary.get('excluded_legal_disclosure_table_count', 0)}",
        f"- excluded_company_profile_table_count: {summary.get('excluded_company_profile_table_count', 0)}",
        "",
        "## Safety",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_official_asset_modification_during_337b: {summary.get('no_official_asset_modification_during_337b', False)}",
        "",
        "## QA",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    lines.append("")
    return "\n".join(lines)
