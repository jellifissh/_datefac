from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEET_ORDER = [
    "00_README",
    "01_REVIEWED_CORE_METRICS",
    "02_NEEDS_REVIEW",
    "03_REJECTED_OR_EXCLUDED",
    "04_SOURCE_TRACE",
    "05_RUN_SUMMARY",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {"evidence", "notes", "message", "reason", "excerpt", "error"}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 90)
            if "excerpt" in header or "notes" in header or "message" in header:
                width = min(max(max_len + 2, 24), 100)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in WORKBOOK_SHEET_ORDER:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def real_test_full_flow_336a_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    pdf_failures = qa_json.get("pdf_failures", [])
    page_failures = qa_json.get("page_failures", [])
    blocked_reasons = qa_json.get("blocked_reasons", [])
    lines = [
        "# Real Test Full Flow 336A",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Run Counts",
        f"- pdf_found_count: {summary.get('pdf_found_count', 0)}",
        f"- pdf_processed_count: {summary.get('pdf_processed_count', 0)}",
        f"- reviewed_count: {summary.get('reviewed_count', 0)}",
        f"- needs_review_count: {summary.get('needs_review_count', 0)}",
        f"- rejected_or_excluded_count: {summary.get('rejected_or_excluded_count', 0)}",
        "",
        "## Safety",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_official_asset_modification_during_336a: {summary.get('no_official_asset_modification_during_336a', False)}",
        "",
        "## Failures",
        f"- pdf_failure_count: {len(pdf_failures)}",
        f"- page_failure_count: {len(page_failures)}",
    ]
    if blocked_reasons:
        lines.extend(["", "## Blocked Reasons"])
        lines.extend([f"- {reason}" for reason in blocked_reasons])
    if pdf_failures:
        lines.extend(["", "## PDF Failures"])
        lines.extend([f"- {item.get('document', '')}: {item.get('stage', '')} -> {item.get('error', '')}" for item in pdf_failures[:20]])
    if page_failures:
        lines.extend(["", "## Page Failures"])
        lines.extend([f"- {item.get('document', '')} page {item.get('page', '')}: {item.get('error', '')}" for item in page_failures[:20]])
    lines.append("")
    return "\n".join(lines)
