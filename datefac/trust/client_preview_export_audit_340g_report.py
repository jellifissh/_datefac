from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_AUDIT_SUMMARY",
    "02_CORE_METRIC_AUDIT",
    "03_UNIT_AUDIT",
    "04_DUPLICATE_AUDIT",
    "05_SOURCE_TRACE_AUDIT",
    "06_NEEDS_REVIEW_AUDIT",
    "07_REJECTED_AUDIT",
    "08_CLAIMS_AUDIT",
    "09_NO_WRITE_BACK_PROOF",
    "10_NEXT_STEP_RECOMMENDATION",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "evidence",
        "notes",
        "status",
        "warning",
        "reference",
        "hash",
        "action",
        "decision",
        "risk",
        "recommendation",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 160)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 180)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# Client Preview Export Audit 340G",
        "",
        "340G audits the 340F client preview workbook for preview suitability, traceability, and claim safety.",
        "A passed audit still does not mean production-ready delivery.",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- client_preview_audit_passed: {summary.get('client_preview_audit_passed', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        "",
        "## Counts",
        f"- audited_core_metric_count: {summary.get('audited_core_metric_count', 0)}",
        f"- confirmed_count: {summary.get('confirmed_count', 0)}",
        f"- corrected_count: {summary.get('corrected_count', 0)}",
        f"- needs_review_count: {summary.get('needs_review_count', 0)}",
        f"- rejected_count: {summary.get('rejected_count', 0)}",
        f"- duplicate_issue_count: {summary.get('duplicate_issue_count', 0)}",
        f"- unit_issue_count: {summary.get('unit_issue_count', 0)}",
        f"- missing_source_trace_count: {summary.get('missing_source_trace_count', 0)}",
        f"- unsafe_claim_count: {summary.get('unsafe_claim_count', 0)}",
        "",
        "## Safety",
        f"- no_write_back_proof_passed: {summary.get('no_write_back_proof_passed', False)}",
        f"- output_workbook_path: {summary.get('output_workbook_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    lines.append("")
    return "\n".join(lines)
