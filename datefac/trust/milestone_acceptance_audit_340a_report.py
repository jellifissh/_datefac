from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_AUDIT_SUMMARY",
    "02_INPUT_PDF_AUDIT",
    "03_OUTPUT_ARTIFACT_AUDIT",
    "04_KEY_METRIC_AUDIT",
    "05_337D_REVIEWED_SAMPLE",
    "06_338D_AI_ADOPTION_AUDIT",
    "07_DOC_CONSISTENCY_AUDIT",
    "08_UNSAFE_CLAIM_AUDIT",
    "09_QA_CHECKS",
    "10_NEXT_STEP",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "excerpt",
        "evidence",
        "notes",
        "docs",
        "matched",
        "line",
        "warning",
        "placeholder",
    }
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 130)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# Milestone Acceptance Audit 340A",
        "",
        "## Judgment",
        f"- milestone_judgment: {summary.get('milestone_judgment', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- suitable_for_demo_research_preview: {summary.get('suitable_for_demo_research_preview', False)}",
        "- not client-ready",
        "- not production-ready",
        "- AI decisions are dry-run only",
        "- human review remains necessary",
        "",
        "## Pipeline Checks",
        f"- input_pdf_count: {summary.get('input_pdf_count', 0)}",
        f"- expected_pdf_present_count: {summary.get('expected_pdf_present_count', 0)}",
        f"- required_pipeline_output_file_count: {summary.get('required_pipeline_output_file_count', 0)}",
        f"- existing_pipeline_output_file_count: {summary.get('existing_pipeline_output_file_count', 0)}",
        "",
        "## Key Metrics",
        f"- 337A parsed_pdf_count: {summary.get('parsed_pdf_count_337a', 0)}",
        f"- 337B reviewed_after_count: {summary.get('reviewed_after_count_337b', 0)}",
        f"- 337C reviewed_after_count: {summary.get('reviewed_after_count_337c', 0)}",
        f"- 337D reviewed_after_count: {summary.get('reviewed_after_count_337d', 0)}",
        f"- 338D input_row_count: {summary.get('input_row_count_338d', 0)}",
        f"- 338D accept_model_confirm_count: {summary.get('accept_model_confirm_count_338d', 0)}",
        f"- 338D accept_model_reject_count: {summary.get('accept_model_reject_count_338d', 0)}",
        f"- 338D hold_for_human_review_count: {summary.get('hold_for_human_review_count_338d', 0)}",
        f"- 338D invalid_model_response_count: {summary.get('invalid_model_response_count_338d', 0)}",
        f"- 338D deterministic_rule_override_count: {summary.get('deterministic_rule_override_count_338d', 0)}",
        "",
        "## Audit Sheets",
        f"- 337D reviewed sample rows: {summary.get('sample_reviewed_row_count_337d', 0)}",
        f"- 338D AI adoption audit rows: {summary.get('ai_adoption_audit_row_count', 0)}",
        "",
        "## Documentation Audit",
        f"- documentation_consistency_passed: {summary.get('documentation_consistency_passed', False)}",
        f"- unsafe_claim_audit_passed: {summary.get('unsafe_claim_audit_passed', False)}",
        "",
        "## Recommendation",
        f"- next_step_recommendation: {summary.get('next_step_recommendation', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    lines.append("")
    return "\n".join(lines)
