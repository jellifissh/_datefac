from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_AB_SUMMARY",
    "02_NEW_MODEL_ADJUDICATION_PLAN",
    "03_DEEPSEEK_338A_BASELINE",
    "04_ROW_LEVEL_COMPARISON",
    "05_CHANGED_DECISIONS",
    "06_INVALID_OR_LOW_CONFIDENCE",
    "07_RULE_MODEL_CONFLICTS",
    "08_PROMPT_CONTEXT_UPGRADE",
    "09_CACHE_AND_COST_SUMMARY",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {"evidence", "notes", "message", "reason", "excerpt", "prompt", "context", "risk", "action", "quote"}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 110)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 120)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# AI Review Model A/B Evaluation 338B",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- recommendation: {summary.get('recommendation', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Runtime",
        f"- env_source: {summary.get('env_source', '')}",
        f"- api_env_ready: {summary.get('api_env_ready', False)}",
        f"- baseline_model_name: {summary.get('baseline_model_name', '')}",
        f"- new_model_name: {summary.get('new_model_name', '')}",
        f"- row_count: {summary.get('row_count', 0)}",
        f"- api_call_count: {summary.get('api_call_count', 0)}",
        f"- cache_hit_count: {summary.get('cache_hit_count', 0)}",
        "",
        "## Comparison",
        f"- low_confidence_count: {summary.get('low_confidence_count_baseline', 0)} -> {summary.get('low_confidence_count_new', 0)}",
        f"- needs_more_context_count: {summary.get('needs_more_context_count_baseline', 0)} -> {summary.get('needs_more_context_count_new', 0)}",
        f"- confirm_reviewed_count: {summary.get('confirm_reviewed_count_baseline', 0)} -> {summary.get('confirm_reviewed_count_new', 0)}",
        f"- downgrade_count: {summary.get('downgrade_count_baseline', 0)} -> {summary.get('downgrade_count_new', 0)}",
        f"- reject_count: {summary.get('reject_count_baseline', 0)} -> {summary.get('reject_count_new', 0)}",
        f"- invalid_response_count: {summary.get('invalid_response_count_baseline', 0)} -> {summary.get('invalid_response_count_new', 0)}",
        f"- evidence_quote_valid_count: {summary.get('evidence_quote_valid_count', 0)}",
        f"- evidence_quote_invalid_count: {summary.get('evidence_quote_invalid_count', 0)}",
        "",
        "## QA",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    lines.append("")
    return "\n".join(lines)
