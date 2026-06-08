from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable, List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


COMBINED_WORKBOOK_SHEETS = [
    "00_README",
    "01_REVIEWED_CORE_METRICS",
    "02_NEEDS_REVIEW",
    "03_REJECTED_OR_EXCLUDED",
    "04_SOURCE_TRACE",
    "05_DOCUMENT_SUMMARY",
    "06_FINANCIAL_TABLE_CANDIDATES",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {"evidence", "notes", "message", "reason", "excerpt", "text", "command", "preview", "warning"}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 110)
            if any(token in header for token in {"excerpt", "notes", "message", "command", "preview"}):
                width = min(max(max_len + 2, 24), 120)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str] | None = None) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    ordered_names = list(sheet_order) if sheet_order is not None else list(sheets.keys())
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in ordered_names:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def batch_report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any], document_rows: List[Dict[str, Any]]) -> str:
    lines = [
        "# MinerU First Real PDF Intake 337A",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Run Counts",
        f"- pdf_found_count: {summary.get('pdf_found_count', 0)}",
        f"- pdf_processed_count: {summary.get('pdf_processed_count', 0)}",
        f"- mineru_success_count: {summary.get('mineru_success_count', 0)}",
        f"- reviewed_count: {summary.get('reviewed_count', 0)}",
        f"- needs_review_count: {summary.get('needs_review_count', 0)}",
        f"- rejected_or_excluded_count: {summary.get('rejected_or_excluded_count', 0)}",
        "",
        "## Safety",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_official_asset_modification_during_337a: {summary.get('no_official_asset_modification_during_337a', False)}",
        "",
        "## Per PDF",
    ]
    for row in document_rows:
        lines.append(
            "- {document}: pages={page_count}, mineru_tables={mineru_table_count}, financial_candidates={financial_table_candidate_count}, metrics={metric_candidate_count}, reviewed={reviewed_count}, needs_review={needs_review_count}, rejected={rejected_count}, status={parse_status}".format(
                **row
            )
        )
    blocked_reasons = qa_json.get("blocked_reasons", [])
    if blocked_reasons:
        lines.extend(["", "## Blocked Reasons"])
        lines.extend([f"- {reason}" for reason in blocked_reasons])
    manual_commands = summary.get("manual_mineru_commands", [])
    if manual_commands:
        lines.extend(["", "## Manual MinerU Commands"])
        lines.extend([f"- `{command}`" for command in manual_commands])
    lines.append("")
    return "\n".join(lines)


def document_report_markdown(document_summary: Dict[str, Any]) -> str:
    likely_pages = document_summary.get("likely_financial_pages", [])
    lines = [
        f"# Debug Report: {document_summary.get('document', '')}",
        "",
        "## Summary",
        f"- parse_status: {document_summary.get('parse_status', '')}",
        f"- page_count: {document_summary.get('page_count', '')}",
        f"- mineru_table_count: {document_summary.get('mineru_table_count', 0)}",
        f"- financial_table_candidate_count: {document_summary.get('financial_table_candidate_count', 0)}",
        f"- metric_candidate_count: {document_summary.get('metric_candidate_count', 0)}",
        f"- reviewed_count: {document_summary.get('reviewed_count', 0)}",
        f"- needs_review_count: {document_summary.get('needs_review_count', 0)}",
        f"- rejected_count: {document_summary.get('rejected_count', 0)}",
        "",
        "## Likely Financial Pages",
        f"- pages: {', '.join(str(page) for page in likely_pages) if likely_pages else 'none detected'}",
        "",
        "## MinerU",
        f"- parse_dir: {document_summary.get('mineru_parse_dir', '')}",
        f"- manual_command: {document_summary.get('manual_mineru_command', '')}",
        f"- failure_reason: {document_summary.get('failure_reason', '')}",
        "",
        "## Notes",
        "- `mineru_artifact_inventory.xlsx` shows raw MinerU source files, table assets, and warnings.",
        "- `extracted_page_text.xlsx` shows per-page text evidence and keyword hits.",
        "- `financial_table_candidates.xlsx` ranks tables that look most like financial forecast tables.",
        "- `metric_candidates.xlsx` and `routing_preview.xlsx` explain what was extracted and why it was routed.",
        "",
    ]
    return "\n".join(lines)
