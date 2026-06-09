from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_REVIEW_QUEUE",
    "02_HOLD_FOR_HUMAN_REVIEW",
    "03_INVALID_MODEL_RESPONSES",
    "04_REJECTED_BY_RULE_FOR_CHECK",
    "05_ACCEPTED_CONFIRM_SPOT_CHECK",
    "06_ACCEPTED_REJECT_SPOT_CHECK",
    "07_SOURCE_TRACE_CONTEXT",
    "08_REVIEW_GUIDE",
    "09_SUMMARY",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "excerpt",
        "evidence",
        "quote",
        "notes",
        "context",
        "action",
        "risk",
        "guide",
    }
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 120)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 140)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# Human Review Package After AI Adoption 340B",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        f"- no_write_back: {summary.get('no_write_back', False)}",
        "",
        "## Workbook",
        f"- review_workbook_path: {summary.get('review_workbook_path', '')}",
        f"- total_review_queue_count: {summary.get('total_review_queue_count', 0)}",
        f"- hold_for_human_count: {summary.get('hold_for_human_count', 0)}",
        f"- invalid_model_response_count: {summary.get('invalid_model_response_count', 0)}",
        f"- rejected_by_rule_check_count: {summary.get('rejected_by_rule_check_count', 0)}",
        f"- accepted_confirm_spot_check_count: {summary.get('accepted_confirm_spot_check_count', 0)}",
        f"- accepted_reject_spot_check_count: {summary.get('accepted_reject_spot_check_count', 0)}",
        "",
        "## Source Counts",
        f"- source_337d_reviewed_count: {summary.get('source_337d_reviewed_count', 0)}",
        f"- source_338d_accept_confirm_count: {summary.get('source_338d_accept_confirm_count', 0)}",
        f"- source_338d_accept_reject_count: {summary.get('source_338d_accept_reject_count', 0)}",
        f"- source_338d_hold_count: {summary.get('source_338d_hold_count', 0)}",
        f"- source_338d_invalid_count: {summary.get('source_338d_invalid_count', 0)}",
        "",
        "## Safety",
        f"- upstream_workbooks_unchanged: {summary.get('upstream_workbooks_unchanged', False)}",
        f"- no_apply_proof_generated: {summary.get('no_apply_proof_generated', False)}",
        f"- reviewer_fields_present: {summary.get('reviewer_fields_present', False)}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    lines.append("")
    return "\n".join(lines)
