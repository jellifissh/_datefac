from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_ADJUDICATION_SUMMARY",
    "02_MODEL_ADJUDICATION_PLAN",
    "03_PROMPT_PREVIEW",
    "04_INVALID_OR_LOW_CONFIDENCE",
    "05_RULE_MODEL_CONFLICTS",
    "06_COST_AND_CACHE_SUMMARY",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {"evidence", "notes", "message", "reason", "excerpt", "preview", "prompt", "context", "risk", "action"}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 110)
            if any(token in header for token in {"evidence", "notes", "message", "preview", "prompt", "context", "reason", "risk", "action"}):
                width = min(max(max_len + 2, 24), 120)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# DeepSeek Text Adjudicator 338A",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Runtime",
        f"- api_env_ready: {summary.get('api_env_ready', False)}",
        f"- model_name: {summary.get('model_name', '')}",
        f"- dry_run_prompts_only: {summary.get('dry_run_prompts_only', False)}",
        f"- adjudication_row_count: {summary.get('adjudication_row_count', 0)}",
        f"- cache_hit_count: {summary.get('cache_hit_count', 0)}",
        "",
        "## Decisions",
        f"- confirm_reviewed_count: {summary.get('confirm_reviewed_count', 0)}",
        f"- downgrade_to_needs_review_count: {summary.get('downgrade_to_needs_review_count', 0)}",
        f"- reject_count: {summary.get('reject_count', 0)}",
        f"- needs_more_context_count: {summary.get('needs_more_context_count', 0)}",
        f"- invalid_response_count: {summary.get('invalid_response_count', 0)}",
        f"- low_confidence_count: {summary.get('low_confidence_count', 0)}",
        f"- rule_model_conflict_count: {summary.get('rule_model_conflict_count', 0)}",
        "",
        "## QA",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    lines.append("")
    return "\n".join(lines)
