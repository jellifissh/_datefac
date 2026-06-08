from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


CUSTOMER_WORKBOOK_SHEETS = [
    "00_README",
    "01_REVIEWED_CORE_METRICS",
    "02_NEEDS_REVIEW",
    "03_REJECTED_OR_EXCLUDED",
    "04_SOURCE_TRACE",
    "05_DOCUMENT_SUMMARY",
    "06_TABLE_CLASSIFICATION_SUMMARY",
    "07_CONTEXT_REPAIR_SUMMARY",
    "08_SUSPICIOUS_REVIEWED_AUDIT",
    "09_ROUTE_CHANGE_TRACE",
]

BEFORE_AFTER_SHEETS = [
    "00_SUMMARY",
    "01_337C_COUNTS",
    "02_337D_COUNTS",
    "03_YEAR_ALIGNMENT_ACTIONS",
    "04_PERCENT_AMOUNT_GUARD",
    "05_UNIT_STRICTNESS",
    "06_REVIEWED_DEDUP",
    "07_REVIEWED_AFTER_337D",
    "08_NEEDS_REVIEW_AFTER_337D",
    "09_REJECTED_AFTER_337D",
    "10_SUSPICIOUS_REVIEWED_AUDIT",
    "11_ROUTE_CHANGE_TRACE",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {"evidence", "notes", "message", "reason", "excerpt", "preview", "trace", "warning", "summary", "action"}
    for sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 110)
            if any(token in header for token in {"excerpt", "notes", "message", "preview", "trace", "evidence", "reason", "action"}):
                width = min(max(max_len + 2, 24), 120)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# Reviewed Strictness, Year Alignment, and Suspicious Row QA 337D",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        "",
        "## Before / After",
        f"- reviewed_before_337c: {summary.get('reviewed_before_count', 0)}",
        f"- reviewed_after_337d: {summary.get('reviewed_after_count', 0)}",
        f"- needs_review_after_337d: {summary.get('needs_review_after_count', 0)}",
        f"- rejected_after_337d: {summary.get('rejected_after_count', 0)}",
        "",
        "## Strictness Actions",
        f"- year_alignment_repaired_count: {summary.get('year_alignment_repaired_count', 0)}",
        f"- year_alignment_downgraded_count: {summary.get('year_alignment_downgraded_count', 0)}",
        f"- percent_amount_guard_remapped_count: {summary.get('percent_amount_guard_remapped_count', 0)}",
        f"- percent_amount_guard_downgraded_count: {summary.get('percent_amount_guard_downgraded_count', 0)}",
        f"- unit_strictness_filled_count: {summary.get('unit_strictness_filled_count', 0)}",
        f"- unit_strictness_downgraded_count: {summary.get('unit_strictness_downgraded_count', 0)}",
        f"- reviewed_duplicate_removed_count: {summary.get('reviewed_duplicate_removed_count', 0)}",
        "",
        "## 356439 Audit",
        f"- reviewed_356439_before_count: {summary.get('reviewed_356439_before_count', 0)}",
        f"- reviewed_356439_after_count: {summary.get('reviewed_356439_after_count', 0)}",
        f"- reviewed_356439_downgraded_count: {summary.get('reviewed_356439_downgraded_count', 0)}",
        "",
        "## QA",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    lines.append("")
    return "\n".join(lines)
