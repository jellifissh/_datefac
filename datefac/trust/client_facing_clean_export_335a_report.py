from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEET_ORDER = [
    "00_README_FOR_CUSTOMER",
    "01_CORE_METRICS_REVIEWED",
    "02_NEEDS_REVIEW",
    "03_EXCLUDED_OR_REJECTED",
    "04_SOURCE_TRACE",
    "05_DELIVERY_SUMMARY",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {"evidence", "notes", "message", "reason", "value"}
    for sheet_name in workbook.sheetnames:
        ws = workbook[sheet_name]
        ws.freeze_panes = "A2"
        if ws.max_row >= 1 and ws.max_column >= 1:
            ws.auto_filter.ref = ws.dimensions
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center")

        for column_cells in ws.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                if len(value) > max_len:
                    max_len = len(value)
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 80)
            if "evidence" in header:
                width = min(max(max_len + 2, 24), 100)
            ws.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in WORKBOOK_SHEET_ORDER:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def client_facing_clean_export_335a_markdown(summary: Dict[str, Any]) -> str:
    return "\n".join(
        [
            "# Client-Facing Clean Export 335A",
            "",
            "## Decision",
            f"- decision: {summary.get('decision', '')}",
            "",
            "## Preview State",
            f"- project_status: {summary.get('project_status', '')}",
            f"- client_facing_preview: {summary.get('client_facing_preview', False)}",
            f"- client_ready: {summary.get('client_ready', False)}",
            f"- production_ready: {summary.get('production_ready', False)}",
            "",
            "## Customer Workbook Counts",
            f"- source_reviewed_trusted_preview_row_count: {summary.get('source_reviewed_trusted_preview_row_count', 0)}",
            f"- core_metrics_reviewed_row_count: {summary.get('core_metrics_reviewed_row_count', 0)}",
            f"- needs_review_row_count: {summary.get('needs_review_row_count', 0)}",
            f"- excluded_or_rejected_row_count: {summary.get('excluded_or_rejected_row_count', 0)}",
            f"- source_trace_row_count: {summary.get('source_trace_row_count', 0)}",
            "",
            "## Safety",
            f"- source_page_missing_count: {summary.get('source_page_missing_count', 0)}",
            f"- no_official_asset_modification_during_335a: {summary.get('no_official_asset_modification_during_335a', False)}",
            f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
            "",
        ]
    )
