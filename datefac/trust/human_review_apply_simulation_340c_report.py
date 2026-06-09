from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Dict, Iterable

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font


WORKBOOK_SHEETS = [
    "00_README",
    "01_APPLY_PLAN",
    "02_FILLED_REVIEW_ROWS",
    "03_PENDING_REVIEW_ROWS",
    "04_VALIDATION_WARNINGS",
    "05_NO_APPLY_PROOF",
    "06_SUMMARY",
]


def _to_jsonable(value: Any) -> Any:
    if isinstance(value, dict):
        return {str(key): _to_jsonable(item) for key, item in value.items()}
    if isinstance(value, list):
        return [_to_jsonable(item) for item in value]
    if isinstance(value, tuple):
        return [_to_jsonable(item) for item in value]
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value


def write_json(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(json.dumps(_to_jsonable(payload), ensure_ascii=False, indent=2), encoding="utf-8")


def _format_workbook(path: Path) -> None:
    workbook = load_workbook(path)
    wrap_keywords = {
        "message",
        "detail",
        "reason",
        "evidence",
        "notes",
        "status",
        "warning",
        "reference",
        "hash",
        "action",
        "decision",
    }
    for worksheet in workbook.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row >= 1 and worksheet.max_column >= 1:
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        for column_cells in worksheet.columns:
            header = str(column_cells[0].value or "").strip().lower()
            max_len = len(header)
            for cell in column_cells[1:]:
                value = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(value))
                if any(token in header for token in wrap_keywords):
                    cell.alignment = Alignment(vertical="top", wrap_text=True)
                else:
                    cell.alignment = Alignment(vertical="top")
            width = min(max(max_len + 2, 12), 140)
            if any(token in header for token in wrap_keywords):
                width = min(max(max_len + 2, 24), 160)
            worksheet.column_dimensions[column_cells[0].column_letter].width = width
    workbook.save(path)


def write_excel(path: Path, sheets: Dict[str, pd.DataFrame], sheet_order: Iterable[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name in sheet_order:
            sheets.get(sheet_name, pd.DataFrame()).to_excel(writer, sheet_name=sheet_name, index=False)
    _format_workbook(path)


def report_markdown(summary: Dict[str, Any], qa_json: Dict[str, Any]) -> str:
    lines = [
        "# Human Review Apply Simulation After AI Adoption 340C",
        "",
        "## Decision",
        f"- decision: {summary.get('decision', '')}",
        f"- qa_fail_count: {summary.get('qa_fail_count', 0)}",
        f"- validation_warning_count: {summary.get('validation_warning_count', 0)}",
        f"- no_write_back: {summary.get('no_write_back', False)}",
        f"- client_ready: {summary.get('client_ready', False)}",
        f"- production_ready: {summary.get('production_ready', False)}",
        "",
        "## Counts",
        f"- total_review_queue_count: {summary.get('total_review_queue_count', 0)}",
        f"- filled_review_row_count: {summary.get('filled_review_row_count', 0)}",
        f"- pending_review_row_count: {summary.get('pending_review_row_count', 0)}",
        f"- confirm_as_reviewed_count: {summary.get('confirm_as_reviewed_count', 0)}",
        f"- correct_and_confirm_count: {summary.get('correct_and_confirm_count', 0)}",
        f"- keep_needs_review_count: {summary.get('keep_needs_review_count', 0)}",
        f"- reject_count: {summary.get('reject_count', 0)}",
        f"- needs_more_context_count: {summary.get('needs_more_context_count', 0)}",
        "",
        "## Safety",
        f"- review_workbook_unchanged: {summary.get('review_workbook_unchanged', False)}",
        f"- upstream_workbooks_unchanged: {summary.get('upstream_workbooks_unchanged', False)}",
        f"- no_apply_proof_passed: {summary.get('no_apply_proof_passed', False)}",
        f"- apply_plan_workbook_path: {summary.get('apply_plan_workbook_path', '')}",
        "",
        "## QA Checks",
    ]
    for check in qa_json.get("checks", []):
        lines.append(f"- {check.get('check_name', '')}: {check.get('status', '')} ({check.get('detail', '')})")
    lines.append("")
    return "\n".join(lines)
