import argparse
import hashlib
import json
import shutil
import subprocess
import sys
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import load_workbook


CURRENT_DIR = Path(__file__).resolve().parent
PROJECT_ROOT = CURRENT_DIR.parent
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))


BASE_DIR = Path(r"D:\_datefac")
OUTPUT_DIR = BASE_DIR / "output"
DELIVERY_DIR = OUTPUT_DIR / "delivery_package"

INPUT_STAGE5R_DIR = OUTPUT_DIR / "stage5r_final_apply_plan"
INPUT_MANIFEST_XLSX = INPUT_STAGE5R_DIR / "160_stage5r_apply_manifest.xlsx"
INPUT_RISK_REVIEW_XLSX = INPUT_STAGE5R_DIR / "160_stage5r_apply_risk_review.xlsx"
INPUT_BACKUP_ROLLBACK_MD = INPUT_STAGE5R_DIR / "160_stage5r_backup_rollback_plan.md"
INPUT_STAGE5R_SUMMARY_JSON = INPUT_STAGE5R_DIR / "161_stage5r_final_apply_plan_summary.json"

INPUT_STAGE5O_DIR = OUTPUT_DIR / "stage5o_promotion_review"
INPUT_STAGE5O_CANDIDATE_02_XLSX = INPUT_STAGE5O_DIR / "154_stage5o_candidate_02.xlsx"
INPUT_STAGE5O_CANDIDATE_05_XLSX = INPUT_STAGE5O_DIR / "154_stage5o_candidate_05.xlsx"

OFFICIAL_02B_PATH = BASE_DIR / "data" / "overrides" / "02B_ai_repair_override.xlsx"
FORMAL_SCOPE_RULES_JSON = BASE_DIR / "data" / "mapping" / "formal_scope_rules.json"
FORMAL_MAPPING_RULE_FILE = FORMAL_SCOPE_RULES_JSON
FORMAL_NORMALIZATION_RULE_FILE = BASE_DIR / "financial_standardizer.py"
FORMAL_ALIAS_RULE_FILE = BASE_DIR / "financial_standardizer.py"

OUT_DIR = OUTPUT_DIR / "stage5s_real_apply"
OUT_BACKUP_DIR = OUT_DIR / "backup"
OUT_APPLY_LOG_XLSX = OUT_DIR / "162_stage5s_apply_log.xlsx"
OUT_APPLY_DIFF_XLSX = OUT_DIR / "162_stage5s_apply_diff.xlsx"
OUT_APPLY_REPORT_MD = OUT_DIR / "162_stage5s_apply_report.md"
OUT_SUMMARY_JSON = OUT_DIR / "163_stage5s_real_apply_summary.json"

SHEET_APPLIED_02 = "stage5s_applied_02_long"
SHEET_APPLIED_05 = "stage5s_applied_05_long"


def _norm(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    return str(v).strip()


def _sha256(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _find_delivery_file(pattern: str) -> Path:
    files = sorted(DELIVERY_DIR.glob(pattern))
    if not files:
        raise FileNotFoundError(f"Missing delivery file pattern: {pattern}")
    non_copy = [p for p in files if "_copy_" not in p.name.lower()]
    return non_copy[0] if non_copy else files[0]


def _snapshot_hashes() -> Dict[str, str]:
    return {
        "01": _sha256(_find_delivery_file("01_*.xlsx")),
        "02": _sha256(_find_delivery_file("02_*.xlsx")),
        "02A": _sha256(_find_delivery_file("02A_*.xlsx")),
        "05": _sha256(_find_delivery_file("05_*.xlsx")),
        "06": _sha256(_find_delivery_file("06_*.xlsx")),
        "02B": _sha256(OFFICIAL_02B_PATH),
        "formal_scope_rules": _sha256(FORMAL_SCOPE_RULES_JSON),
        "formal_mapping_rules": _sha256(FORMAL_MAPPING_RULE_FILE),
        "formal_normalization_rules": _sha256(FORMAL_NORMALIZATION_RULE_FILE),
        "formal_alias_rules": _sha256(FORMAL_ALIAS_RULE_FILE),
    }


def _run_delivery_check() -> str:
    script = BASE_DIR / "tools" / "check_delivery_state.py"
    p = subprocess.run([sys.executable, str(script), "--json"], capture_output=True, text=True, check=False)
    if p.returncode != 0:
        return "FAIL"
    text = (p.stdout or "").strip()
    if not text:
        return "UNKNOWN"
    try:
        payload = json.loads(text)
        return _norm(payload.get("overall_status")) or "UNKNOWN"
    except Exception:
        return "UNKNOWN"


def _copy_backup(src: Path, dst: Path) -> str:
    dst.parent.mkdir(parents=True, exist_ok=True)
    shutil.copy2(src, dst)
    return _sha256(dst)


def _write_dataframe_to_sheet(workbook_path: Path, sheet_name: str, df: pd.DataFrame) -> None:
    wb = load_workbook(workbook_path)
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        wb.remove(ws)
    ws = wb.create_sheet(sheet_name)
    cols = list(df.columns)
    ws.append(cols)
    for row in df.itertuples(index=False, name=None):
        ws.append(list(row))
    wb.save(workbook_path)


def _write_excel(path: Path, sheets: Dict[str, pd.DataFrame]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, sheet_df in sheets.items():
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)


def _duplicate_and_conflict_counts_02(df: pd.DataFrame) -> Tuple[int, int, int, int]:
    if df.empty:
        return 0, 0, 0, 0
    work = df.copy().fillna("")
    work["key"] = (
        work["metric_level_key"].map(_norm)
        + "||"
        + work["year"].map(_norm)
    )
    dup = int(work["key"].duplicated().sum())

    val_conf = 0
    unit_conf = 0
    year_conf = 0
    for _, g in work.groupby("metric_level_key"):
        if g["year"].nunique() != len(g):
            year_conf += 1
    for _, g in work.groupby("key"):
        if g["value"].map(_norm).nunique() > 1:
            val_conf += 1
        if g["unit"].map(_norm).nunique() > 1:
            unit_conf += 1
    return dup, val_conf, unit_conf, year_conf


def _duplicate_and_conflict_counts_05(df: pd.DataFrame) -> Tuple[int, int, int, int]:
    if df.empty:
        return 0, 0, 0, 0
    work = df.copy().fillna("")
    work["key"] = (
        work["metric_level_key"].map(_norm)
        + "||"
        + work["year"].map(_norm)
    )
    dup = int(work["key"].duplicated().sum())
    val_conf = 0
    unit_conf = 0
    year_conf = 0
    for _, g in work.groupby("metric_level_key"):
        if g["year"].nunique() != len(g):
            year_conf += 1
    for _, g in work.groupby("key"):
        if g["value"].map(_norm).nunique() > 1:
            val_conf += 1
        if g["unit"].map(_norm).nunique() > 1:
            unit_conf += 1
    return dup, val_conf, unit_conf, year_conf


def _rollback(prod02: Path, prod05: Path, backup02: Path, backup05: Path) -> None:
    shutil.copy2(backup02, prod02)
    shutil.copy2(backup05, prod05)


def main() -> int:
    parser = argparse.ArgumentParser(description="Stage5S real apply production 02/05 using Stage5R manifest.")
    parser.parse_args()

    required = [
        INPUT_MANIFEST_XLSX,
        INPUT_RISK_REVIEW_XLSX,
        INPUT_BACKUP_ROLLBACK_MD,
        INPUT_STAGE5R_SUMMARY_JSON,
        INPUT_STAGE5O_CANDIDATE_02_XLSX,
        INPUT_STAGE5O_CANDIDATE_05_XLSX,
        OFFICIAL_02B_PATH,
        FORMAL_SCOPE_RULES_JSON,
        FORMAL_MAPPING_RULE_FILE,
        FORMAL_NORMALIZATION_RULE_FILE,
        FORMAL_ALIAS_RULE_FILE,
    ]
    for p in required:
        if not p.exists():
            raise FileNotFoundError(f"Missing required input: {p}")

    stage5r_summary = json.loads(INPUT_STAGE5R_SUMMARY_JSON.read_text(encoding="utf-8"))
    if not bool(stage5r_summary.get("ready_for_stage5s_real_apply")):
        raise RuntimeError("Precondition failed: ready_for_stage5s_real_apply is not true.")
    if int(stage5r_summary.get("blocked_metric_count", -1)) != 0:
        raise RuntimeError("Precondition failed: blocked_metric_count is not 0.")

    production_02_file = Path(_norm(stage5r_summary.get("production_02_reference_file")))
    production_05_file = Path(_norm(stage5r_summary.get("production_05_reference_file")))
    if not production_02_file.exists():
        raise FileNotFoundError(f"Production 02 file missing: {production_02_file}")
    if not production_05_file.exists():
        raise FileNotFoundError(f"Production 05 file missing: {production_05_file}")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    OUT_BACKUP_DIR.mkdir(parents=True, exist_ok=True)

    before_snapshot = _snapshot_hashes()
    production_02_hash_before = _sha256(production_02_file)
    production_05_hash_before = _sha256(production_05_file)

    # Guard hash check against Stage5R
    stage5r_h02 = _norm(stage5r_summary.get("production_02_hash_before"))
    stage5r_h05 = _norm(stage5r_summary.get("production_05_hash_before"))
    if stage5r_h02 and stage5r_h02 != production_02_hash_before:
        raise RuntimeError("Hash guard failed before apply: production_02 hash mismatch vs Stage5R.")
    if stage5r_h05 and stage5r_h05 != production_05_hash_before:
        raise RuntimeError("Hash guard failed before apply: production_05 hash mismatch vs Stage5R.")

    # Backup
    backup_02_file = OUT_BACKUP_DIR / "02_研报全量结构化数据.before_stage5s.xlsx"
    backup_05_file = OUT_BACKUP_DIR / "05_核心财务指标标准化.before_stage5s.xlsx"
    production_02_backup_hash = _copy_backup(production_02_file, backup_02_file)
    production_05_backup_hash = _copy_backup(production_05_file, backup_05_file)

    # Load manifest apply rows.
    apply_02_rows = pd.read_excel(INPUT_MANIFEST_XLSX, sheet_name="apply_02_rows").fillna("")
    apply_05_rows = pd.read_excel(INPUT_MANIFEST_XLSX, sheet_name="apply_05_rows").fillna("")

    manifest_promote_to_02_metric_count = int(apply_02_rows["metric_level_key"].map(_norm).nunique()) if not apply_02_rows.empty else 0
    manifest_promote_to_02_row_count = int(len(apply_02_rows))
    manifest_promote_to_05_metric_count = int(apply_05_rows["metric_level_key"].map(_norm).nunique()) if not apply_05_rows.empty else 0
    manifest_promote_to_05_row_count = int(len(apply_05_rows))

    # Validate excluded queue counts from stage5r summary
    excluded_need_mapping_metric_count = int(stage5r_summary.get("excluded_need_mapping_metric_count", -1))
    excluded_need_scope_review_metric_count = int(stage5r_summary.get("excluded_need_scope_review_metric_count", -1))
    blocked_metric_count = int(stage5r_summary.get("blocked_metric_count", -1))

    # Real apply: write dedicated applied sheets into production files.
    applied_to_02_rows_df = apply_02_rows.copy()
    applied_to_05_rows_df = apply_05_rows.copy()

    # Basic apply-time duplicate/conflict checks on candidate rows before writing.
    duplicate_key_count_02, value_conflict_count_02, unit_conflict_count_02, year_conflict_count_02 = _duplicate_and_conflict_counts_02(
        applied_to_02_rows_df
    )
    duplicate_key_count_05, value_conflict_count_05, unit_conflict_count_05, year_conflict_count_05 = _duplicate_and_conflict_counts_05(
        applied_to_05_rows_df
    )

    skipped_existing_02_row_count = 0
    skipped_existing_05_row_count = 0

    # Stop & rollback trigger pre-write if manifest rows are already invalid.
    if any(
        x != 0
        for x in [
            duplicate_key_count_02,
            duplicate_key_count_05,
            value_conflict_count_02,
            value_conflict_count_05,
            unit_conflict_count_02,
            unit_conflict_count_05,
            year_conflict_count_02,
            year_conflict_count_05,
        ]
    ):
        _rollback(production_02_file, production_05_file, backup_02_file, backup_05_file)
        raise RuntimeError("Pre-write conflict check failed. Rolled back from backups.")

    # Write now
    _write_dataframe_to_sheet(production_02_file, SHEET_APPLIED_02, applied_to_02_rows_df)
    _write_dataframe_to_sheet(production_05_file, SHEET_APPLIED_05, applied_to_05_rows_df)

    production_02_hash_after = _sha256(production_02_file)
    production_05_hash_after = _sha256(production_05_file)
    production_02_changed = bool(production_02_hash_after != production_02_hash_before)
    production_05_changed = bool(production_05_hash_after != production_05_hash_before)

    # Post-write global guard.
    after_snapshot = _snapshot_hashes()
    production_01_unchanged = bool(before_snapshot["01"] == after_snapshot["01"])
    production_02A_unchanged = bool(before_snapshot["02A"] == after_snapshot["02A"])
    production_06_unchanged = bool(before_snapshot["06"] == after_snapshot["06"])
    official_02B_unchanged = bool(before_snapshot["02B"] == after_snapshot["02B"])
    formal_rules_unchanged = bool(
        before_snapshot["formal_scope_rules"] == after_snapshot["formal_scope_rules"]
        and before_snapshot["formal_mapping_rules"] == after_snapshot["formal_mapping_rules"]
        and before_snapshot["formal_normalization_rules"] == after_snapshot["formal_normalization_rules"]
        and before_snapshot["formal_alias_rules"] == after_snapshot["formal_alias_rules"]
    )

    # If non-target changed or target didn't change unexpectedly => rollback.
    must_rollback = bool(
        (not production_02_changed)
        or (not production_05_changed)
        or (not production_01_unchanged)
        or (not production_02A_unchanged)
        or (not production_06_unchanged)
        or (not official_02B_unchanged)
        or (not formal_rules_unchanged)
    )
    if must_rollback:
        _rollback(production_02_file, production_05_file, backup_02_file, backup_05_file)
        # refresh hash after rollback
        production_02_hash_after = _sha256(production_02_file)
        production_05_hash_after = _sha256(production_05_file)
        production_02_changed = bool(production_02_hash_after != production_02_hash_before)
        production_05_changed = bool(production_05_hash_after != production_05_hash_before)
        raise RuntimeError("Post-write guard failed. Rollback completed.")

    check_delivery_state_after = _run_delivery_check()

    applied_to_02_metric_count = int(applied_to_02_rows_df["metric_level_key"].map(_norm).nunique()) if not applied_to_02_rows_df.empty else 0
    applied_to_02_row_count = int(len(applied_to_02_rows_df))
    applied_to_05_metric_count = int(applied_to_05_rows_df["metric_level_key"].map(_norm).nunique()) if not applied_to_05_rows_df.empty else 0
    applied_to_05_row_count = int(len(applied_to_05_rows_df))

    rollback_possible = bool(backup_02_file.exists() and backup_05_file.exists())

    summary = {
        "production_02_file": str(production_02_file),
        "production_05_file": str(production_05_file),
        "production_02_hash_before": production_02_hash_before,
        "production_05_hash_before": production_05_hash_before,
        "production_02_backup_file": str(backup_02_file),
        "production_05_backup_file": str(backup_05_file),
        "production_02_backup_hash": production_02_backup_hash,
        "production_05_backup_hash": production_05_backup_hash,
        "production_02_hash_after": production_02_hash_after,
        "production_05_hash_after": production_05_hash_after,
        "manifest_promote_to_02_metric_count": manifest_promote_to_02_metric_count,
        "manifest_promote_to_02_row_count": manifest_promote_to_02_row_count,
        "manifest_promote_to_05_metric_count": manifest_promote_to_05_metric_count,
        "manifest_promote_to_05_row_count": manifest_promote_to_05_row_count,
        "applied_to_02_metric_count": applied_to_02_metric_count,
        "applied_to_02_row_count": applied_to_02_row_count,
        "applied_to_05_metric_count": applied_to_05_metric_count,
        "applied_to_05_row_count": applied_to_05_row_count,
        "skipped_existing_02_row_count": skipped_existing_02_row_count,
        "skipped_existing_05_row_count": skipped_existing_05_row_count,
        "excluded_need_mapping_metric_count": excluded_need_mapping_metric_count,
        "excluded_need_scope_review_metric_count": excluded_need_scope_review_metric_count,
        "blocked_metric_count": blocked_metric_count,
        "duplicate_key_count_02": duplicate_key_count_02,
        "duplicate_key_count_05": duplicate_key_count_05,
        "value_conflict_count_02": value_conflict_count_02,
        "value_conflict_count_05": value_conflict_count_05,
        "unit_conflict_count_02": unit_conflict_count_02,
        "unit_conflict_count_05": unit_conflict_count_05,
        "year_conflict_count_02": year_conflict_count_02,
        "year_conflict_count_05": year_conflict_count_05,
        "production_02_changed": production_02_changed,
        "production_05_changed": production_05_changed,
        "production_01_unchanged": production_01_unchanged,
        "production_02A_unchanged": production_02A_unchanged,
        "production_06_unchanged": production_06_unchanged,
        "official_02B_unchanged": official_02B_unchanged,
        "formal_rules_unchanged": formal_rules_unchanged,
        "rollback_possible": rollback_possible,
        "check_delivery_state_after": check_delivery_state_after,
        "ai_called": False,
        "internet_called": False,
        "factory_core_called": False,
        "ocr_called": False,
        "stage5s_real_apply_pass": False,
    }

    summary["stage5s_real_apply_pass"] = bool(
        summary["applied_to_02_metric_count"] == 59
        and summary["applied_to_02_row_count"] == 295
        and summary["applied_to_05_metric_count"] == 9
        and summary["applied_to_05_row_count"] == 45
        and summary["excluded_need_mapping_metric_count"] == 4
        and summary["excluded_need_scope_review_metric_count"] == 4
        and summary["blocked_metric_count"] == 0
        and summary["duplicate_key_count_02"] == 0
        and summary["duplicate_key_count_05"] == 0
        and summary["value_conflict_count_02"] == 0
        and summary["value_conflict_count_05"] == 0
        and summary["unit_conflict_count_02"] == 0
        and summary["unit_conflict_count_05"] == 0
        and summary["year_conflict_count_02"] == 0
        and summary["year_conflict_count_05"] == 0
        and summary["production_02_changed"]
        and summary["production_05_changed"]
        and summary["production_01_unchanged"]
        and summary["production_02A_unchanged"]
        and summary["production_06_unchanged"]
        and summary["official_02B_unchanged"]
        and summary["formal_rules_unchanged"]
        and summary["rollback_possible"]
        and summary["check_delivery_state_after"] == "PASS"
        and (summary["ai_called"] is False)
        and (summary["internet_called"] is False)
        and (summary["factory_core_called"] is False)
        and (summary["ocr_called"] is False)
    )

    apply_log_df = pd.DataFrame(
        [
            {"metric": k, "value": v}
            for k, v in summary.items()
        ]
    )
    apply_diff_sheets = {
        "apply_02_rows": applied_to_02_rows_df,
        "apply_05_rows": applied_to_05_rows_df,
    }

    _write_excel(
        OUT_APPLY_LOG_XLSX,
        {
            "apply_summary": apply_log_df,
            "manifest_02": apply_02_rows,
            "manifest_05": apply_05_rows,
        },
    )
    _write_excel(
        OUT_APPLY_DIFF_XLSX,
        apply_diff_sheets,
    )

    md_lines = [
        "# Stage5S Real Apply Report",
        "",
        "## Apply Result",
        f"- applied_to_02_metric_count: {summary['applied_to_02_metric_count']}",
        f"- applied_to_02_row_count: {summary['applied_to_02_row_count']}",
        f"- applied_to_05_metric_count: {summary['applied_to_05_metric_count']}",
        f"- applied_to_05_row_count: {summary['applied_to_05_row_count']}",
        "",
        "## Excluded Queue",
        f"- excluded_need_mapping_metric_count: {summary['excluded_need_mapping_metric_count']}",
        f"- excluded_need_scope_review_metric_count: {summary['excluded_need_scope_review_metric_count']}",
        f"- blocked_metric_count: {summary['blocked_metric_count']}",
        "",
        "## Conflict Check",
        f"- duplicate_key_count_02: {summary['duplicate_key_count_02']}",
        f"- duplicate_key_count_05: {summary['duplicate_key_count_05']}",
        f"- value_conflict_count_02: {summary['value_conflict_count_02']}",
        f"- value_conflict_count_05: {summary['value_conflict_count_05']}",
        f"- unit_conflict_count_02: {summary['unit_conflict_count_02']}",
        f"- unit_conflict_count_05: {summary['unit_conflict_count_05']}",
        f"- year_conflict_count_02: {summary['year_conflict_count_02']}",
        f"- year_conflict_count_05: {summary['year_conflict_count_05']}",
        "",
        "## Guard",
        f"- production_01_unchanged: {summary['production_01_unchanged']}",
        f"- production_02A_unchanged: {summary['production_02A_unchanged']}",
        f"- production_06_unchanged: {summary['production_06_unchanged']}",
        f"- official_02B_unchanged: {summary['official_02B_unchanged']}",
        f"- formal_rules_unchanged: {summary['formal_rules_unchanged']}",
        f"- check_delivery_state_after: {summary['check_delivery_state_after']}",
        f"- stage5s_real_apply_pass: {summary['stage5s_real_apply_pass']}",
    ]
    OUT_APPLY_REPORT_MD.write_text("\n".join(md_lines), encoding="utf-8")
    OUT_SUMMARY_JSON.write_text(json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"stage5s_apply_log_xlsx: {OUT_APPLY_LOG_XLSX}")
    print(f"stage5s_apply_diff_xlsx: {OUT_APPLY_DIFF_XLSX}")
    print(f"stage5s_apply_report_md: {OUT_APPLY_REPORT_MD}")
    print(f"stage5s_summary_json: {OUT_SUMMARY_JSON}")
    print(f"stage5s_real_apply_pass: {summary['stage5s_real_apply_pass']}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
