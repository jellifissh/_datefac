import argparse
import hashlib
import json
import os
import shutil
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from openpyxl import load_workbook


DELIVERY_DIR = Path(r"D:\_datefac\output\delivery_package")
PRODUCTION_01 = next(iter(sorted(DELIVERY_DIR.glob("01_*.xlsx"))), None)
PRODUCTION_02 = next(iter(sorted([p for p in DELIVERY_DIR.glob("02_*.xlsx") if "backup" not in p.name.lower()])), None)
PRODUCTION_02A = next(iter(sorted(DELIVERY_DIR.glob("02A_*.xlsx"))), None)
PRODUCTION_06 = Path(r"D:\_datefac\output\delivery_package\06_最终核心财务指标.xlsx")
BACKUP_DIR = DELIVERY_DIR / "backup_before_real_apply"
BACKUP_06 = BACKUP_DIR / "06_最终核心财务指标.before_ai_extract_real_apply.xlsx"
APPROVAL_68 = DELIVERY_DIR / "68_ai_extract_real_apply_approval_review.xlsx"
READINESS_69 = DELIVERY_DIR / "69_ai_extract_real_apply_readiness_summary.json"


def _norm(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    return str(v).strip()


def _to_float(v: Any) -> Optional[float]:
    s = _norm(v).replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except Exception:
        return None


def _file_hash(path: Path) -> str:
    h = hashlib.sha256()
    with open(path, "rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest()


def _load_summary() -> Dict[str, Any]:
    return json.loads(READINESS_69.read_text(encoding="utf-8"))


def _load_approval_rows() -> pd.DataFrame:
    df = pd.read_excel(APPROVAL_68, sheet_name="approval_review")
    return df


def _current_production_snapshot() -> Dict[str, str]:
    snap = {}
    for key, path in [
        ("01", PRODUCTION_01),
        ("02", PRODUCTION_02),
        ("02A", PRODUCTION_02A),
        ("06", PRODUCTION_06),
    ]:
        snap[key] = _file_hash(path) if path and path.exists() else ""
    return snap


def _validate_readiness(summary: Dict[str, Any]) -> None:
    checks = [
        (bool(summary.get("ready_for_real_apply")), "ready_for_real_apply must be true"),
        (int(summary.get("reviewed_candidate_count", -1)) == 13, "reviewed_candidate_count must be 13"),
        (int(summary.get("blocked_count", -1)) == 0, "blocked_count must be 0"),
        (int(summary.get("need_manual_review_count", -1)) == 0, "need_manual_review_count must be 0"),
    ]
    for ok, msg in checks:
        if not ok:
            raise RuntimeError(f"BLOCKED: {msg}")


def _candidate_key(asset: str, metric: str, year: str) -> str:
    return "|".join([_norm(asset), _norm(metric), _norm(year)])


def _build_new_row(headers: List[str], candidate: Dict[str, Any]) -> Dict[str, Any]:
    new_row = {h: "" for h in headers}
    new_row.update(
        {
            "source_pdf": "",
            "asset_package": _norm(candidate.get("target_asset_package")),
            "report_type": "stage1_ai_extract_real_apply",
            "data_usability_tier": "ai_extract_approved",
            "standard_metric": _norm(candidate.get("metric")),
            "year": _norm(candidate.get("year")),
            "final_value": _norm(candidate.get("new_value")),
            "final_unit": _norm(candidate.get("new_unit")),
            "final_value_source": "ai_extract_real_apply",
            "final_review_status": "approved_auto_applied",
            "original_auto_value": "",
            "original_auto_unit": "",
            "corrected_value": "",
            "corrected_unit": "",
            "value_validation_status": "approved_from_sandbox_dry_run",
            "value_repair_applied": "",
            "source_row_label": _norm(candidate.get("metric")),
            "source_table_index": "",
            "source_row_index": "",
            "source_column": _norm(candidate.get("source_reference")),
            "evidence_crop_path": "",
            "trace_note": f"real_apply_from_{_norm(candidate.get('candidate_id'))}",
            "reviewer": "codex_real_apply",
            "reviewed_at": datetime.now().strftime("%Y-%m-%d"),
            "reviewer_note": f"approved via 68/69 package; source={_norm(candidate.get('source_reference'))}",
        }
    )
    return new_row


def main() -> int:
    parser = argparse.ArgumentParser(description="Real-apply approved AI extract candidates to production 06 with backup and hash guard.")
    parser.add_argument("--approval-xlsx", default=str(APPROVAL_68))
    parser.add_argument("--readiness-json", default=str(READINESS_69))
    parser.add_argument("--production-06", default=str(PRODUCTION_06))
    parser.add_argument("--delivery-dir", default=str(DELIVERY_DIR))
    args = parser.parse_args()

    production_06 = Path(args.production_06)
    delivery_dir = Path(args.delivery_dir)
    approval_xlsx = Path(args.approval_xlsx)
    readiness_json = Path(args.readiness_json)

    if not readiness_json.exists() or not approval_xlsx.exists() or not production_06.exists():
        print("BLOCKED_REQUIRED_INPUT_MISSING")
        return 3

    summary = json.loads(readiness_json.read_text(encoding="utf-8"))
    _validate_readiness(summary)

    if int(summary.get("reviewed_candidate_count", -1)) != 13:
        raise RuntimeError("BLOCKED: reviewed_candidate_count != 13")
    if int(summary.get("blocked_count", -1)) != 0:
        raise RuntimeError("BLOCKED: blocked_count != 0")
    if int(summary.get("need_manual_review_count", -1)) != 0:
        raise RuntimeError("BLOCKED: need_manual_review_count != 0")

    approval_df = _load_approval_rows()
    approved_df = approval_df[
        approval_df["review_decision"].astype(str).str.strip().str.lower().isin({"auto_approve", "approved"})
    ].copy()
    if len(approved_df) != 13:
        raise RuntimeError(f"BLOCKED: approved_candidate_count={len(approved_df)} != 13")

    production_before = _current_production_snapshot()
    hash_06_before = production_before["06"]

    BACKUP_DIR.mkdir(parents=True, exist_ok=True)
    shutil.copy2(production_06, BACKUP_06)

    workbook = load_workbook(production_06)
    sheet = workbook[workbook.sheetnames[0]]
    headers = [c.value for c in sheet[1]]
    if not headers:
        raise RuntimeError("BLOCKED: production 06 header row is empty")

    existing_rows: List[Dict[str, Any]] = []
    for r in sheet.iter_rows(min_row=2, values_only=True):
        existing_rows.append(dict(zip(headers, r)))

    existing_keys = {
        _candidate_key(row.get("asset_package"), row.get("standard_metric"), row.get("year"))
        for row in existing_rows
    }

    log_rows: List[Dict[str, Any]] = []
    diff_rows: List[Dict[str, Any]] = []
    real_applied_count = 0
    skipped_count = 0
    failed_count = 0

    for idx, row in approved_df.reset_index(drop=True).iterrows():
        candidate = row.to_dict()
        candidate_id = _norm(candidate.get("candidate_id"))
        metric_key = _norm(candidate.get("metric_key"))
        target_sheet = _norm(candidate.get("target_sheet")) or workbook.sheetnames[0]
        target_hint = _norm(candidate.get("target_cell / row / column"))
        source_reference = _norm(candidate.get("source_reference"))
        asset = _norm(candidate.get("target_asset_package"))
        metric = _norm(candidate.get("metric"))
        year = _norm(candidate.get("year"))
        new_value = _norm(candidate.get("new_value"))
        new_unit = _norm(candidate.get("new_unit"))

        old_value = ""
        old_unit = ""
        reason = ""
        apply_status = "applied"

        if not asset or not metric or not year:
            apply_status = "failed"
            reason = "missing_metric_key_components"
        elif new_value == "":
            apply_status = "failed"
            reason = "new_value_empty"
        elif _to_float(new_value) is None:
            apply_status = "failed"
            reason = "new_value_non_numeric"
        elif not source_reference:
            apply_status = "failed"
            reason = "source_reference_missing"

        key = _candidate_key(asset, metric, year)
        if apply_status == "applied" and key in existing_keys:
            apply_status = "failed"
            reason = "duplicate_metric_year_exists_in_production_06"

        if apply_status == "failed":
            failed_count += 1
            log_rows.append(
                {
                    "candidate_id": candidate_id,
                    "metric_key": metric_key,
                    "target_sheet": target_sheet,
                    "target_cell / row / column": target_hint,
                    "old_value": old_value,
                    "new_value": new_value,
                    "apply_status": apply_status,
                    "reason": reason,
                    "source_reference": source_reference,
                }
            )
            break

        new_row = _build_new_row(headers, candidate)
        sheet.append([new_row.get(h, "") for h in headers])
        existing_keys.add(key)
        existing_rows.append(new_row)
        real_applied_count += 1

        target_row = f"append_row_{sheet.max_row}"
        log_rows.append(
            {
                "candidate_id": candidate_id,
                "metric_key": metric_key,
                "target_sheet": target_sheet,
                "target_cell / row / column": target_hint or target_row,
                "old_value": old_value,
                "new_value": new_value,
                "apply_status": apply_status,
                "reason": "append_new_metric_year_row",
                "source_reference": source_reference,
            }
        )
        diff_rows.append(
            {
                "diff_action": "ADD",
                "candidate_id": candidate_id,
                "metric_key": metric_key,
                "asset_package": asset,
                "standard_metric": metric,
                "year": year,
                "old_value": old_value,
                "new_value": new_value,
                "old_unit": old_unit,
                "new_unit": new_unit,
                "target_sheet": target_sheet,
                "target_cell / row / column": target_hint or target_row,
                "source_reference": source_reference,
            }
        )

    # write only if everything succeeded
    if failed_count == 0:
        temp_path = production_06.with_name(f"{production_06.stem}.real_apply_tmp{production_06.suffix}")
        workbook.save(temp_path)
        os.replace(temp_path, production_06)
    else:
        skipped_count = len(approved_df) - real_applied_count - failed_count

    production_after = _current_production_snapshot()
    hash_06_after = production_after["06"]

    prod01_unchanged = production_before["01"] == production_after["01"]
    prod02_unchanged = production_before["02"] == production_after["02"]
    prod02a_unchanged = production_before["02A"] == production_after["02A"]
    prod06_changed = hash_06_before != hash_06_after

    summary70 = pd.DataFrame(
        [
            {"field": "approved_candidate_count", "value": int(len(approved_df))},
            {"field": "real_applied_count", "value": int(real_applied_count)},
            {"field": "skipped_count", "value": int(skipped_count)},
            {"field": "failed_count", "value": int(failed_count)},
            {"field": "production_06_hash_before", "value": hash_06_before},
            {"field": "production_06_hash_after", "value": hash_06_after},
            {"field": "production_06_changed", "value": str(prod06_changed).lower()},
            {"field": "production_01_unchanged", "value": str(prod01_unchanged).lower()},
            {"field": "production_02_unchanged", "value": str(prod02_unchanged).lower()},
            {"field": "production_02A_unchanged", "value": str(prod02a_unchanged).lower()},
            {"field": "ai_called", "value": "false"},
            {"field": "factory_core_called", "value": "false"},
            {"field": "ocr_called", "value": "false"},
            {"field": "backup_06_path", "value": str(BACKUP_06)},
        ]
    )

    log70 = pd.DataFrame(log_rows)
    diff71 = pd.DataFrame(diff_rows)

    out70xlsx = delivery_dir / "70_ai_extract_real_apply_log.xlsx"
    out70md = delivery_dir / "70_ai_extract_real_apply_log.md"
    out71 = delivery_dir / "71_ai_extract_real_apply_diff.xlsx"
    out72 = delivery_dir / "72_ai_extract_real_apply_summary.json"

    with pd.ExcelWriter(out70xlsx, engine="openpyxl") as writer:
        log70.to_excel(writer, sheet_name="apply_log", index=False)
        summary70.to_excel(writer, sheet_name="summary", index=False)

    with pd.ExcelWriter(out71, engine="openpyxl") as writer:
        diff71.to_excel(writer, sheet_name="real_apply_diff", index=False)

    out70md.write_text(
        "\n".join(
            [
                "# AI Extract Real Apply Log",
                "",
                f"- approved_candidate_count: {len(approved_df)}",
                f"- real_applied_count: {real_applied_count}",
                f"- skipped_count: {skipped_count}",
                f"- failed_count: {failed_count}",
                f"- production_06_hash_before: {hash_06_before}",
                f"- production_06_hash_after: {hash_06_after}",
                f"- production_06_changed: {str(prod06_changed).lower()}",
                f"- production_01_unchanged: {str(prod01_unchanged).lower()}",
                f"- production_02_unchanged: {str(prod02_unchanged).lower()}",
                f"- production_02A_unchanged: {str(prod02a_unchanged).lower()}",
                f"- backup_06_path: {BACKUP_06}",
                f"- failure_mode: {'none' if failed_count == 0 else 'stopped_on_first_failure'}",
            ]
        ),
        encoding="utf-8",
    )

    summary72 = {
        "approved_candidate_count": int(len(approved_df)),
        "real_applied_count": int(real_applied_count),
        "skipped_count": int(skipped_count),
        "failed_count": int(failed_count),
        "production_06_hash_before": hash_06_before,
        "production_06_hash_after": hash_06_after,
        "production_06_changed": bool(prod06_changed),
        "production_01_unchanged": bool(prod01_unchanged),
        "production_02_unchanged": bool(prod02_unchanged),
        "production_02A_unchanged": bool(prod02a_unchanged),
        "ai_called": False,
        "factory_core_called": False,
        "ocr_called": False,
        "backup_06_path": str(BACKUP_06),
        "failure_mode": "none" if failed_count == 0 else "stopped_on_first_failure",
    }
    out72.write_text(json.dumps(summary72, ensure_ascii=False, indent=2), encoding="utf-8")

    print(f"approved_candidate_count: {len(approved_df)}")
    print(f"real_applied_count: {real_applied_count}")
    print(f"skipped_count: {skipped_count}")
    print(f"failed_count: {failed_count}")
    print(f"production_06_hash_before: {hash_06_before}")
    print(f"production_06_hash_after: {hash_06_after}")
    print(f"production_06_changed: {prod06_changed}")
    print(f"production_01_unchanged: {prod01_unchanged}")
    print(f"production_02_unchanged: {prod02_unchanged}")
    print(f"production_02A_unchanged: {prod02a_unchanged}")
    print(f"backup_06_path: {BACKUP_06}")
    print(f"out70xlsx: {out70xlsx}")
    print(f"out70md: {out70md}")
    print(f"out71: {out71}")
    print(f"out72: {out72}")

    return 0 if failed_count == 0 else 4


if __name__ == "__main__":
    raise SystemExit(main())
