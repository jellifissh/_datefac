from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from datefac.benchmark.review_queue_source_check_evidence_review_ingestion_344b import (
    READY_DECISION_344B,
    VALIDATION_FAILED_DECISION_344B,
    build_review_queue_source_check_evidence_review_ingestion_344b,
)
from tools.run_review_queue_source_check_evidence_review_ingestion_344b import (
    _resolve_filled_workbook,
)


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SOURCE_CHECK_EVIDENCE_ENRICHMENT_344A2_DIR = (
    PROJECT_ROOT / "output" / "review_queue_source_check_evidence_enrichment_344a2"
)
SOURCE_CHECK_BACKLOG_PACKAGE_344A_DIR = (
    PROJECT_ROOT / "output" / "review_queue_source_check_backlog_package_344a"
)
DEMO_AUDIT_SNAPSHOT_343O_DIR = (
    PROJECT_ROOT / "output" / "review_queue_demo_audit_snapshot_343o"
)
REVIEW_QUEUE_SCHEMA_343A_DIR = PROJECT_ROOT / "output" / "review_queue_schema_343a"
FILLED_WORKBOOK = (
    PROJECT_ROOT
    / "input"
    / "review_queue_source_check_evidence_344a2_filled"
    / "review_queue_source_check_evidence_enrichment_344a2_enriched_review_template_filled_independent.xlsx"
)


def _make_output_dir() -> Path:
    tmp_root = Path(
        tempfile.mkdtemp(prefix="codex_344b_", dir=str(Path(tempfile.gettempdir())))
    )
    output_dir = tmp_root / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def _build_artifacts(filled_workbook: Path, output_dir: Path):
    return build_review_queue_source_check_evidence_review_ingestion_344b(
        source_check_evidence_enrichment_344a2_dir=SOURCE_CHECK_EVIDENCE_ENRICHMENT_344A2_DIR,
        source_check_backlog_package_344a_dir=SOURCE_CHECK_BACKLOG_PACKAGE_344A_DIR,
        demo_audit_snapshot_343o_dir=DEMO_AUDIT_SNAPSHOT_343O_DIR,
        review_queue_schema_343a_dir=REVIEW_QUEUE_SCHEMA_343A_DIR,
        filled_workbook=filled_workbook,
        output_dir=output_dir,
        repo_root=PROJECT_ROOT,
    )


def test_344b_ready_path() -> None:
    output_dir = _make_output_dir()
    try:
        artifacts = _build_artifacts(FILLED_WORKBOOK, output_dir)
        summary = artifacts["summary"]
        assert summary["decision"] == READY_DECISION_344B
        assert summary["review_queue_schema_version"] == "343A.review_queue.v1"
        assert summary["filled_row_count"] == 19
        assert summary["valid_row_count"] == 19
        assert summary["invalid_row_count"] == 0
        assert summary["source_confirm_count"] == 10
        assert summary["source_correct_count"] == 9
        assert summary["source_reject_count"] == 0
        assert summary["source_still_insufficient_count"] == 0
        assert summary["source_defer_count"] == 0
        assert summary["validated_sidecar_row_count"] == 19
        assert summary["correction_row_count"] == 9
        assert summary["source_check_result_ingested"] is True
        assert summary["source_check_backlog_resolved"] is True
        assert summary["validated_sidecar_generated"] is True
        assert summary["correction_sidecar_generated"] is True
        assert summary["audit_gate_generated"] is True
        assert summary["formal_client_export_allowed"] is False
        assert summary["client_ready"] is False
        assert summary["production_ready"] is False
        assert summary["global_strict_human_review_completed"] is False
        assert summary["ready_for_344c"] is True
        assert (
            summary["recommended_344c_scope"]
            == "source_check_confirmed_sidecar_apply_simulation_and_expanded_trust_gate"
        )
        assert summary["qa_fail_count"] == 0
        assert summary["no_write_back_proof_passed"] is True
        assert len(artifacts["validated_sidecar_rows"]) == 19
        assert len(artifacts["correction_rows"]) == 9
        assert artifacts["audit_gate"]["source_check_result_ingested"] is True
        assert artifacts["audit_gate"]["source_check_backlog_resolved"] is True
        assert any(
            row["corrected_metric_standardized"] == "YOY"
            and row["corrected_normalized_unit"] == "%"
            for row in artifacts["correction_rows"]
        )
    finally:
        shutil.rmtree(output_dir.parent, ignore_errors=True)


def test_344b_validation_failed_if_decision_invalid() -> None:
    output_dir = _make_output_dir()
    tmp_dir = Path(
        tempfile.mkdtemp(prefix="codex_344b_invalid_", dir=str(Path(tempfile.gettempdir())))
    )
    try:
        workbook_path = tmp_dir / "filled_invalid.xlsx"
        shutil.copy2(FILLED_WORKBOOK, workbook_path)
        workbook = load_workbook(workbook_path)
        sheet = workbook["04_REVIEW_TEMPLATE"]
        headers = {str(cell.value): cell.column for cell in sheet[1]}
        sheet.cell(row=2, column=headers["source_check_decision"]).value = "SOURCE_UNKNOWN"
        workbook.save(workbook_path)

        artifacts = _build_artifacts(workbook_path, output_dir)
        summary = artifacts["summary"]
        assert summary["decision"] == VALIDATION_FAILED_DECISION_344B
        assert summary["invalid_row_count"] == 1
        assert summary["source_check_result_ingested"] is False
        assert summary["ready_for_344c"] is False
        assert summary["qa_fail_count"] > 0
        assert any(
            "invalid source_check_decision" in issue["message"]
            for issue in artifacts["validation_errors"]
        )
    finally:
        shutil.rmtree(output_dir.parent, ignore_errors=True)
        shutil.rmtree(tmp_dir, ignore_errors=True)


def test_344b_validation_failed_if_correction_semantics_break() -> None:
    output_dir = _make_output_dir()
    tmp_dir = Path(
        tempfile.mkdtemp(
            prefix="codex_344b_bad_correction_", dir=str(Path(tempfile.gettempdir()))
        )
    )
    try:
        workbook_path = tmp_dir / "filled_bad_correction.xlsx"
        shutil.copy2(FILLED_WORKBOOK, workbook_path)
        workbook = load_workbook(workbook_path)
        sheet = workbook["04_REVIEW_TEMPLATE"]
        headers = {str(cell.value): cell.column for cell in sheet[1]}
        for row_index in range(2, sheet.max_row + 1):
            if sheet.cell(row=row_index, column=headers["source_check_decision"]).value == "SOURCE_CORRECT":
                sheet.cell(
                    row=row_index,
                    column=headers["source_check_metric_standardized"],
                ).value = "revenue"
                break
        workbook.save(workbook_path)

        artifacts = _build_artifacts(workbook_path, output_dir)
        summary = artifacts["summary"]
        assert summary["decision"] == VALIDATION_FAILED_DECISION_344B
        assert summary["invalid_row_count"] == 1
        assert summary["source_check_result_ingested"] is False
        assert summary["correction_row_count"] == 8
        assert any(
            "SOURCE_CORRECT revenue rows must correct metric to YOY" in issue["message"]
            for issue in artifacts["validation_errors"]
        )
    finally:
        shutil.rmtree(output_dir.parent, ignore_errors=True)
        shutil.rmtree(tmp_dir, ignore_errors=True)


def test_resolve_filled_workbook_uses_explicit_path() -> None:
    resolved = _resolve_filled_workbook(
        str(FILLED_WORKBOOK),
        SOURCE_CHECK_EVIDENCE_ENRICHMENT_344A2_DIR,
    )
    assert resolved == FILLED_WORKBOOK

