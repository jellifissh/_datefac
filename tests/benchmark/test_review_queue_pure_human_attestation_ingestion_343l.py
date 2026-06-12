from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from datefac.benchmark.review_queue_pure_human_attestation_ingestion_343l import (
    NOT_READY_DECISION_343L,
    READY_DECISION_343L,
    build_review_queue_pure_human_attestation_ingestion_343l,
)


PROJECT_ROOT = Path(__file__).resolve().parents[2]
PACKAGE_343K_DIR = PROJECT_ROOT / "output" / "review_queue_pure_human_attestation_package_343k"
STRICT_REVIEW_343J_DIR = PROJECT_ROOT / "output" / "review_queue_strict_review_ingestion_343j"
SOURCE_EVIDENCE_343I2_DIR = PROJECT_ROOT / "output" / "review_queue_source_evidence_enrichment_343i2"
SCHEMA_343A_DIR = PROJECT_ROOT / "output" / "review_queue_schema_343a"
FILLED_WORKBOOK = (
    PROJECT_ROOT
    / "input"
    / "review_queue_pure_human_attestation_343k_filled"
    / "review_queue_pure_human_attestation_package_343k_attestation_template_filled.xlsx"
)


def _make_output_dir() -> Path:
    tmp_root = Path(
        tempfile.mkdtemp(prefix="codex_343l_", dir=str(Path(tempfile.gettempdir())))
    )
    output_dir = tmp_root / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def _build_artifacts(*, filled_workbook: Path, output_dir: Path):
    return build_review_queue_pure_human_attestation_ingestion_343l(
        pure_human_attestation_package_343k_dir=PACKAGE_343K_DIR,
        strict_review_ingestion_343j_dir=STRICT_REVIEW_343J_DIR,
        source_evidence_enrichment_343i2_dir=SOURCE_EVIDENCE_343I2_DIR,
        review_queue_schema_343a_dir=SCHEMA_343A_DIR,
        filled_workbook=filled_workbook,
        output_dir=output_dir,
        repo_root=PROJECT_ROOT,
    )


def test_343l_ready_path() -> None:
    output_dir = _make_output_dir()
    try:
        artifacts = _build_artifacts(filled_workbook=FILLED_WORKBOOK, output_dir=output_dir)
        summary = artifacts["summary"]
        assert summary["decision"] == READY_DECISION_343L
        assert summary["review_queue_schema_version"] == "343A.review_queue.v1"
        assert summary["filled_row_count"] == 10
        assert summary["valid_row_count"] == 10
        assert summary["invalid_row_count"] == 0
        assert summary["human_accept_count"] == 10
        assert summary["human_correct_count"] == 0
        assert summary["human_reject_count"] == 0
        assert summary["human_needs_source_check_count"] == 0
        assert summary["human_defer_count"] == 0
        assert summary["human_source_evidence_checked_true_count"] == 10
        assert summary["human_independent_check_attested_true_count"] == 10
        assert summary["pure_human_attestation_result_ingested"] is True
        assert summary["pure_strict_human_confirm_count"] == 10
        assert summary["pure_strict_human_correct_count"] == 0
        assert summary["pure_strict_human_review_completed_for_package"] is True
        assert summary["strict_human_review_completed_scope"] == "343K_PACKAGE_ONLY"
        assert summary["global_strict_human_review_completed"] is False
        assert summary["formal_client_export_allowed"] is False
        assert summary["client_ready"] is False
        assert summary["production_ready"] is False
        assert summary["ready_for_343m"] is True
        assert (
            summary["recommended_343m_scope"]
            == "human_confirmed_sidecar_apply_simulation_and_limited_export_gate"
        )
        assert summary["qa_fail_count"] == 0
        assert summary["no_write_back_proof_passed"] is True
        assert artifacts["client_export_gate"]["formal_client_export_allowed"] is False
        assert artifacts["client_export_gate"]["package_strict_human_review_completed"] is True
        assert "343K_PACKAGE_ONLY" in artifacts["scope_boundary_markdown"]
    finally:
        shutil.rmtree(output_dir.parent, ignore_errors=True)


def test_343l_not_ready_if_accept_missing_independent_attestation() -> None:
    output_dir = _make_output_dir()
    tmp_dir = Path(
        tempfile.mkdtemp(prefix="codex_343l_invalid_", dir=str(Path(tempfile.gettempdir())))
    )
    try:
        workbook_path = tmp_dir / "filled_invalid.xlsx"
        shutil.copy2(FILLED_WORKBOOK, workbook_path)
        workbook = load_workbook(workbook_path)
        sheet = workbook["04_ATTESTATION_TEMPLATE"]
        headers = {str(cell.value): cell.column for cell in sheet[1]}
        sheet.cell(row=2, column=headers["human_independent_check_attested"]).value = False
        workbook.save(workbook_path)

        artifacts = _build_artifacts(filled_workbook=workbook_path, output_dir=output_dir)
        summary = artifacts["summary"]
        assert summary["decision"] == NOT_READY_DECISION_343L
        assert summary["invalid_row_count"] == 1
        assert summary["pure_human_attestation_result_ingested"] is False
        assert summary["pure_strict_human_review_completed_for_package"] is False
        assert summary["ready_for_343m"] is False
        assert summary["qa_fail_count"] > 0
        assert any(
            "human_independent_check_attested=true" in issue["message"]
            for issue in artifacts["validation_errors"]
            if issue["issue_type"] == "ERROR"
        )
    finally:
        shutil.rmtree(output_dir.parent, ignore_errors=True)
        shutil.rmtree(tmp_dir, ignore_errors=True)
