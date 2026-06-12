from __future__ import annotations

import shutil
import tempfile
from pathlib import Path

from openpyxl import load_workbook

from datefac.benchmark.review_queue_strict_review_ingestion_343j import (
    NOT_READY_DECISION_343J,
    READY_DECISION_343J,
    build_review_queue_strict_review_ingestion_343j,
)
from tools.run_review_queue_strict_review_ingestion_343j import _resolve_filled_workbook


PROJECT_ROOT = Path(__file__).resolve().parents[2]
SOURCE_EVIDENCE_ENRICHMENT_343I2_DIR = PROJECT_ROOT / "output" / "review_queue_source_evidence_enrichment_343i2"
STRICT_HUMAN_REVIEW_PACKAGE_343I_DIR = PROJECT_ROOT / "output" / "review_queue_strict_human_review_package_343i"
AUDIT_SUMMARY_343H_DIR = PROJECT_ROOT / "output" / "review_queue_audit_summary_343h"
REVIEW_QUEUE_SCHEMA_343A_DIR = PROJECT_ROOT / "output" / "review_queue_schema_343a"
FILLED_WORKBOOK = PROJECT_ROOT / "input" / "review_queue_strict_human_review_343i2_filled" / "review_queue_source_evidence_enrichment_343i2_enriched_review_template_filled.xlsx"


def _make_output_dir() -> Path:
    tmp_root = Path(tempfile.mkdtemp(prefix="codex_343j_", dir=str(Path(tempfile.gettempdir()))))
    output_dir = tmp_root / "output"
    output_dir.mkdir(parents=True, exist_ok=True)
    return output_dir


def _build_artifacts(filled_workbook: Path, output_dir: Path):
    return build_review_queue_strict_review_ingestion_343j(
        source_evidence_enrichment_343i2_dir=SOURCE_EVIDENCE_ENRICHMENT_343I2_DIR,
        strict_human_review_package_343i_dir=STRICT_HUMAN_REVIEW_PACKAGE_343I_DIR,
        audit_summary_343h_dir=AUDIT_SUMMARY_343H_DIR,
        review_queue_schema_343a_dir=REVIEW_QUEUE_SCHEMA_343A_DIR,
        filled_workbook=filled_workbook,
        output_dir=output_dir,
        repo_root=PROJECT_ROOT,
    )


def test_343j_ready_path() -> None:
    output_dir = _make_output_dir()
    try:
        artifacts = _build_artifacts(FILLED_WORKBOOK, output_dir)
        summary = artifacts["summary"]
        assert summary["decision"] == READY_DECISION_343J
        assert summary["review_queue_schema_version"] == "343A.review_queue.v1"
        assert summary["filled_row_count"] == 10
        assert summary["valid_row_count"] == 10
        assert summary["invalid_row_count"] == 0
        assert summary["strict_confirm_count"] == 10
        assert summary["strict_correct_count"] == 0
        assert summary["strict_reject_count"] == 0
        assert summary["strict_needs_source_check_count"] == 0
        assert summary["strict_defer_count"] == 0
        assert summary["strict_review_input_source_type"] == "AI_ASSISTED_EVIDENCE_CHECK"
        assert summary["not_pure_human_review"] is True
        assert summary["pure_strict_human_confirm_count"] == 0
        assert summary["ai_assisted_strict_review_confirm_count"] == 10
        assert summary["strict_review_result_ingested"] is True
        assert summary["pure_strict_human_review_completed"] is False
        assert summary["strict_human_review_completed"] is False
        assert summary["requires_strict_human_review"] is True
        assert summary["requires_pure_human_confirmation"] is True
        assert summary["formal_client_export_allowed"] is False
        assert summary["client_ready"] is False
        assert summary["production_ready"] is False
        assert summary["ready_for_343k"] is True
        assert (
            summary["recommended_343k_scope"]
            == "pure_human_confirmation_attestation_package_for_ai_assisted_strict_confirmed_rows"
        )
        assert summary["qa_fail_count"] == 0
        assert summary["no_write_back_proof_passed"] is True
        assert artifacts["client_export_gate"]["formal_client_export_allowed"] is False
        assert artifacts["result_rows"][0]["strict_review_input_source_type"] == "AI_ASSISTED_EVIDENCE_CHECK"
        assert "AI-assisted evidence check" in artifacts["result_rows"][0]["review_source_disclosure"]
        assert artifacts["validation_errors"] == []
        assert "03_REVIEW_RESULTS" in artifacts["workbook_sheets"]
        assert "07_SOURCE_DISCLOSURE" in artifacts["workbook_sheets"]
    finally:
        shutil.rmtree(output_dir.parent, ignore_errors=True)


def test_343j_not_ready_if_decision_invalid() -> None:
    output_dir = _make_output_dir()
    tmp_dir = Path(tempfile.mkdtemp(prefix="codex_343j_invalid_", dir=str(Path(tempfile.gettempdir()))))
    try:
        workbook_path = tmp_dir / "filled_invalid.xlsx"
        shutil.copy2(FILLED_WORKBOOK, workbook_path)
        workbook = load_workbook(workbook_path)
        sheet = workbook["04_REVIEW_TEMPLATE"]
        headers = {str(cell.value): cell.column for cell in sheet[1]}
        sheet.cell(row=2, column=headers["strict_review_decision"]).value = "STRICT_UNKNOWN"
        workbook.save(workbook_path)

        artifacts = _build_artifacts(workbook_path, output_dir)
        summary = artifacts["summary"]
        assert summary["decision"] == NOT_READY_DECISION_343J
        assert summary["invalid_row_count"] == 1
        assert summary["ready_for_343k"] is False
        assert summary["strict_review_result_ingested"] is False
        assert summary["qa_fail_count"] > 0
        assert artifacts["validation_errors"]
        assert any("invalid strict_review_decision" in issue["message"] for issue in artifacts["validation_errors"])
    finally:
        shutil.rmtree(output_dir.parent, ignore_errors=True)
        shutil.rmtree(tmp_dir, ignore_errors=True)


def test_resolve_filled_workbook_uses_explicit_path() -> None:
    resolved = _resolve_filled_workbook(str(FILLED_WORKBOOK), SOURCE_EVIDENCE_ENRICHMENT_343I2_DIR)
    assert resolved == FILLED_WORKBOOK
